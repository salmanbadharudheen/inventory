from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, DetailView, UpdateView, FormView, View, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.db.models import Q
from django.contrib import messages
import csv
import io
from .models import (Asset, AssetAttachment, Category, SubCategory, Vendor, generate_asset_tag,
                     Group, SubGroup, Brand, Company, Supplier, Custodian, AssetRemarks, AssetTransfer, AssetDisposal)
from .forms import (AssetForm, CategoryForm, SubCategoryForm, VendorForm, AssetImportForm,
                    GroupForm, SubGroupForm, BrandForm, CompanyForm, SupplierForm, CustodianForm, AssetRemarksForm, AssetTransferForm, AssetTransferReceiveForm, AssetDisposalForm, AssetDisposalManagerApprovalForm, AssetDisposalApprovalForm)
from django.db import transaction, models, IntegrityError
from apps.locations.models import (Branch, Building, Floor, Room, 
                                   Region, Site, Location, SubLocation, Department)
from django.urls import reverse
from django.views.decorators.http import require_POST
import json
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.template.loader import render_to_string
from django.utils import timezone
from decimal import Decimal
from datetime import date, datetime
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.core.cache import cache
from uuid import uuid4
import openpyxl

try:
    from weasyprint import HTML as WeasyprintHTML
except (ImportError, OSError):
    WeasyprintHTML = None


def invalidate_dashboard_cache_for_org(org):
    if org:
        cache.delete(f'dashboard_data_{org.id}')

def get_subcategories(request):
    category_id = request.GET.get('category_id')
    if category_id:
        subcategories = SubCategory.objects.filter(category_id=category_id).values('id', 'name')
        return JsonResponse(list(subcategories), safe=False)
    return JsonResponse([], safe=False)

def get_departments(request):
    branch_id = request.GET.get('branch_id')
    if branch_id:
        from apps.locations.models import Department
        departments = Department.objects.filter(branch_id=branch_id).values('id', 'name')
        return JsonResponse(list(departments), safe=False)
    return JsonResponse([], safe=False)

def get_buildings(request):
    branch_id = request.GET.get('branch_id')
    if branch_id:
        buildings = Building.objects.filter(branch_id=branch_id).values('id', 'name')
        return JsonResponse(list(buildings), safe=False)
    return JsonResponse([], safe=False)


def get_buildings_by_site(request):
    """Return buildings associated with a given site (via Location->building link)."""
    site_id = request.GET.get('site_id')
    if site_id:
        buildings = Building.objects.filter(locations__site_id=site_id).distinct().values('id', 'name')
        return JsonResponse(list(buildings), safe=False)
    return JsonResponse([], safe=False)

def get_floors(request):
    building_id = request.GET.get('building_id')
    if building_id:
        floors = Floor.objects.filter(building_id=building_id).values('id', 'name')
        return JsonResponse(list(floors), safe=False)
    return JsonResponse([], safe=False)

def get_rooms(request):
    floor_id = request.GET.get('floor_id')
    if floor_id:
        rooms = Room.objects.filter(floor_id=floor_id).values('id', 'name')
        return JsonResponse(list(rooms), safe=False)
    return JsonResponse(list(rooms), safe=False)

def get_locations(request):
    """Return locations filtered by building_id (optional)."""
    building_id = request.GET.get('building_id')
    if building_id:
        locations = Location.objects.filter(building_id=building_id).values('id', 'name')
        return JsonResponse(list(locations), safe=False)
    return JsonResponse([], safe=False)

def lookup_asset(request):
    # Support lookup by free-text `q` (tag/name/code), `asset_tag`, or by explicit `asset_id`.
    asset_id = request.GET.get('asset_id')
    query = request.GET.get('q')
    asset_tag = request.GET.get('asset_tag')

    asset = None
    org = getattr(request.user, 'organization', None)

    if asset_id:
        try:
            asset = Asset.objects.select_related(
                'department', 'branch', 'building', 'floor', 'room',
                'region', 'site', 'location', 'sub_location', 'assigned_to',
                'company', 'custodian'
            ).get(id=asset_id, organization=org)
        except Asset.DoesNotExist:
            asset = None

    elif asset_tag:
        # Search by asset tag, asset code, ERP number (exact) then partial
        asset = Asset.objects.filter(
            Q(asset_tag__iexact=asset_tag) |
            Q(asset_code__iexact=asset_tag) |
            Q(erp_asset_number__iexact=asset_tag),
            organization=org
        ).select_related(
            'department', 'branch', 'building', 'floor', 'room',
            'region', 'site', 'location', 'sub_location', 'assigned_to',
            'company', 'custodian'
        ).first()

        if not asset:
            # Partial match fallback
            asset = Asset.objects.filter(
                Q(asset_tag__icontains=asset_tag) |
                Q(asset_code__icontains=asset_tag) |
                Q(name__icontains=asset_tag),
                organization=org
            ).select_related(
                'department', 'branch', 'building', 'floor', 'room',
                'region', 'site', 'location', 'sub_location', 'assigned_to',
                'company', 'custodian'
            ).first()

    elif query:
        # Prioritize exact match on tags, then partial on name
        asset = Asset.objects.filter(
            Q(asset_tag__iexact=query) |
            Q(asset_code__iexact=query) |
            Q(erp_asset_number__iexact=query),
            organization=org
        ).select_related(
            'department', 'branch', 'building', 'floor', 'room',
            'region', 'site', 'location', 'sub_location', 'assigned_to',
            'company', 'custodian'
        ).first()

        if not asset:
            # Fallback to name search if no exact tag match
            asset = Asset.objects.filter(
                name__icontains=query,
                organization=org
            ).select_related(
                'department', 'branch', 'building', 'floor', 'room',
                'region', 'site', 'location', 'sub_location', 'assigned_to',
                'company', 'custodian'
            ).first()

    if not asset:
        return JsonResponse({'error': 'Not found'}, status=404)

    # Current location/ownership details
    assigned = None
    if asset.assigned_to:
        assigned = {'id': asset.assigned_to.id, 'name': asset.assigned_to.get_full_name() or asset.assigned_to.username}

    company = None
    if hasattr(asset, 'company') and asset.company:
        company = {'id': asset.company.id, 'name': asset.company.name}

    custodian = None
    if hasattr(asset, 'custodian') and asset.custodian:
        custodian = {'id': asset.custodian.id, 'name': str(asset.custodian)}

    current = {
        'id': str(asset.id),
        'name': asset.name,
        'asset_tag': asset.asset_tag,
        'department': {'id': asset.department.id, 'name': asset.department.name} if asset.department else None,
        'branch': {'id': asset.branch.id, 'name': asset.branch.name} if asset.branch else None,
        'building': {'id': asset.building.id, 'name': asset.building.name} if asset.building else None,
        'floor': {'id': asset.floor.id, 'name': asset.floor.name} if asset.floor else None,
        'room': {'id': asset.room.id, 'name': asset.room.name} if asset.room else None,
        'region': {'id': asset.region.id, 'name': asset.region.name} if getattr(asset, 'region', None) else None,
        'site': {'id': asset.site.id, 'name': asset.site.name} if getattr(asset, 'site', None) else None,
        'location': {'id': asset.location.id, 'name': asset.location.name} if getattr(asset, 'location', None) else None,
        'sub_location': {'id': asset.sub_location.id, 'name': asset.sub_location.name} if getattr(asset, 'sub_location', None) else None,
        'assigned_to': assigned,
        'company': company,
        'custodian': custodian,
        'status': asset.status,
    }

    # Available transfer options (simple lists to populate selects)
    departments = list(Department.objects.filter(organization=org).values('id', 'name')) if org else []
    locations = list(Location.objects.filter(organization=org).values('id', 'name')) if org else []
    users = []
    try:
        from django.contrib.auth import get_user_model
        User = get_user_model()
        users = list(User.objects.filter(organization=org).values('id', 'first_name', 'last_name', 'username')) if org else []
        # normalize to id/name
        users = [{'id': u['id'], 'name': (u['first_name'] + ' ' + u['last_name']).strip() or u['username']} for u in users]
    except Exception:
        users = []

    return JsonResponse({'asset': current, 'departments': departments, 'locations': locations, 'users': users})

def ajax_search_assets(request):
    """Return a JSON list of assets matching a search query. Used by the transfer browse panel."""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)

    org = getattr(request.user, 'organization', None)
    q = request.GET.get('q', '').strip()
    limit = min(int(request.GET.get('limit', 50)), 100)

    _active_disposal_statuses = [
        AssetDisposal.Status.PENDING,
        AssetDisposal.Status.MANAGER_APPROVED,
        AssetDisposal.Status.APPROVED,
        AssetDisposal.Status.COMPLETED,
    ]
    qs = Asset.objects.filter(organization=org, is_deleted=False).exclude(
        disposals__status__in=_active_disposal_statuses
    ).select_related('category', 'department')

    if q:
        qs = qs.filter(
            Q(asset_tag__icontains=q) |
            Q(name__icontains=q) |
            Q(custom_asset_tag__icontains=q) |
            Q(serial_number__icontains=q)
        )

    assets = []
    for a in qs.order_by('asset_tag')[:limit]:
        assets.append({
            'id': str(a.id),
            'asset_tag': a.asset_tag or '',
            'name': a.name or '',
            'category': a.category.name if a.category else '',
            'department': a.department.name if a.department else '',
        })

    return JsonResponse({'assets': assets, 'total': qs.count()})

ASSET_IMPORT_FIELDS = [
    'name', 'description', 'short_description', 'asset_tag',
    'asset_code', 'erp_asset_number', 'quantity', 'label_type', 'serial_number', 
    'category', 'sub_category', 'asset_type', 'group', 'sub_group', 'brand', 
    'model', 'condition', 'status', 'department', 'cost_center', 'company', 
    'supplier', 'vendor', 'custodian', 'employee_number', 'branch', 'building', 
    'floor', 'room', 'region', 'site', 'location', 'sub_location', 'purchase_date', 
    'purchase_price', 'currency', 'invoice_number', 'invoice_date', 'po_number', 
    'po_date', 'do_number', 'do_date', 'grn_number', 'warranty_start', 'warranty_end', 
    'tagged_date', 'date_placed_in_service', 'insurance_start_date', 'insurance_end_date', 
    'maintenance_start_date', 'maintenance_end_date', 'next_maintenance_date', 
    'maintenance_frequency_days', 'expected_units', 'useful_life_years', 
    'salvage_value', 'depreciation_method', 'remarks', 'notes'
]

def download_sample_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_assets.csv"'

    writer = csv.writer(response)
    writer.writerow(ASSET_IMPORT_FIELDS)
    
    # Sample Row
    writer.writerow([
        'Laptop Dell XPS', 'High-end laptop', 'Dell XPS 15', '', 'TAG-001',
        'C001', 'ERP-100', '1', 'BARCODE', 'SN123456', 
        'IT', 'Laptops', 'TAGGABLE', 'IT Equipment', 'Computers', 'Dell', 
        'XPS 15', 'NEW', 'ACTIVE', 'IT Dept', 'CC-101', 'ABC Corp', 
        'Tech Supplies Ltd', 'Main Vendor', 'EMP001', 'E123', 'Main Branch', 'HQ Building', 
        '2nd Floor', 'Room 201', 'North Region', 'Main Site', 'Main Location', 'Sub 1', '2023-01-01', 
        '5000', 'AED', 'INV-001', '2023-01-01', 'PO-100', 
        '2022-12-15', 'DO-100', '2022-12-28', 'GRN-100', '2023-01-01', '2026-01-01', 
        '2023-01-02', '2023-01-10', '2023-01-01', '2024-01-01', 
        '2023-01-01', '2024-01-01', '2023-06-01', '180', '1000', '5', '500', 
        'STRAIGHT_LINE', 'Needs Setup', 'Initial deployment'
    ])
    
    return response

def download_sample_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assets"

    # Style definitions
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Calibri', size=11)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    alt_row_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

    # Header row
    ws.append(ASSET_IMPORT_FIELDS)

    # Sample row
    sample_data = [
        'Laptop Dell XPS', 'High-end laptop', 'Dell XPS 15', '', 'TAG-001',
        'C001', 'ERP-100', 1, 'BARCODE', 'SN123456',
        'IT', 'Laptops', 'TAGGABLE', 'IT Equipment', 'Computers', 'Dell',
        'XPS 15', 'NEW', 'ACTIVE', 'IT Dept', 'CC-101', 'ABC Corp',
        'Tech Supplies Ltd', 'Main Vendor', 'EMP001', 'E123', 'Main Branch', 'HQ Building',
        '2nd Floor', 'Room 201', 'North Region', 'Main Site', 'Main Location', 'Sub 1', '2023-01-01',
        5000, 'AED', 'INV-001', '2023-01-01', 'PO-100',
        '2022-12-15', 'DO-100', '2022-12-28', 'GRN-100', '2023-01-01', '2026-01-01',
        '2023-01-02', '2023-01-10', '2023-01-01', '2024-01-01',
        '2023-01-01', '2024-01-01', '2023-06-01', '180', '1000', '5', '500',
        'STRAIGHT_LINE', 'Needs Setup', 'Initial deployment'
    ]
    ws.append(sample_data)

    # Apply header styles
    for col_idx, header in enumerate(ASSET_IMPORT_FIELDS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Apply data row styles
    for col_idx in range(1, len(ASSET_IMPORT_FIELDS) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        cell.fill = alt_row_fill

    # Auto-fit column widths based on content
    for col_idx, header in enumerate(ASSET_IMPORT_FIELDS, 1):
        col_letter = get_column_letter(col_idx)
        header_len = len(str(header))
        data_val = sample_data[col_idx - 1] if col_idx - 1 < len(sample_data) else ''
        data_len = len(str(data_val))
        optimal_width = max(header_len, data_len) + 4
        optimal_width = min(optimal_width, 40)
        optimal_width = max(optimal_width, 12)
        ws.column_dimensions[col_letter].width = optimal_width

    # Set row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22

    # Freeze header row
    ws.freeze_panes = 'A2'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sample_assets.xlsx"'
    wb.save(response)
    return response

# --- ASSET VIEWS ---
class AssetListView(LoginRequiredMixin, ListView):
    model = Asset
    template_name = 'assets/asset_list.html'
    context_object_name = 'assets'
    paginate_by = 25

    def get_template_names(self):
        if self.request.GET.get('view') == 'depreciation':
            return ['assets/depreciation_report.html']
        return ['assets/asset_list.html']

    def get_queryset(self):
        org = getattr(self.request.user, 'organization', None)
        # Tenant isolation guard: never show organization-less assets.
        if not org:
            return Asset.objects.none()

        queryset = Asset.objects.filter(
            organization=org,
            is_deleted=False
        ).select_related(
            'category', 'sub_category', 'branch', 'assigned_to', 
            'site', 'building', 'brand_new', 'room', 'department',
            'sub_location', 'group'
        ).prefetch_related('attachments')

        # Search across many asset fields and common related names
        query = self.request.GET.get('q')
        if query:
            q = (
                Q(name__icontains=query) |
                Q(asset_tag__icontains=query) |
                Q(asset_code__icontains=query) |
                Q(erp_asset_number__icontains=query) |
                Q(serial_number__icontains=query) |
                Q(model__icontains=query) |
                Q(brand__icontains=query) |
                Q(brand_new__name__icontains=query) |
                Q(category__name__icontains=query) |
                Q(sub_category__name__icontains=query) |
                Q(vendor__name__icontains=query) |
                Q(supplier__name__icontains=query) |
                Q(company__name__icontains=query) |
                Q(invoice_number__icontains=query) |
                Q(po_number__icontains=query) |
                Q(grn_number__icontains=query) |
                Q(do_number__icontains=query) |
                Q(tagged_date__icontains=query) |
                Q(purchase_date__icontains=query) |
                Q(department__name__icontains=query) |
                Q(branch__name__icontains=query) |
                Q(building__name__icontains=query) |
                Q(room__name__icontains=query) |
                Q(site__name__icontains=query) |
                Q(region__name__icontains=query) |
                Q(location__name__icontains=query) |
                Q(sub_location__name__icontains=query) |
                Q(assigned_to__username__icontains=query) |
                Q(assigned_to__first_name__icontains=query) |
                Q(assigned_to__last_name__icontains=query) |
                Q(notes__icontains=query)
            )

            queryset = queryset.filter(q)

        product_name = self.request.GET.get('product_name')
        if product_name:
            queryset = queryset.filter(name__icontains=product_name)
            
        # Advanced Filters
        filters = {
            'status': 'status',
            'condition': 'condition',
            'category': 'category_id',
            'subcategory': 'sub_category_id',
            'group': 'group_id',
            'sub_group': 'sub_group_id',
            'site': 'site_id',
            'building': 'building_id',
            'floor': 'floor_id',
            'room': 'room_id',
            'location': 'location_id',
            'sub_location': 'sub_location_id',
            'brand': 'brand_new_id',
            'department': 'department_id',
            'supplier': 'supplier_id',
            'created_by': 'created_by_id',
            'label_type': 'label_type',
        }
        
        for param, field in filters.items():
            val = self.request.GET.get(param)
            if val:
                queryset = queryset.filter(**{field: val})

        is_tagged = (self.request.GET.get('is_tagged') or '').strip().lower()
        if is_tagged == 'tagged':
            queryset = queryset.filter(tagging_status='TAGGED')
        elif is_tagged == 'untagged':
            queryset = queryset.filter(
                Q(tagging_status='UNTAGGED') |
                Q(tagging_status__isnull=True) |
                Q(tagging_status='')
            )

        tag_type = self.request.GET.get('tag_type')
        if tag_type == 'BARCODE':
            queryset = queryset.filter(barcode_image__isnull=False)
        elif tag_type == 'RFID':
            queryset = queryset.filter(label_type__iexact='RFID')
        
        # Date Range Filters
        purchase_date_from = self.request.GET.get('purchase_date_from')
        purchase_date_to = self.request.GET.get('purchase_date_to')
        registered_date_from = self.request.GET.get('registered_date_from')
        registered_date_to = self.request.GET.get('registered_date_to')
        
        if purchase_date_from:
            queryset = queryset.filter(purchase_date__gte=purchase_date_from)
        if purchase_date_to:
            queryset = queryset.filter(purchase_date__lte=purchase_date_to)
        if registered_date_from:
            queryset = queryset.filter(created_at__date__gte=registered_date_from)
        if registered_date_to:
            queryset = queryset.filter(created_at__date__lte=registered_date_to)

        # Price Range Filters
        purchase_price_from = self.request.GET.get('purchase_price_from')
        purchase_price_to = self.request.GET.get('purchase_price_to')

        if purchase_price_from:
            try:
                queryset = queryset.filter(purchase_price__gte=Decimal(purchase_price_from))
            except Exception:
                pass
        if purchase_price_to:
            try:
                queryset = queryset.filter(purchase_price__lte=Decimal(purchase_price_to))
            except Exception:
                pass

        # Current Value (Net Book Value) Range Filters
        current_value_from = self.request.GET.get('current_value_from')
        current_value_to = self.request.GET.get('current_value_to')

        if current_value_from or current_value_to:
            try:
                # current_value is computed on the model, so filter after evaluating the queryset.
                assets_list = list(queryset)
                current_value_from_decimal = Decimal(current_value_from) if current_value_from else None
                current_value_to_decimal = Decimal(current_value_to) if current_value_to else None

                filtered_assets = []
                for asset in assets_list:
                    asset_current_value = asset.current_value
                    include = True

                    if current_value_from_decimal is not None and asset_current_value < current_value_from_decimal:
                        include = False
                    if current_value_to_decimal is not None and asset_current_value > current_value_to_decimal:
                        include = False

                    if include:
                        filtered_assets.append(asset)

                asset_ids = [asset.id for asset in filtered_assets]
                queryset = queryset.filter(id__in=asset_ids)
            except Exception:
                pass
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        org = self.request.user.organization
        user = self.request.user
        
        # Determine if financial data should be shown - for CHECKER and above roles
        show_financial = user.role in [user.Role.CHECKER, user.Role.SENIOR_MANAGER, user.Role.ADMIN] or user.is_superuser
        context['show_financial'] = show_financial
        
        # Calculate total values for filtered assets
        if show_financial:
            from django.db.models import Sum, Count
            from django.db.models.functions import Coalesce
            
            filtered_queryset = self.get_queryset()
            
            # Calculate aggregates
            aggregates = filtered_queryset.aggregate(
                total_purchase_value=Coalesce(Sum('purchase_price'), Decimal('0')),
                total_asset_count=Count('id')
            )
            
            context['total_purchase_value'] = aggregates['total_purchase_value']
            context['filtered_asset_count'] = aggregates['total_asset_count']
            
            # Calculate current value (NBV) for filtered assets
            # For performance, only calculate if there are reasonable number of assets
            if aggregates['total_asset_count'] <= 1000:
                # Direct calculation for small datasets
                filtered_assets = list(filtered_queryset)
                total_current_value = sum(asset.current_value for asset in filtered_assets if hasattr(asset, 'current_value'))
                total_accumulated_depreciation = sum(asset.accumulated_depreciation for asset in filtered_assets if hasattr(asset, 'accumulated_depreciation'))
            else:
                # Estimate for large datasets
                sample_size = 500
                sample_assets = list(filtered_queryset[:sample_size])
                if sample_assets:
                    avg_depreciation_ratio = sum(asset.accumulated_depreciation for asset in sample_assets) / sum(asset.purchase_price or Decimal('0') for asset in sample_assets if asset.purchase_price) if any(asset.purchase_price for asset in sample_assets) else 0
                    total_accumulated_depreciation = aggregates['total_purchase_value'] * Decimal(str(avg_depreciation_ratio))
                    total_current_value = aggregates['total_purchase_value'] - total_accumulated_depreciation
                else:
                    total_accumulated_depreciation = Decimal('0')
                    total_current_value = aggregates['total_purchase_value']
            
            context['total_current_value'] = total_current_value
            context['total_accumulated_depreciation'] = total_accumulated_depreciation
            
            # Pass date filter values to context
            context['purchase_date_from'] = self.request.GET.get('purchase_date_from', '')
            context['purchase_date_to'] = self.request.GET.get('purchase_date_to', '')
        
        # Use only() to reduce database load for filter dropdowns
        context['categories'] = Category.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['sites'] = Site.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['buildings'] = Building.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['brands'] = Brand.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['departments'] = Department.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['subcategories'] = SubCategory.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['subgroups'] = SubGroup.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['groups'] = Group.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['branches'] = Branch.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['floors'] = Floor.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['rooms'] = Room.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['locations'] = Location.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['sublocations'] = SubLocation.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['suppliers'] = Supplier.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['conditions'] = Asset.Condition.choices
        context['label_types'] = Asset.LabelType.choices
        context['statuses'] = Asset.Status.choices

        from django.contrib.auth import get_user_model
        User = get_user_model()
        context['creators'] = User.objects.filter(organization=org).only('id', 'first_name', 'last_name', 'username').order_by('first_name', 'last_name', 'username')

        if self.request.GET.get('view') == 'depreciation' and show_financial:
            # Depreciation report with efficient aggregation
            from django.db.models import Sum, Count, Q
            from django.db.models.functions import Coalesce
            from datetime import datetime
            
            queryset = self.get_queryset()
            
            # CRITICAL: Only show assets with purchase price (exclude assets without value)
            queryset = queryset.filter(purchase_price__isnull=False, purchase_price__gt=0)
            
            # Add date filtering for depreciation report
            depr_date_from = self.request.GET.get('depr_date_from')
            depr_date_to = self.request.GET.get('depr_date_to')
            
            opening_date = None
            closing_date = None
            
            if depr_date_from:
                try:
                    opening_date = datetime.strptime(depr_date_from, '%Y-%m-%d').date()
                    # Don't filter by purchase date, we want all assets that existed at opening date
                except (ValueError, TypeError):
                    pass
            
            if depr_date_to:
                try:
                    closing_date = datetime.strptime(depr_date_to, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    pass
            
            # Filter to only show assets purchased on or before the closing date (if provided)
            if closing_date:
                queryset = queryset.filter(Q(purchase_date__lte=closing_date) | Q(purchase_date__isnull=True))
            
            # Add location and category filters for depreciation report
            depr_filters = {
                'depr_category': 'category_id',
                'depr_group': 'group_id',
                'depr_department': 'department_id',
                'depr_site': 'site_id',
                'depr_branch': 'branch_id',
                'depr_building': 'building_id',
                'depr_location': 'location_id',
                'depr_tagging_status': 'tagging_status',
            }
            
            for param, field in depr_filters.items():
                val = self.request.GET.get(param)
                if val:
                    queryset = queryset.filter(**{field: val})
                    context[param] = val
            
            # Use database aggregation directly for efficiency
            # Only aggregate purchase_price (a database field)
            agg = queryset.aggregate(
                total_cost=Coalesce(Sum('purchase_price'), Decimal('0')),
                total_count=Count('id')
            )
            
            total_cost = agg['total_cost']
            total_count = agg['total_count']
            
            # Calculate opening and closing values
            total_opening_value = Decimal('0')
            total_closing_value = Decimal('0')
            
            # Calculate exact values by iterating all assets in batches
            # This matches the dashboard calculation for consistency
            BATCH_SIZE = 1000
            
            total_acc_dep = Decimal('0')
            total_nbv = Decimal('0')
            
            if opening_date or closing_date:
                # Need to calculate period-based values
                for i in range(0, total_count, BATCH_SIZE):
                    batch = list(queryset[i:i+BATCH_SIZE])
                    for asset in batch:
                        if opening_date:
                            total_opening_value += asset.get_value_at_date(opening_date)
                        else:
                            total_opening_value += asset.current_value
                        
                        if closing_date:
                            total_closing_value += asset.get_value_at_date(closing_date)
                        else:
                            total_closing_value += asset.current_value
                        
                        total_acc_dep += asset.get_accumulated_dep_at_date(closing_date) if closing_date else asset.accumulated_depreciation
                
                total_nbv = total_closing_value
            else:
                # No date range - just compute current values
                for i in range(0, total_count, BATCH_SIZE):
                    batch = list(queryset[i:i+BATCH_SIZE])
                    for asset in batch:
                        cv = asset.current_value
                        total_nbv += cv
                        total_acc_dep += asset.accumulated_depreciation
                
                total_opening_value = total_nbv
                total_closing_value = total_nbv
            
            # Calculate depreciation for the period
            period_depreciation = total_opening_value - total_closing_value
            
            context['total_cost'] = total_cost
            context['total_opening_value'] = total_opening_value
            context['total_closing_value'] = total_closing_value
            context['period_depreciation'] = period_depreciation
            context['total_acc_dep'] = total_acc_dep
            context['total_nbv'] = total_nbv
            context['is_report'] = True
            context['total_assets_report'] = total_count
            context['depr_date_from'] = depr_date_from
            context['depr_date_to'] = depr_date_to
            context['opening_date'] = opening_date
            context['closing_date'] = closing_date
            
            # Support grouped summaries
            group_by = self.request.GET.get('group_by')
            context['group_by'] = group_by
            
            if group_by == 'category':
                # Efficient category grouping without loading all assets
                grouped_data = queryset.values('category', 'category__name').annotate(
                    count=Count('id'),
                    total_cost=Sum('purchase_price')
                ).order_by('-total_cost')[:100]  # Limit to top 100 categories
                
                # Calculate exact depreciation for each category
                grouped_list = []
                for group in grouped_data:
                    cat_id = group['category']
                    cat_qs = queryset.filter(category_id=cat_id)
                    
                    cat_assets = list(cat_qs)
                    total_cat_dep = sum(a.get_accumulated_dep_at_date(closing_date) if closing_date else a.accumulated_depreciation for a in cat_assets) if cat_assets else Decimal('0')
                    
                    grouped_list.append({
                        'id': cat_id,
                        'name': group['category__name'] or 'Uncategorized',
                        'total_cost': group['total_cost'] or Decimal('0'),
                        'total_acc_dep': total_cat_dep,
                        'total_nbv': (group['total_cost'] or Decimal('0')) - total_cat_dep,
                        'count': group['count'],
                    })
                context['grouped_data'] = grouped_list
            
            # Paginate the filtered depreciation assets
            from django.core.paginator import Paginator
            paginator = Paginator(queryset, 25)  # 25 items per page
            page_number = self.request.GET.get('page', 1)
            page_obj = paginator.get_page(page_number)
            
            # Add opening and closing values to each asset
            assets_with_values = []
            for asset in page_obj.object_list:
                asset_dict = asset
                if opening_date:
                    asset_dict.opening_value = asset.get_value_at_date(opening_date)
                else:
                    asset_dict.opening_value = asset.current_value
                
                if closing_date:
                    asset_dict.closing_value = asset.get_value_at_date(closing_date)
                else:
                    asset_dict.closing_value = asset.current_value
                
                asset_dict.period_depreciation = asset_dict.opening_value - asset_dict.closing_value
                asset_dict.display_acc_dep = asset.get_accumulated_dep_at_date(closing_date) if closing_date else asset.accumulated_depreciation
                assets_with_values.append(asset_dict)
            
            context['assets'] = assets_with_values
            context['page_obj'] = page_obj
            context['is_paginated'] = page_obj.has_other_pages()
        
        # Ensure filters persist during pagination
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['query_params'] = query_params.urlencode()
            
        return context

class BulkAssetActionView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        asset_ids = request.POST.getlist('asset_ids')
        action = request.POST.get('action')
        
        if not asset_ids:
            messages.warning(request, "No assets selected.")
            return redirect('asset-list')
            
        assets = Asset.objects.filter(
            id__in=asset_ids, 
            organization=request.user.organization
        )
        
        count = assets.count()
        
        if action == 'delete':
            assets.update(is_deleted=True)
            invalidate_dashboard_cache_for_org(request.user.organization)
            messages.success(request, f"Successfully deleted {count} assets.")
        elif action.startswith('status_'):
            new_status = action.replace('status_', '').upper()
            assets.update(status=new_status)
            invalidate_dashboard_cache_for_org(request.user.organization)
            messages.success(request, f"Successfully updated status for {count} assets.")
            
        return redirect('asset-list')

class ExportAssetExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        # We reuse the logic from AssetListView to respect current filters
        view = AssetListView()
        view.request = request
        queryset = view.get_queryset()

        is_depreciation_view = request.GET.get('view') == 'depreciation'

        opening_date = None
        closing_date = None

        if is_depreciation_view:
            # Match depreciation report behavior: only assets with value
            queryset = queryset.filter(purchase_price__isnull=False, purchase_price__gt=0)

            # Apply depreciation dimension filters from report pages
            depr_filters = {
                'depr_category': 'category_id',
                'depr_group': 'group_id',
                'depr_department': 'department_id',
                'depr_site': 'site_id',
                'depr_branch': 'branch_id',
                'depr_building': 'building_id',
                'depr_location': 'location_id',
                'depr_tagging_status': 'tagging_status',
            }

            for param, field in depr_filters.items():
                val = request.GET.get(param)
                if val:
                    queryset = queryset.filter(**{field: val})

            depr_date_from = request.GET.get('depr_date_from')
            depr_date_to = request.GET.get('depr_date_to')

            if depr_date_from:
                try:
                    opening_date = datetime.strptime(depr_date_from, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    opening_date = None

            if depr_date_to:
                try:
                    closing_date = datetime.strptime(depr_date_to, '%Y-%m-%d').date()
                except (ValueError, TypeError):
                    closing_date = None

            # Keep only assets existing up to closing date
            if closing_date:
                queryset = queryset.filter(Q(purchase_date__lte=closing_date) | Q(purchase_date__isnull=True))
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Depreciation Export" if is_depreciation_view else "Assets Export"

        def resolve_name(model_cls, value):
            if not value:
                return ''
            obj = model_cls.objects.filter(pk=value).only('id', 'name').first()
            return obj.name if obj else str(value)

        def resolve_creator(value):
            if not value:
                return ''
            from django.contrib.auth import get_user_model
            User = get_user_model()
            user_obj = User.objects.filter(pk=value).only('id', 'first_name', 'last_name', 'username').first()
            if not user_obj:
                return str(value)
            full_name = (user_obj.get_full_name() or '').strip()
            return full_name or user_obj.username

        status_map = dict(Asset.Status.choices)
        condition_map = dict(Asset.Condition.choices)
        label_type_map = dict(Asset.LabelType.choices)

        applied_filters = []
        if request.GET.get('q'):
            applied_filters.append(("Search", request.GET.get('q')))
        if request.GET.get('product_name'):
            applied_filters.append(("Product Name", request.GET.get('product_name')))
        if request.GET.get('status'):
            status_val = request.GET.get('status')
            applied_filters.append(("Status", status_map.get(status_val, status_val)))
        if request.GET.get('condition'):
            condition_val = request.GET.get('condition')
            applied_filters.append(("Condition", condition_map.get(condition_val, condition_val)))
        if request.GET.get('is_tagged'):
            tagged_val = request.GET.get('is_tagged')
            applied_filters.append(("Tagged Status", "Tagged" if tagged_val == 'tagged' else "Untagged"))
        if request.GET.get('tag_type'):
            applied_filters.append(("Type of Tag", request.GET.get('tag_type')))
        if request.GET.get('label_type'):
            label_type_val = request.GET.get('label_type')
            applied_filters.append(("Label Type", label_type_map.get(label_type_val, label_type_val)))

        fk_filters = [
            ("Category", 'category', Category),
            ("Sub Category", 'subcategory', SubCategory),
            ("Group", 'group', Group),
            ("Sub Group", 'sub_group', SubGroup),
            ("Brand", 'brand', Brand),
            ("Site", 'site', Site),
            ("Building", 'building', Building),
            ("Floor", 'floor', Floor),
            ("Room", 'room', Room),
            ("Location", 'location', Location),
            ("Sub Location", 'sub_location', SubLocation),
            ("Department", 'department', Department),
            ("Supplier", 'supplier', Supplier),
        ]

        for label, param, model_cls in fk_filters:
            raw_val = request.GET.get(param)
            if raw_val:
                applied_filters.append((label, resolve_name(model_cls, raw_val)))

        if request.GET.get('created_by'):
            applied_filters.append(("Registered By", resolve_creator(request.GET.get('created_by'))))

        if request.GET.get('purchase_date_from') or request.GET.get('purchase_date_to'):
            from_val = request.GET.get('purchase_date_from', '')
            to_val = request.GET.get('purchase_date_to', '')
            applied_filters.append(("Purchase Date Range", f"{from_val} to {to_val}"))

        if request.GET.get('purchase_price_from') or request.GET.get('purchase_price_to'):
            from_val = request.GET.get('purchase_price_from', '')
            to_val = request.GET.get('purchase_price_to', '')
            applied_filters.append(("Purchase Price Range", f"{from_val} to {to_val}"))

        if request.GET.get('current_value_from') or request.GET.get('current_value_to'):
            from_val = request.GET.get('current_value_from', '')
            to_val = request.GET.get('current_value_to', '')
            applied_filters.append(("Current Value Range", f"{from_val} to {to_val}"))

        if request.GET.get('registered_date_from') or request.GET.get('registered_date_to'):
            from_val = request.GET.get('registered_date_from', '')
            to_val = request.GET.get('registered_date_to', '')
            applied_filters.append(("Registered Date Range", f"{from_val} to {to_val}"))

        if is_depreciation_view:
            depr_fk_filters = [
                ("Depreciation Category", 'depr_category', Category),
                ("Depreciation Group", 'depr_group', Group),
                ("Depreciation Department", 'depr_department', Department),
                ("Depreciation Site", 'depr_site', Site),
                ("Depreciation Branch", 'depr_branch', Branch),
                ("Depreciation Building", 'depr_building', Building),
                ("Depreciation Location", 'depr_location', Location),
            ]
            for label, param, model_cls in depr_fk_filters:
                raw_val = request.GET.get(param)
                if raw_val:
                    applied_filters.append((label, resolve_name(model_cls, raw_val)))

            if request.GET.get('depr_date_from') or request.GET.get('depr_date_to'):
                from_val = request.GET.get('depr_date_from', '')
                to_val = request.GET.get('depr_date_to', '')
                applied_filters.append(("Depreciation Date Range", f"{from_val} to {to_val}"))

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
        from openpyxl.utils import get_column_letter

        # Style palette
        TITLE_FILL = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
        META_LABEL_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        META_VALUE_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        HEADER_FILL = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
        ROW_ALT_FILL = PatternFill(start_color='F8F9FB', end_color='F8F9FB', fill_type='solid')
        TOTAL_FILL = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
        SUMMARY_HEADER_FILL = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
        SUMMARY_LABEL_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

        TITLE_FONT = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        META_LABEL_FONT = Font(name='Calibri', size=10, bold=True, color='1F3864')
        META_VALUE_FONT = Font(name='Calibri', size=10, color='000000')
        HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        BODY_FONT = Font(name='Calibri', size=10, color='000000')
        TOTAL_FONT = Font(name='Calibri', size=11, bold=True, color='000000')
        SUMMARY_HEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')

        thin = Side(border_style='thin', color='BFBFBF')
        BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
        thick = Side(border_style='medium', color='1F3864')
        TOTAL_BORDER = Border(left=thin, right=thin, top=thick, bottom=thick)

        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right = Alignment(horizontal='right', vertical='center')

        # Build headers first so we know the column count for the title merge
        if is_depreciation_view:
            headers = [
                'Asset Tag', 'Name', 'Category', 'Status',
                'Purchase Date', 'Purchase Price', 'Currency'
            ]
            if opening_date:
                headers.append(f'Opening Value ({opening_date.strftime("%Y-%m-%d")})')
            if closing_date:
                headers.append(f'Closing Value ({closing_date.strftime("%Y-%m-%d")})')
            if opening_date and closing_date:
                headers.append('Period Depreciation')
            headers.extend(['Accumulated Depreciation', 'Current NBV'])
        else:
            headers = [
                # Identification
                'Asset Tag', 'Asset Code', 'ERP Asset Number', 'Name', 'Short Description', 'Description',
                'Serial Number', 'Quantity', 'Label Type', 'Asset Type', 'Status', 'Condition',
                # Classification
                'Category', 'Sub Category', 'Group', 'Sub Group', 'Brand', 'Model',
                # Ownership
                'Company', 'Department', 'Cost Center', 'Assigned To', 'Custodian', 'Employee Number', 'Supplier', 'Vendor',
                # Location
                'Region', 'Site', 'Branch', 'Building', 'Floor', 'Room', 'Location', 'Sub Location',
                # Financial
                'Currency', 'Purchase Price', 'Salvage Value', 'Purchase Date', 'Invoice Number', 'Invoice Date',
                'PO Number', 'PO Date', 'DO Number', 'DO Date', 'GRN Number', 'Date Placed in Service', 'Tagged Date',
                # Warranty / Insurance / AMC
                'Warranty Start', 'Warranty End', 'Insurance Start', 'Insurance End', 'Maintenance Start', 'Maintenance End',
                # Depreciation
                'Depreciation Method', 'Useful Life (Years)', 'Accumulated Depreciation', 'Current NBV',
                # Maintenance
                'Maintenance Required', 'Maintenance Frequency (days)', 'Next Maintenance Date',
                # Misc / Audit
                'Notes', 'Registered By', 'Registered On',
            ]

        n_cols = len(headers)
        last_col_letter = get_column_letter(n_cols)

        # ---- Title row ----
        title_text = ('Asset Depreciation Report' if is_depreciation_view
                      else 'Asset Inventory Report')
        ws.append([title_text])
        ws.merge_cells(f'A1:{last_col_letter}1')
        title_cell = ws.cell(row=1, column=1)
        title_cell.font = TITLE_FONT
        title_cell.fill = TITLE_FILL
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # ---- Metadata block (label | value spans rest) ----
        def add_meta(label, value):
            ws.append([label, value] + [''] * (n_cols - 2))
            r = ws.max_row
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=n_cols)
            lc = ws.cell(row=r, column=1)
            vc = ws.cell(row=r, column=2)
            lc.font = META_LABEL_FONT
            lc.fill = META_LABEL_FILL
            lc.alignment = left
            lc.border = BORDER
            vc.font = META_VALUE_FONT
            vc.fill = META_VALUE_FILL
            vc.alignment = left
            vc.border = BORDER

        add_meta('Export Type', 'Depreciation' if is_depreciation_view else 'Assets')
        add_meta('Exported At', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        if applied_filters:
            add_meta('Applied Filters', '')
            for label, value in applied_filters:
                add_meta(label, str(value))
        else:
            add_meta('Applied Filters', 'None')

        # Spacer row
        ws.append([])

        # ---- Header row ----
        ws.append(headers)
        header_row_idx = ws.max_row
        for cell in ws[header_row_idx]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = center
            cell.border = BORDER
        ws.row_dimensions[header_row_idx].height = 26

        # Identify numeric columns (1-based indexes) to total in the summary row
        if is_depreciation_view:
            numeric_col_indexes = [6]  # Purchase Price
            col = 8  # next column after Currency (7)
            if opening_date:
                numeric_col_indexes.append(col); col += 1
            if closing_date:
                numeric_col_indexes.append(col); col += 1
            if opening_date and closing_date:
                numeric_col_indexes.append(col); col += 1
            numeric_col_indexes.extend([col, col + 1])  # Accumulated Depreciation, Current NBV
        else:
            # Quantity, Purchase Price, Salvage Value, Useful Life, Accumulated Depreciation, Current NBV, Maintenance Frequency
            numeric_col_indexes = [8, 36, 37, 55, 56, 57, 59]

        totals = {idx: 0.0 for idx in numeric_col_indexes}
        row_count = 0

        for asset in queryset:
            if is_depreciation_view:
                opening_value = asset.get_value_at_date(opening_date) if opening_date else asset.current_value
                closing_value = asset.get_value_at_date(closing_date) if closing_date else asset.current_value

                # If opening date is provided and asset has no opening value, skip it
                if opening_date and opening_value <= 0:
                    continue

                row = [
                    asset.asset_tag,
                    asset.name,
                    asset.category.name if asset.category else '',
                    asset.get_status_display(),
                    asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '',
                    float(asset.purchase_price) if asset.purchase_price else 0,
                    asset.currency,
                ]

                if opening_date:
                    row.append(float(opening_value))
                if closing_date:
                    row.append(float(closing_value))
                if opening_date and closing_date:
                    row.append(float(opening_value - closing_value))

                row.extend([
                    float(asset.get_accumulated_dep_at_date(closing_date) if closing_date else asset.accumulated_depreciation),
                    float(closing_value),
                ])

                ws.append(row)
            else:
                def _d(d):
                    return d.strftime('%Y-%m-%d') if d else ''
                creator = ''
                if asset.created_by:
                    creator = (asset.created_by.get_full_name() or '').strip() or asset.created_by.username
                assigned = ''
                if asset.assigned_to:
                    assigned = (asset.assigned_to.get_full_name() or '').strip() or asset.assigned_to.username
                elif asset.custodian and asset.custodian.user:
                    assigned = (asset.custodian.user.get_full_name() or '').strip() or asset.custodian.user.username
                brand_val = ''
                if getattr(asset, 'brand_new', None):
                    brand_val = asset.brand_new.name
                elif asset.brand:
                    brand_val = asset.brand
                row = [
                    # Identification
                    asset.asset_tag,
                    asset.asset_code or '',
                    asset.erp_asset_number or '',
                    asset.name,
                    asset.short_description or '',
                    asset.description or '',
                    asset.serial_number or '',
                    int(asset.quantity or 0),
                    asset.get_label_type_display() if asset.label_type else '',
                    asset.get_asset_type_display() if asset.asset_type else '',
                    asset.get_status_display(),
                    asset.get_condition_display(),
                    # Classification
                    asset.category.name if asset.category else '',
                    asset.sub_category.name if asset.sub_category else '',
                    asset.group.name if asset.group else '',
                    asset.sub_group.name if asset.sub_group else '',
                    brand_val,
                    asset.model or '',
                    # Ownership
                    asset.company.name if asset.company else '',
                    asset.department.name if asset.department else '',
                    asset.cost_center or '',
                    assigned,
                    str(asset.custodian) if asset.custodian else '',
                    asset.employee_number or '',
                    asset.supplier.name if asset.supplier else '',
                    asset.vendor.name if asset.vendor else '',
                    # Location
                    asset.region.name if asset.region else '',
                    asset.site.name if asset.site else '',
                    asset.branch.name if asset.branch else '',
                    asset.building.name if asset.building else '',
                    asset.floor.name if asset.floor else '',
                    str(asset.room) if asset.room else '',
                    asset.location.name if asset.location else '',
                    asset.sub_location.name if asset.sub_location else '',
                    # Financial
                    asset.currency or '',
                    float(asset.purchase_price) if asset.purchase_price else 0,
                    float(asset.salvage_value) if asset.salvage_value else 0,
                    _d(asset.purchase_date),
                    asset.invoice_number or '',
                    _d(asset.invoice_date),
                    asset.po_number or '',
                    _d(asset.po_date),
                    asset.do_number or '',
                    _d(asset.do_date),
                    asset.grn_number or '',
                    _d(asset.date_placed_in_service),
                    _d(asset.tagged_date),
                    # Warranty / Insurance / AMC
                    _d(asset.warranty_start),
                    _d(asset.warranty_end),
                    _d(asset.insurance_start_date),
                    _d(asset.insurance_end_date),
                    _d(asset.maintenance_start_date),
                    _d(asset.maintenance_end_date),
                    # Depreciation
                    asset.get_depreciation_method_display() if asset.depreciation_method else '',
                    int(asset.useful_life_years) if asset.useful_life_years else 0,
                    float(asset.accumulated_depreciation),
                    float(asset.current_value),
                    # Maintenance
                    'Yes' if asset.maintenance_required else 'No',
                    int(asset.maintenance_frequency_days or 0),
                    _d(asset.next_maintenance_date),
                    # Misc / Audit
                    asset.notes or '',
                    creator,
                    asset.created_at.strftime('%Y-%m-%d %H:%M') if getattr(asset, 'created_at', None) else '',
                ]
                ws.append(row)

            row_count += 1
            for idx in numeric_col_indexes:
                if idx <= len(row):
                    val = row[idx - 1]
                    if isinstance(val, (int, float)):
                        totals[idx] += float(val)

        data_start_row = header_row_idx + 1
        data_end_row = ws.max_row
        NUMBER_FORMAT = '#,##0.00'
        numeric_set = set(numeric_col_indexes)

        # Style data rows: borders, alignment, banding, number formats
        if row_count > 0:
            for r in range(data_start_row, data_end_row + 1):
                is_alt = ((r - data_start_row) % 2 == 1)
                for c in range(1, n_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.font = BODY_FONT
                    cell.border = BORDER
                    if c in numeric_set:
                        cell.alignment = right
                        cell.number_format = NUMBER_FORMAT
                    else:
                        cell.alignment = left
                    if is_alt:
                        cell.fill = ROW_ALT_FILL
                ws.row_dimensions[r].height = 18

        # ---- TOTAL row ----
        if row_count > 0:
            total_row = [''] * n_cols
            total_row[0] = 'TOTAL'
            for idx in numeric_col_indexes:
                if idx <= len(total_row):
                    total_row[idx - 1] = round(totals[idx], 2)
            ws.append(total_row)
            total_row_idx = ws.max_row
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=total_row_idx, column=c)
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
                cell.border = TOTAL_BORDER
                if c in numeric_set:
                    cell.alignment = right
                    cell.number_format = NUMBER_FORMAT
                elif c == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = left
            ws.row_dimensions[total_row_idx].height = 22

            # ---- Summary card ----
            ws.append([])  # spacer

            # Summary header (merged)
            ws.append(['Summary'] + [''] * (n_cols - 1))
            sh_row = ws.max_row
            ws.merge_cells(start_row=sh_row, start_column=1, end_row=sh_row, end_column=min(3, n_cols))
            sh_cell = ws.cell(row=sh_row, column=1)
            sh_cell.font = SUMMARY_HEADER_FONT
            sh_cell.fill = SUMMARY_HEADER_FILL
            sh_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[sh_row].height = 24
            for c in range(1, min(3, n_cols) + 1):
                ws.cell(row=sh_row, column=c).border = BORDER

            def add_summary_row(label, value, is_number=False):
                ws.append([label, value] + [''] * (n_cols - 2))
                r = ws.max_row
                lc = ws.cell(row=r, column=1)
                vc = ws.cell(row=r, column=2)
                lc.font = META_LABEL_FONT
                lc.fill = SUMMARY_LABEL_FILL
                lc.alignment = left
                lc.border = BORDER
                vc.font = Font(name='Calibri', size=10, bold=True, color='000000')
                vc.fill = META_VALUE_FILL
                vc.alignment = right if is_number else left
                vc.border = BORDER
                if is_number:
                    vc.number_format = NUMBER_FORMAT

            add_summary_row('Total Records', row_count)
            for idx in numeric_col_indexes:
                label = headers[idx - 1] if idx - 1 < len(headers) else f'Column {idx}'
                add_summary_row(f'Total {label}', round(totals[idx], 2), is_number=True)
        else:
            # No data row notice
            ws.append(['No records found for the selected filters.'] + [''] * (n_cols - 1))
            r = ws.max_row
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
            cell = ws.cell(row=r, column=1)
            cell.font = Font(italic=True, color='808080')
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # ---- Column widths (auto-fit, capped) ----
        for col_idx in range(1, n_cols + 1):
            max_len = len(str(headers[col_idx - 1]))
            for r in range(data_start_row, ws.max_row + 1):
                v = ws.cell(row=r, column=col_idx).value
                if v is None:
                    continue
                # Format numbers similar to display for width calc
                if isinstance(v, (int, float)):
                    s = f'{v:,.2f}'
                else:
                    s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 4, 12), 40)

        # Freeze header row
        ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

        # Add auto-filter on the data range
        if row_count > 0:
            ws.auto_filter.ref = f'A{header_row_idx}:{last_col_letter}{data_end_row}'

        # Print setup
        ws.print_options.horizontalCentered = True
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.print_title_rows = f'{header_row_idx}:{header_row_idx}'

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        file_prefix = 'depreciation_export' if is_depreciation_view else 'assets_export'
        response['Content-Disposition'] = f'attachment; filename="{file_prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response

class AssetCreateView(LoginRequiredMixin, CreateView):
    model = Asset
    form_class = AssetForm
    template_name = 'assets/asset_form.html'
    success_url = reverse_lazy('asset-list')

    def get(self, request, *args, **kwargs):
        try:
            return super().get(request, *args, **kwargs)
        except Exception:
            import traceback
            print(traceback.format_exc())
            raise

    def post(self, request, *args, **kwargs):
        try:
            return super().post(request, *args, **kwargs)
        except Exception:
            import traceback
            print(traceback.format_exc())
            raise

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        user = self.request.user

        # Data Entry users submit for approval instead of direct inventory creation
        if getattr(user, 'role', None) == user.Role.EMPLOYEE:
            approval_data = self._build_approval_payload(form)
            ApprovalRequest.objects.create(
                organization=user.organization,
                requester=user,
                request_type=ApprovalRequest.RequestType.ASSET_CREATE,
                status=ApprovalRequest.Status.PENDING,
                data=approval_data,
                comments='Asset uploaded by employee and waiting for checker approval.'
            )
            messages.success(
                self.request,
                'Asset upload submitted for checker approval. It will be added to inventory after approval.'
            )
            return redirect('approval_list')

        form.instance.organization = self.request.user.organization
        form.instance.created_by = self.request.user
        # Keep assignment in sync with selected custodian on asset creation.
        form.instance.assigned_to = form.instance.custodian.user if form.instance.custodian and form.instance.custodian.user else None
        
        # Handle bulk creation based on quantity
        quantity = form.instance.quantity if form.instance.quantity else 1
        
        if quantity > 1:
            # Create multiple assets with different auto-generated tags
            # Set quantity to 1 for each individual asset
            original_quantity = form.instance.quantity
            form.instance.quantity = 1
            
            # Create first asset with form data and auto-generated tag
            if not form.instance.asset_tag:
                form.instance.asset_tag = generate_asset_tag(
                    self.request.user.organization,
                    form.instance.category,
                    form.instance.company
                )
            
            response = super().form_valid(form)
            first_asset = form.instance
            
            # Create remaining assets one by one with their own generated tags
            # This ensures each call to generate_asset_tag sees the previously created assets
            for i in range(1, original_quantity):
                # Clone the first saved instance
                new_asset = Asset.objects.get(pk=first_asset.pk)
                new_asset.pk = None  # Clear primary key to create new record
                new_asset.id = None  # Clear UUID
                
                # Generate new tag for this asset
                new_asset.asset_tag = generate_asset_tag(
                    self.request.user.organization,
                    form.instance.category,
                    form.instance.company
                )
                new_asset.save()

            invalidate_dashboard_cache_for_org(self.request.user.organization)
            
            return response
        else:
            # Single asset creation with auto-generated tag if not provided
            if not form.instance.asset_tag:
                form.instance.asset_tag = generate_asset_tag(
                    self.request.user.organization,
                    form.instance.category,
                    form.instance.company
                )
            response = super().form_valid(form)
            invalidate_dashboard_cache_for_org(self.request.user.organization)
            return response

    def _build_approval_payload(self, form):
        """Serialize asset form data for deferred creation after checker approval."""
        from django.db import models

        cleaned_data = form.cleaned_data
        payload = {}
        file_fields = {}

        model_fields = {
            field.name: field
            for field in Asset._meta.get_fields()
            if isinstance(field, models.Field)
        }
        excluded_fields = {
            'id', 'organization', 'created_by', 'asset_tag', 'custom_fields',
            'is_deleted', 'deleted_at', 'created_at', 'updated_at'
        }

        # Only serialize fields that are actually submitted in the form
        for field_name, value in cleaned_data.items():
            field = model_fields.get(field_name)
            if not field:
                continue
            if field.auto_created or field_name in excluded_fields:
                continue

            if isinstance(field, models.FileField):
                uploaded = self.request.FILES.get(field_name)
                if uploaded:
                    saved_name = default_storage.save(
                        f"approval_uploads/{uuid4()}_{uploaded.name}",
                        ContentFile(uploaded.read())
                    )
                    file_fields[field_name] = saved_name
                continue

            if isinstance(field, models.ForeignKey):
                payload[field_name] = str(value.pk) if value else None
            elif value is None:
                payload[field_name] = None
            elif isinstance(value, Decimal):
                payload[field_name] = str(value)
            elif isinstance(value, (date, datetime)):
                payload[field_name] = value.isoformat()
            else:
                payload[field_name] = value

        return {
            'asset_name': cleaned_data.get('name', ''),
            'asset_category': cleaned_data.get('category').name if cleaned_data.get('category') else '',
            'asset_quantity': cleaned_data.get('quantity', 1),
            'asset_description': cleaned_data.get('description', ''),
            'asset_payload': payload,
            'file_fields': file_fields,
        }


class AssetImportView(LoginRequiredMixin, FormView):
    template_name = 'assets/asset_import.html'
    form_class = AssetImportForm
    success_url = reverse_lazy('asset-list')

    def post(self, request, *args, **kwargs):
        # Handle the confirm step (no file upload, data is in session)
        if 'confirm_create' in request.POST:
            return self._handle_confirm_import(request)
        return super().post(request, *args, **kwargs)

    def get_file_data(self, uploaded_file):
        """Extract data rows from CSV or Excel file efficiently."""
        data = []
        filename = uploaded_file.name.lower()
        
        if filename.endswith('.csv'):
            try:
                # Use a generator or list but decoded efficiently
                decoded_file = uploaded_file.read().decode('utf-8-sig')
                io_string = io.StringIO(decoded_file)
                reader = csv.DictReader(io_string)
                raw_data = list(reader)
                # Normalize headers to lowercase with underscores
                for row in raw_data:
                    data.append(self._normalize_row_keys(row))
            except Exception as e:
                raise ValueError(f"Error reading CSV: {str(e)}")
        
        elif filename.endswith('.xlsx'):
            try:
                # Optimized Excel reading for large files
                wb = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True)
                sheet = wb.active
                
                # Get headers from the first row and normalize
                rows_gen = sheet.iter_rows(values_only=True)
                raw_headers = [h for h in next(rows_gen) if h]
                headers = [self._normalize_header(h) for h in raw_headers]
                
                for row in rows_gen:
                    if any(row):  # Skip empty rows
                        row_dict = dict(zip(headers, row))
                        data.append(row_dict)
                wb.close() # Important for read_only=True
            except Exception as e:
                raise ValueError(f"Error reading Excel: {str(e)}")
        
        return data

    # Header alias mapping: maps common variations to the expected field names
    HEADER_ALIASES = {
        'building name': 'building', 'bldg': 'building', 'bldg name': 'building',
        'floor name': 'floor', 'floor no': 'floor', 'floor number': 'floor',
        'room name': 'room', 'room no': 'room', 'room number': 'room',
        'branch name': 'branch', 'site name': 'site', 'region name': 'region',
        'location name': 'location', 'sub location': 'sub_location',
        'sub location name': 'sub_location', 'sublocation': 'sub_location',
        'sub category': 'sub_category', 'subcategory': 'sub_category',
        'sub group': 'sub_group', 'subgroup': 'sub_group',
        'department name': 'department', 'dept': 'department',
        'brand name': 'brand', 'company name': 'company',
        'supplier name': 'supplier', 'vendor name': 'vendor',
        'asset name': 'name', 'asset description': 'description',
        'asset tag': 'asset_tag',
        'erp asset number': 'erp_asset_number', 'erp number': 'erp_asset_number',
        'asset code': 'asset_code', 'asset type': 'asset_type',
        'label type': 'label_type', 'serial number': 'serial_number',
        'serial no': 'serial_number', 'cost center': 'cost_center',
        'employee number': 'employee_number', 'employee no': 'employee_number',
        'employee id': 'employee_number', 'emp no': 'employee_number',
        'purchase date': 'purchase_date', 'purchase price': 'purchase_price',
        'invoice number': 'invoice_number', 'invoice no': 'invoice_number',
        'invoice date': 'invoice_date', 'po number': 'po_number',
        'po date': 'po_date', 'do number': 'do_number', 'do date': 'do_date',
        'grn number': 'grn_number', 'grn no': 'grn_number',
        'warranty start': 'warranty_start', 'warranty end': 'warranty_end',
        'tagged date': 'tagged_date', 'short description': 'short_description',
        'date placed in service': 'date_placed_in_service',
        'insurance start date': 'insurance_start_date',
        'insurance end date': 'insurance_end_date',
        'maintenance start date': 'maintenance_start_date',
        'maintenance end date': 'maintenance_end_date',
        'next maintenance date': 'next_maintenance_date',
        'maintenance frequency days': 'maintenance_frequency_days',
        'maintenance frequency': 'maintenance_frequency_days',
        'expected units': 'expected_units',
        'useful life years': 'useful_life_years', 'useful life': 'useful_life_years',
        'salvage value': 'salvage_value', 'depreciation method': 'depreciation_method',
    }

    def _normalize_header(self, header):
        """Normalize a single header: lowercase, strip, replace spaces with underscores, apply aliases."""
        if not header:
            return ''
        h = str(header).strip().lower()
        # Check alias mapping first (before replacing spaces with underscores)
        if h in self.HEADER_ALIASES:
            return self.HEADER_ALIASES[h]
        # Replace spaces with underscores for standard field matching
        h = h.replace(' ', '_')
        if h in self.HEADER_ALIASES:
            return self.HEADER_ALIASES[h]
        return h

    def _normalize_row_keys(self, row):
        """Normalize all keys in a row dict."""
        return {self._normalize_header(k): v for k, v in row.items()}

    def parse_date(self, value):
        if not value:
            return None
        if isinstance(value, (datetime, date)):
            return value
        val_str = str(value).strip()
        if not val_str or val_str.lower() in ('none', 'null', 'nan'):
            return None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S',
                    '%d/%m/%Y %I:%M:%S %p', '%m/%d/%Y %I:%M:%S %p',
                    '%d/%m/%Y %H:%M:%S', '%m/%d/%Y %H:%M:%S',
                    '%Y-%m-%d %I:%M:%S %p', '%d-%m-%Y', '%d-%m-%Y %H:%M:%S',
                    '%d-%m-%Y %I:%M:%S %p'):
            try:
                return datetime.strptime(val_str, fmt).date()
            except (ValueError, TypeError):
                continue
        return None

    def _build_cache(self, model, field='name', org=None):
        qs = model.objects.all()
        if org and hasattr(model, 'organization'):
            qs = qs.filter(organization=org)
        cache = {}
        for obj in qs:
            val = getattr(obj, field)
            if val:
                cache[str(val).strip().lower()] = obj
        return cache

    def _detect_new_entities(self, rows, org):
        """Scan rows and return sets of entity names that don't exist in the system."""
        categories_by_code = self._build_cache(Category, 'code', org)
        categories_by_name = self._build_cache(Category, 'name', org)
        subcategories = self._build_cache(SubCategory, 'name', org)
        groups = self._build_cache(Group, 'name', org)
        subgroups = self._build_cache(SubGroup, 'name', org)
        brands = self._build_cache(Brand, 'name', org)
        regions = self._build_cache(Region, 'name', org)
        sites = self._build_cache(Site, 'name', org)
        buildings = self._build_cache(Building, 'name', org)
        floors = self._build_cache(Floor, 'name', org)

        new_entities = {
            'categories': set(),
            'subcategories': set(),
            'groups': set(),
            'sub_groups': set(),
            'brands': set(),
            'regions': set(),
            'sites': set(),
            'buildings': set(),
            'floors': set(),
        }

        for row in rows:
            cat_val = str(row.get('category') or '').strip()
            if cat_val and not (categories_by_code.get(cat_val.lower()) or categories_by_name.get(cat_val.lower())):
                new_entities['categories'].add(cat_val)

            sub_val = str(row.get('sub_category') or '').strip()
            if sub_val and not subcategories.get(sub_val.lower()):
                new_entities['subcategories'].add(sub_val)

            grp_val = str(row.get('group') or '').strip()
            if grp_val and not groups.get(grp_val.lower()):
                new_entities['groups'].add(grp_val)

            sgrp_val = str(row.get('sub_group') or '').strip()
            if sgrp_val and not subgroups.get(sgrp_val.lower()):
                new_entities['sub_groups'].add(sgrp_val)

            brand_val = str(row.get('brand') or '').strip()
            if brand_val and not brands.get(brand_val.lower()):
                new_entities['brands'].add(brand_val)

            region_val = str(row.get('region') or '').strip()
            if region_val and not regions.get(region_val.lower()):
                new_entities['regions'].add(region_val)

            site_val = str(row.get('site') or '').strip()
            if site_val and not sites.get(site_val.lower()):
                new_entities['sites'].add(site_val)

            building_val = str(row.get('building') or '').strip()
            if building_val and not buildings.get(building_val.lower()):
                new_entities['buildings'].add(building_val)

            floor_val = str(row.get('floor') or '').strip()
            if floor_val and not floors.get(floor_val.lower()):
                new_entities['floors'].add(floor_val)

        # Filter out empty sets
        return {k: sorted(v) for k, v in new_entities.items() if v}

    def form_valid(self, form):
        import_file = form.cleaned_data['import_file']
        org = self.request.user.organization
        
        try:
            rows = self.get_file_data(import_file)
        except ValueError as e:
            messages.error(self.request, str(e))
            return self.form_invalid(form)

        if not rows:
            messages.warning(self.request, "No data found in the file.")
            return self.form_invalid(form)

        # --- DETECT NEW ENTITIES ---
        new_entities = self._detect_new_entities(rows, org)

        if new_entities:
            # Store file data in session for the confirmation step
            serializable_rows = []
            for row in rows:
                serializable_row = {}
                for k, v in row.items():
                    if isinstance(v, (date, datetime)):
                        serializable_row[k] = v.isoformat()
                    elif v is None:
                        serializable_row[k] = ''
                    else:
                        serializable_row[k] = str(v)
                serializable_rows.append(serializable_row)

            self.request.session['import_rows'] = serializable_rows
            self.request.session['import_new_entities'] = new_entities

            return render(self.request, 'assets/asset_import_preview.html', {
                'new_entities': new_entities,
                'total_rows': len(rows),
            })

        # No new entities, proceed with import directly
        return self._process_import(rows, org)

    def _handle_confirm_import(self, request):
        """Handle the confirm step: auto-create entities, then import."""
        org = request.user.organization
        stored_rows = request.session.pop('import_rows', None)
        stored_entities = request.session.pop('import_new_entities', None)

        if not stored_rows:
            messages.error(request, "Import session expired. Please upload the file again.")
            return redirect('asset-import')

        rows = stored_rows
        new_entities = stored_entities or {}

        # Auto-create selected new entities
        selected = request.POST.getlist('create_entities')

        # Groups must be created before SubGroups (dependency order)
        if 'groups' in selected:
            for name in new_entities.get('groups', []):
                Group.objects.get_or_create(organization=org, name=name)

        if 'sub_groups' in selected:
            for name in new_entities.get('sub_groups', []):
                # Find the parent group from the first matching row
                parent_group = None
                for row in rows:
                    if str(row.get('sub_group') or '').strip().lower() == name.lower():
                        grp_val = str(row.get('group') or '').strip()
                        if grp_val:
                            parent_group = Group.objects.filter(
                                organization=org, name__iexact=grp_val
                            ).first()
                        break
                obj, created = SubGroup.objects.get_or_create(organization=org, name=name)
                if created and parent_group:
                    obj.group = parent_group
                    obj.save()

        # Categories must be created before SubCategories (dependency order)
        if 'categories' in selected:
            for name in new_entities.get('categories', []):
                # Find the sub_group from the first matching row
                parent_sub_group = None
                for row in rows:
                    if str(row.get('category') or '').strip().lower() == name.lower():
                        sgrp_val = str(row.get('sub_group') or '').strip()
                        if sgrp_val:
                            parent_sub_group = SubGroup.objects.filter(
                                organization=org, name__iexact=sgrp_val
                            ).first()
                        break
                obj, created = Category.objects.get_or_create(organization=org, name=name)
                if created and parent_sub_group:
                    obj.sub_group = parent_sub_group
                    obj.save()

        if 'subcategories' in selected:
            for name in new_entities.get('subcategories', []):
                parent_cat = None
                for row in rows:
                    if str(row.get('sub_category') or '').strip().lower() == name.lower():
                        cat_val = str(row.get('category') or '').strip()
                        if cat_val:
                            parent_cat = Category.objects.filter(
                                organization=org
                            ).filter(
                                models.Q(name__iexact=cat_val) | models.Q(code__iexact=cat_val)
                            ).first()
                        break
                if parent_cat:
                    SubCategory.objects.get_or_create(
                        organization=org, category=parent_cat, name=name
                    )

        if 'brands' in selected:
            for name in new_entities.get('brands', []):
                Brand.objects.get_or_create(organization=org, name=name)

        if 'regions' in selected:
            for name in new_entities.get('regions', []):
                Region.objects.get_or_create(organization=org, name=name)

        if 'sites' in selected:
            for name in new_entities.get('sites', []):
                # Link site to its region from the file row
                parent_region = None
                for row in rows:
                    if str(row.get('site') or '').strip().lower() == name.lower():
                        region_val = str(row.get('region') or '').strip()
                        if region_val:
                            parent_region = Region.objects.filter(
                                organization=org, name__iexact=region_val
                            ).first()
                        break
                if parent_region:
                    Site.objects.get_or_create(
                        region=parent_region, name=name,
                        defaults={'organization': org}
                    )

        if 'buildings' in selected:
            for name in new_entities.get('buildings', []):
                # Link building to its branch from the file row
                parent_branch = None
                for row in rows:
                    if str(row.get('building') or '').strip().lower() == name.lower():
                        branch_val = str(row.get('branch') or '').strip()
                        if branch_val:
                            parent_branch = Branch.objects.filter(
                                organization=org, name__iexact=branch_val
                            ).first()
                        break
                if parent_branch:
                    Building.objects.get_or_create(
                        branch=parent_branch, name=name,
                        defaults={'organization': org}
                    )

        if 'floors' in selected:
            for name in new_entities.get('floors', []):
                # Link floor to its building from the file row
                parent_building = None
                for row in rows:
                    if str(row.get('floor') or '').strip().lower() == name.lower():
                        building_val = str(row.get('building') or '').strip()
                        if building_val:
                            parent_building = Building.objects.filter(
                                organization=org, name__iexact=building_val
                            ).first()
                        break
                if parent_building:
                    Floor.objects.get_or_create(
                        building=parent_building, name=name,
                        defaults={'organization': org}
                    )

        created_types = [t.replace('_', ' ').title() for t in selected]
        if created_types:
            messages.info(request, f"Auto-created: {', '.join(created_types)}")

        return self._process_import(rows, org)

    def _process_import(self, rows, org):
        def build_cache(model, field='name', org_relevant=True):
            qs = model.objects.all()
            if org_relevant and hasattr(model, 'organization'):
                qs = qs.filter(organization=org)
            elif org_relevant and hasattr(model, 'category') and hasattr(model.category, 'organization'):
                qs = qs.filter(category__organization=org)
            
            cache = {}
            for obj in qs:
                val = getattr(obj, field)
                if val:
                    cache[str(val).strip().lower()] = obj
            return cache

        # Asset-specific master data
        categories_by_code = build_cache(Category, 'code')
        categories_by_name = build_cache(Category, 'name')
        subcategories = build_cache(SubCategory, 'name')
        groups = build_cache(Group, 'name')
        subgroups = build_cache(SubGroup, 'name')
        brands = build_cache(Brand, 'name')
        companies = build_cache(Company, 'name')
        suppliers = build_cache(Supplier, 'name')
        vendors = build_cache(Vendor, 'name')
        remarks = build_cache(AssetRemarks, 'remark')
        
        # Custodians (Lookup by Employee ID or Username)
        custodians_by_eid = build_cache(Custodian, 'employee_id')
        # Custodian by username requires related lookup
        custodians_by_user = {
            str(c.user.username).lower(): c 
            for c in Custodian.objects.filter(organization=org).select_related('user') 
            if c.user
        }

        # Location master data
        branches = build_cache(Branch, 'name')
        buildings = build_cache(Building, 'name')
        floors = build_cache(Floor, 'name')
        rooms = build_cache(Room, 'name')
        regions = build_cache(Region, 'name')
        sites = build_cache(Site, 'name')
        locations = build_cache(Location, 'name')
        sub_locations = build_cache(SubLocation, 'name')
        departments = build_cache(Department, 'name')

        # Helper to avoid repetitive dictionary lookups
        def get_from_cache(cache, val):
            if val is None: return None
            return cache.get(str(val).strip().lower())

        errors = []
        assets_to_create = []

        # --- STRUCTURED ASSET TAG GENERATION (matches the rest of the app) ---
        # Use the same configuration the org uses for manual asset creation
        # (apps.assets.models.generate_asset_tag), but maintain an in-memory
        # counter per-prefix so rows in the same import don't collide while
        # they're queued for bulk_create (and not yet visible in the DB).
        from datetime import date as _date
        from .models import generate_asset_tag  # noqa: F401  (kept for parity)

        _sep = getattr(org, 'tag_separator', '-') or '-'
        _include_company = getattr(org, 'tag_include_company', True)
        _include_category = getattr(org, 'tag_include_category', True)
        _include_year = getattr(org, 'tag_include_year', True)
        _seq_format = getattr(org, 'tag_sequence_format', 'HEX4') or 'HEX4'
        _fixed_prefix = (getattr(org, 'tag_prefix', '') or '').strip().upper()
        _year_suffix = str(_date.today().year)[-2:] if _include_year else ''

        _fmt_map = {
            'HEX4': lambda n: f"{n:04X}",
            'HEX6': lambda n: f"{n:06X}",
            'NUM4': lambda n: f"{n:04d}",
            'NUM5': lambda n: f"{n:05d}",
            'NUM6': lambda n: f"{n:06d}",
        }
        _fmt_seq = _fmt_map.get(_seq_format, _fmt_map['HEX4'])

        # Cache: prefix -> next available number
        _next_num_cache = {}

        def _build_prefix_parts(_company, _category):
            parts = []
            if _fixed_prefix:
                parts.append(_fixed_prefix)
            elif _include_company:
                # Derive 2-letter code from the ORGANIZATION name (was company name).
                org_name = getattr(org, 'name', '') or ''
                alpha = ''.join(c for c in org_name if c.isalpha()).upper()
                if not alpha and _company and getattr(_company, 'name', ''):
                    alpha = ''.join(c for c in _company.name if c.isalpha()).upper()
                if alpha:
                    code = alpha[:2] if len(alpha) >= 2 else alpha.ljust(2, 'X')[:2]
                else:
                    code = 'XX'
                parts.append(code)
            if _include_category:
                cat_code = (_category.code[:3].upper() if _category and _category.code else 'XXX')
                parts.append(cat_code)
            return parts

        def _seed_counter(prefix, prefix_parts):
            """Return the highest existing sequence number for this prefix."""
            qs = Asset.objects.filter(organization=org, asset_tag__startswith=prefix)
            if _include_year:
                qs = qs.filter(asset_tag__endswith=f"{_sep}{_year_suffix}")
            total_parts = len(prefix_parts) + 1 + (1 if _include_year else 0)
            seq_index = len(prefix_parts)
            max_num = 0
            for tag in qs.values_list('asset_tag', flat=True):
                try:
                    parts = tag.split(_sep)
                    if len(parts) != total_parts:
                        continue
                    counter_str = parts[seq_index]
                    num = int(counter_str, 16) if _seq_format.startswith('HEX') else int(counter_str)
                    if num > max_num:
                        max_num = num
                except (ValueError, IndexError):
                    continue
            return max_num

        def generate_tag_for_row(_company, _category):
            prefix_parts = _build_prefix_parts(_company, _category)
            prefix = _sep.join(prefix_parts)
            if prefix not in _next_num_cache:
                _next_num_cache[prefix] = _seed_counter(prefix, prefix_parts) + 1
            num = _next_num_cache[prefix]
            _next_num_cache[prefix] = num + 1
            all_parts = prefix_parts + [_fmt_seq(num)]
            if _include_year:
                all_parts.append(_year_suffix)
            return _sep.join(all_parts)

        for row_idx, row in enumerate(rows, start=2):
            try:
                # 1. Identification
                name = str(row.get('name') or '').strip()
                if not name:
                    raise ValueError("Asset Name is required.")

                cat_val = str(row.get('category') or '').strip()
                category = get_from_cache(categories_by_code, cat_val) or get_from_cache(categories_by_name, cat_val)
                if not category:
                    raise ValueError(f"Category '{cat_val}' not found.")

                # If the row provides a sub_group and the category doesn't have one yet, link it
                _sgrp_val = str(row.get('sub_group') or '').strip()
                if _sgrp_val and not category.sub_group_id:
                    _linked_sg = get_from_cache(subgroups, _sgrp_val)
                    if _linked_sg:
                        category.sub_group = _linked_sg
                        category.save(update_fields=['sub_group'])

                # Look up company early so we can use it in the tag
                _company_val = row.get('company')
                _row_company = get_from_cache(companies, _company_val)

                asset_tag = str(row.get('asset_tag') or '').strip()
                if not asset_tag:
                    asset_tag = generate_tag_for_row(_row_company, category)

                # 2. Enums / Choices
                def get_choice(val, choices_model, default):
                    if val is None: return default
                    v = str(val).strip().upper()
                    if v in choices_model.values: return v
                    # Search by display name if possible? 
                    # For now just upper-case exact match with choices
                    return default

                status = get_choice(row.get('status'), Asset.Status, Asset.Status.ACTIVE)
                condition = get_choice(row.get('condition'), Asset.Condition, Asset.Condition.NEW)
                asset_type = get_choice(row.get('asset_type'), Asset.Type, Asset.Type.TAGGABLE)
                label_type = get_choice(row.get('label_type'), Asset.LabelType, Asset.LabelType.BARCODE)

                # 3. Master Data Lookups (from cache)
                sub_category = get_from_cache(subcategories, row.get('sub_category'))
                group = get_from_cache(groups, row.get('group'))
                sub_group = get_from_cache(subgroups, row.get('sub_group'))
                brand_new = get_from_cache(brands, row.get('brand'))
                company = get_from_cache(companies, row.get('company'))
                supplier = get_from_cache(suppliers, row.get('supplier'))
                vendor = get_from_cache(vendors, row.get('vendor'))
                
                # Custodian lookup (EID then Username)
                cust_val = row.get('custodian')
                custodian = get_from_cache(custodians_by_eid, cust_val) or get_from_cache(custodians_by_user, cust_val)
                
                asset_remarks = get_from_cache(remarks, row.get('remarks'))

                # 4. Location Details (from cache)
                branch = get_from_cache(branches, row.get('branch'))
                department = get_from_cache(departments, row.get('department'))
                building = get_from_cache(buildings, row.get('building'))
                floor = get_from_cache(floors, row.get('floor'))
                room = get_from_cache(rooms, row.get('room'))
                
                # Special case: Auto-create room if missing but floor exists
                if not room and floor and row.get('room'):
                    room_name = str(row.get('room')).strip()
                    room = Room.objects.create(organization=org, floor=floor, name=room_name)
                    # Add to cache to prevent duplicate creation in this session
                    rooms[room_name.lower()] = room

                region = get_from_cache(regions, row.get('region'))
                site = get_from_cache(sites, row.get('site'))
                location = get_from_cache(locations, row.get('location'))
                sub_location = get_from_cache(sub_locations, row.get('sub_location'))

                # 5. Numerical Parsing
                def parse_decimal(val):
                    """Parse a numeric value robustly from CSV/Excel imports.

                    Accepts values like "5,000.00", "AED 5,000.00", "$5,000", "(5,000.00)" and returns Decimal or None.
                    """
                    if val is None:
                        return None
                    s = str(val).strip()
                    if s == '':
                        return None

                    # Handle parentheses for negative values: (1,234.56)
                    negative = False
                    if s.startswith('(') and s.endswith(')'):
                        negative = True
                        s = s[1:-1].strip()

                    # Remove any currency letters/symbols but keep digits, dot, comma and minus
                    try:
                        import re
                        s = re.sub(r"[^0-9.,\-]", "", s)
                    except Exception:
                        # Fallback: remove common non-numeric chars
                        s = s.replace('AED', '').replace('$', '').replace('USD', '')

                    # Normalize commas and dots (remove thousands separators)
                    s = s.replace(',', '')

                    if s in ('', '-', '.'):
                        return None

                    try:
                        d = Decimal(s)
                        return -d if negative else d
                    except Exception:
                        return None
                
                def parse_int(val, default=None):
                    if val is None or str(val).strip() == '': return default
                    try: return int(float(str(val)))
                    except: return default

                # Create Asset Instance (in memory)
                asset = Asset(
                    organization=org,
                    created_by=self.request.user,
                    name=name,
                    description=str(row.get('description') or ''),
                    short_description=str(row.get('short_description') or ''),
                    asset_tag=asset_tag,
                    asset_code=row.get('asset_code'),
                    erp_asset_number=row.get('erp_asset_number'),
                    quantity=parse_int(row.get('quantity'), 1),
                    label_type=label_type,
                    serial_number=row.get('serial_number'),
                    category=category,
                    sub_category=sub_category,
                    asset_type=asset_type,
                    status=status,
                    condition=condition,
                    group=group,
                    sub_group=sub_group,
                    brand_new=brand_new,
                    brand=str(row.get('brand') or '')[:100],
                    model=str(row.get('model') or '')[:100],
                    department=department,
                    cost_center=str(row.get('cost_center') or '')[:100],
                    company=company,
                    supplier=supplier,
                    vendor=vendor,
                    custodian=custodian,
                    employee_number=str(row.get('employee_number') or '')[:100],
                    branch=branch,
                    building=building,
                    floor=floor,
                    room=room,
                    region=region,
                    site=site,
                    location=location,
                    sub_location=sub_location,
                    purchase_date=self.parse_date(row.get('purchase_date')),
                    purchase_price=parse_decimal(row.get('purchase_price')),
                    currency=str(row.get('currency') or 'AED')[:10],
                    invoice_number=str(row.get('invoice_number') or '')[:100],
                    invoice_date=self.parse_date(row.get('invoice_date')),
                    po_number=str(row.get('po_number') or '')[:100],
                    po_date=self.parse_date(row.get('po_date')),
                    do_number=str(row.get('do_number') or '')[:100],
                    do_date=self.parse_date(row.get('do_date')),
                    grn_number=str(row.get('grn_number') or '')[:100],
                    warranty_start=self.parse_date(row.get('warranty_start')),
                    warranty_end=self.parse_date(row.get('warranty_end')),
                    tagged_date=self.parse_date(row.get('tagged_date')),
                    date_placed_in_service=self.parse_date(row.get('date_placed_in_service')),
                    insurance_start_date=self.parse_date(row.get('insurance_start_date')),
                    insurance_end_date=self.parse_date(row.get('insurance_end_date')),
                    maintenance_start_date=self.parse_date(row.get('maintenance_start_date')),
                    maintenance_end_date=self.parse_date(row.get('maintenance_end_date')),
                    next_maintenance_date=self.parse_date(row.get('next_maintenance_date')),
                    maintenance_frequency_days=parse_int(row.get('maintenance_frequency_days'), 0),
                    expected_units=parse_int(row.get('expected_units')),
                    useful_life_years=parse_int(row.get('useful_life_years')),
                    salvage_value=parse_decimal(row.get('salvage_value')),
                    depreciation_method=str(row.get('depreciation_method') or '').upper() or None,
                    asset_remarks=asset_remarks,
                    notes=str(row.get('notes') or '')
                )

                # Performance: Manually apply inheritance from category (since bulk_create bypasses save())
                if asset.category:
                    if asset.useful_life_years in (None, 0):
                        asset.useful_life_years = asset.category.useful_life_years
                    if not asset.depreciation_method:
                        asset.depreciation_method = asset.category.depreciation_method
                    if asset.salvage_value is None or asset.salvage_value == 0:
                        asset.salvage_value = asset.category.default_salvage_value
                    if asset.depreciation_method == 'UNITS_OF_PRODUCTION' and not asset.expected_units:
                        asset.expected_units = asset.category.default_expected_units

                assets_to_create.append(asset)

            except Exception as e:
                errors.append(f"Row {row_idx}: {str(e)}")

        if errors:
            # Display limited errors to avoid overwhelming the message system
            for err in errors[:15]:
                messages.error(self.request, err)
            if len(errors) > 15:
                messages.error(self.request, f"...and {len(errors) - 15} more errors.")
            return redirect('asset-import')

        # --- BULK SAVE ---
        try:
            with transaction.atomic():
                # Process in batches of 1000 for stability
                Asset.objects.bulk_create(assets_to_create, batch_size=1000)
                messages.success(self.request, f"Successfully imported {len(assets_to_create)} assets.")
        except Exception as e:
            messages.error(self.request, f"Database error during bulk save: {str(e)}")
            return redirect('asset-import')

        return redirect(self.success_url)

class AssetDetailView(LoginRequiredMixin, DetailView):
    model = Asset
    template_name = 'assets/asset_detail.html'
    context_object_name = 'asset'

    def get_queryset(self):
        return Asset.objects.filter(
            organization=self.request.user.organization,
            is_deleted=False,
        ).select_related(
            'category', 'sub_category', 'group', 'sub_group',
            'brand_new', 'department', 'assigned_to', 'company',
            'supplier', 'custodian', 'branch', 'building', 'floor',
            'room', 'region', 'site', 'location', 'sub_location',
            'vendor', 'asset_remarks'
        )
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from datetime import date
        from django.core.files.storage import default_storage
        from .code_generators import generate_codes_for_asset, AssetCodeGenerator

        asset = self.object

        def _file_missing(field):
            if not field:
                return True
            try:
                return not default_storage.exists(field.name)
            except Exception:
                return True

        # Auto-generate / regenerate missing codes so they always show
        if asset.asset_tag and (
            _file_missing(asset.barcode_image)
            or _file_missing(asset.qr_code_image)
            or _file_missing(asset.label_image)
        ):
            try:
                if _file_missing(asset.barcode_image):
                    asset.barcode_image = None
                if _file_missing(asset.qr_code_image):
                    asset.qr_code_image = None
                if _file_missing(asset.label_image):
                    asset.label_image = None
                generate_codes_for_asset(asset)
                asset.refresh_from_db(fields=['barcode_image', 'qr_code_image', 'label_image'])
                context['asset'] = asset
            except Exception:
                pass

        context['today'] = date.today()
        context['attachments'] = asset.attachments.all().order_by('-created_at')
        context['attachment_types'] = AssetAttachment.Type.choices
        return context


class AssetDeleteView(LoginRequiredMixin, View):
    """Soft-delete a single asset within the current tenant."""

    def post(self, request, pk, *args, **kwargs):
        org = getattr(request.user, 'organization', None)
        if not org:
            messages.error(request, 'You are not assigned to an organization.')
            return redirect('asset-list')

        asset = get_object_or_404(Asset, pk=pk, organization=org, is_deleted=False)
        asset.is_deleted = True
        asset.save(update_fields=['is_deleted'])
        invalidate_dashboard_cache_for_org(org)

        messages.success(request, f'Asset {asset.asset_tag} deleted successfully.')
        return redirect('asset-list')

class AssetUpdateView(LoginRequiredMixin, UpdateView):
    model = Asset
    form_class = AssetForm
    template_name = 'assets/asset_form.html'
    success_url = reverse_lazy('asset-list')

    def get_queryset(self):
        return Asset.objects.filter(organization=self.request.user.organization, is_deleted=False)


    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def form_valid(self, form):
        # Auto-generate Asset Tag if empty
        if not form.instance.asset_tag:
            form.instance.asset_tag = generate_asset_tag(
                self.request.user.organization,
                form.instance.category,
                form.instance.company
            )
        response = super().form_valid(form)
        # Post-save bulletproof safety net:
        # After the form saves, verify the saved site belongs to the saved region.
        # If not (JS didn't clear it or form clean() had an issue), force-clear all
        # child location fields directly in the DB so they are never left stale.
        saved = self.object
        if saved.site_id:
            try:
                _site = Site.objects.get(pk=saved.site_id)
                if _site.region_id != saved.region_id:
                    Asset.objects.filter(pk=saved.pk).update(
                        site_id=None, location_id=None, sub_location_id=None,
                        building_id=None, floor_id=None, room_id=None
                    )
            except Site.DoesNotExist:
                pass
        invalidate_dashboard_cache_for_org(self.request.user.organization)
        return response

# --- CATEGORY VIEWS ---
class CategoryListView(LoginRequiredMixin, ListView):
    model = Category
    template_name = 'assets/configuration/category_list.html'
    context_object_name = 'categories'

    def get_queryset(self):
        return Category.objects.filter(organization=self.request.user.organization)

class CategoryCreateView(LoginRequiredMixin, CreateView):
    model = Category
    form_class = CategoryForm
    template_name = 'assets/configuration/category_form.html'
    success_url = reverse_lazy('category-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.organization = self.request.user.organization
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This category already exists in your organization.')
            return self.form_invalid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['groups'] = Group.objects.filter(organization=self.request.user.organization)
        return ctx

class CategoryUpdateView(LoginRequiredMixin, UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = 'assets/configuration/category_form.html'
    success_url = reverse_lazy('category-list')
    
    def get_queryset(self):
        return Category.objects.filter(organization=self.request.user.organization)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['groups'] = Group.objects.filter(organization=self.request.user.organization)
        return ctx

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This category already exists in your organization.')
            return self.form_invalid(form)

# --- SUBCATEGORY VIEWS ---
class SubCategoryListView(LoginRequiredMixin, ListView):
    model = SubCategory
    template_name = 'assets/configuration/subcategory_list.html'
    context_object_name = 'subcategories'

    def get_queryset(self):
        return SubCategory.objects.filter(category__organization=self.request.user.organization).select_related('category')

class SubCategoryCreateView(LoginRequiredMixin, CreateView):
    model = SubCategory
    form_class = SubCategoryForm
    template_name = 'assets/configuration/subcategory_form.html'
    success_url = reverse_lazy('subcategory-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['groups'] = Group.objects.filter(organization=self.request.user.organization)
        return ctx

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This sub-category already exists in the selected category.')
            return self.form_invalid(form)

class SubCategoryUpdateView(LoginRequiredMixin, UpdateView):
    model = SubCategory
    form_class = SubCategoryForm
    template_name = 'assets/configuration/subcategory_form.html'
    success_url = reverse_lazy('subcategory-list')
    
    def get_queryset(self):
        return SubCategory.objects.filter(category__organization=self.request.user.organization)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['groups'] = Group.objects.filter(organization=self.request.user.organization)
        return ctx

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This sub-category already exists in the selected category.')
            return self.form_invalid(form)

# --- VENDOR VIEWS ---
class VendorListView(LoginRequiredMixin, ListView):
    model = Vendor
    template_name = 'assets/configuration/vendor_list.html'
    context_object_name = 'vendors'

    def get_queryset(self):
        return Vendor.objects.filter(organization=self.request.user.organization)

class VendorCreateView(LoginRequiredMixin, CreateView):
    model = Vendor
    form_class = VendorForm
    template_name = 'assets/configuration/vendor_form.html'
    success_url = reverse_lazy('vendor-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.organization = self.request.user.organization
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This vendor already exists in your organization.')
            return self.form_invalid(form)

class VendorUpdateView(LoginRequiredMixin, UpdateView):
    model = Vendor
    form_class = VendorForm
    template_name = 'assets/configuration/vendor_form.html'
    success_url = reverse_lazy('vendor-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This vendor already exists in your organization.')
            return self.form_invalid(form)
    
    def get_queryset(self):
        return Vendor.objects.filter(organization=self.request.user.organization)

# New Master Data CRUD Views

# --- GROUP VIEWS ---
class GroupListView(LoginRequiredMixin, ListView):
    model = Group
    template_name = 'assets/configuration/group_list.html'
    context_object_name = 'groups'

    def get_queryset(self):
        return Group.objects.filter(organization=self.request.user.organization)

class GroupCreateView(LoginRequiredMixin, CreateView):
    model = Group
    form_class = GroupForm
    template_name = 'assets/configuration/group_form.html'
    success_url = reverse_lazy('group-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.organization = self.request.user.organization
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This group already exists in your organization.')
            return self.form_invalid(form)

class GroupUpdateView(LoginRequiredMixin, UpdateView):
    model = Group
    form_class = GroupForm
    template_name = 'assets/configuration/group_form.html'
    success_url = reverse_lazy('group-list')
    
    def get_queryset(self):
        return Group.objects.filter(organization=self.request.user.organization)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This group already exists in your organization.')
            return self.form_invalid(form)

# --- SUBGROUP VIEWS ---
class SubGroupListView(LoginRequiredMixin, ListView):
    model = SubGroup
    template_name = 'assets/configuration/subgroup_list.html'
    context_object_name = 'subgroups'

    def get_queryset(self):
        return SubGroup.objects.filter(group__organization=self.request.user.organization).select_related('group')

class SubGroupCreateView(LoginRequiredMixin, CreateView):
    model = SubGroup
    form_class = SubGroupForm
    template_name = 'assets/configuration/subgroup_form.html'
    success_url = reverse_lazy('subgroup-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This sub-group already exists in the selected group.')
            return self.form_invalid(form)

class SubGroupUpdateView(LoginRequiredMixin, UpdateView):
    model = SubGroup
    form_class = SubGroupForm
    template_name = 'assets/configuration/subgroup_form.html'
    success_url = reverse_lazy('subgroup-list')
    
    def get_queryset(self):
        return SubGroup.objects.filter(group__organization=self.request.user.organization)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This sub-group already exists in the selected group.')
            return self.form_invalid(form)

# --- BRAND VIEWS ---
class BrandListView(LoginRequiredMixin, ListView):
    model = Brand
    template_name = 'assets/configuration/brand_list.html'
    context_object_name = 'brands'

    def get_queryset(self):
        return Brand.objects.filter(organization=self.request.user.organization)

class BrandCreateView(LoginRequiredMixin, CreateView):
    model = Brand
    form_class = BrandForm
    template_name = 'assets/configuration/brand_form.html'
    success_url = reverse_lazy('brand-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.organization = self.request.user.organization
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This brand already exists in your organization.')
            return self.form_invalid(form)

class BrandUpdateView(LoginRequiredMixin, UpdateView):
    model = Brand
    form_class = BrandForm
    template_name = 'assets/configuration/brand_form.html'
    success_url = reverse_lazy('brand-list')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_queryset(self):
        return Brand.objects.filter(organization=self.request.user.organization)

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This brand already exists in your organization.')
            return self.form_invalid(form)

# --- COMPANY VIEWS ---
class CompanyListView(LoginRequiredMixin, ListView):
    model = Company
    template_name = 'assets/configuration/company_list.html'
    context_object_name = 'companies'

    def get_queryset(self):
        return Company.objects.filter(organization=self.request.user.organization)

class CompanyCreateView(LoginRequiredMixin, CreateView):
    model = Company
    form_class = CompanyForm
    template_name = 'assets/configuration/company_form.html'
    success_url = reverse_lazy('company-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.organization = self.request.user.organization
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This company already exists in your organization.')
            return self.form_invalid(form)

class CompanyUpdateView(LoginRequiredMixin, UpdateView):
    model = Company
    form_class = CompanyForm
    template_name = 'assets/configuration/company_form.html'
    success_url = reverse_lazy('company-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This company already exists in your organization.')
            return self.form_invalid(form)
    
    def get_queryset(self):
        return Company.objects.filter(organization=self.request.user.organization)

# --- SUPPLIER VIEWS ---
class SupplierListView(LoginRequiredMixin, ListView):
    model = Supplier
    template_name = 'assets/configuration/supplier_list.html'
    context_object_name = 'suppliers'

    def get_queryset(self):
        return Supplier.objects.filter(organization=self.request.user.organization)

class SupplierCreateView(LoginRequiredMixin, CreateView):
    model = Supplier
    form_class = SupplierForm
    template_name = 'assets/configuration/supplier_form.html'
    success_url = reverse_lazy('supplier-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.organization = self.request.user.organization
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This supplier already exists in your organization.')
            return self.form_invalid(form)

class SupplierUpdateView(LoginRequiredMixin, UpdateView):
    model = Supplier
    form_class = SupplierForm
    template_name = 'assets/configuration/supplier_form.html'
    success_url = reverse_lazy('supplier-list')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def get_queryset(self):
        return Supplier.objects.filter(organization=self.request.user.organization)

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('name', 'This supplier already exists in your organization.')
            return self.form_invalid(form)

# --- CUSTODIAN VIEWS ---
class CustodianListView(LoginRequiredMixin, ListView):
    model = Custodian
    template_name = 'assets/configuration/custodian_list.html'
    context_object_name = 'custodians'

    def get_queryset(self):
        return Custodian.objects.filter(organization=self.request.user.organization).select_related('user')

class CustodianCreateView(LoginRequiredMixin, CreateView):
    model = Custodian
    form_class = CustodianForm
    template_name = 'assets/configuration/custodian_form.html'
    success_url = reverse_lazy('custodian-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.organization = self.request.user.organization
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('user', 'This user is already assigned as a custodian in your organization.')
            return self.form_invalid(form)

class CustodianUpdateView(LoginRequiredMixin, UpdateView):
    model = Custodian
    form_class = CustodianForm
    template_name = 'assets/configuration/custodian_form.html'
    success_url = reverse_lazy('custodian-list')
    
    def get_queryset(self):
        return Custodian.objects.filter(organization=self.request.user.organization)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('user', 'This user is already assigned as a custodian in your organization.')
            return self.form_invalid(form)

# --- ASSET REMARKS VIEWS ---
class AssetRemarksListView(LoginRequiredMixin, ListView):
    model = AssetRemarks
    template_name = 'assets/configuration/assetremarks_list.html'
    context_object_name = 'remarks'

    def get_queryset(self):
        return AssetRemarks.objects.filter(organization=self.request.user.organization)

class AssetRemarksCreateView(LoginRequiredMixin, CreateView):
    model = AssetRemarks
    form_class = AssetRemarksForm
    template_name = 'assets/configuration/assetremarks_form.html'
    success_url = reverse_lazy('assetremarks-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        form.instance.organization = self.request.user.organization
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('remark', 'This remark already exists in your organization.')
            return self.form_invalid(form)

class AssetRemarksUpdateView(LoginRequiredMixin, UpdateView):
    model = AssetRemarks
    form_class = AssetRemarksForm
    template_name = 'assets/configuration/assetremarks_form.html'
    success_url = reverse_lazy('assetremarks-list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def form_valid(self, form):
        try:
            return super().form_valid(form)
        except IntegrityError:
            form.add_error('remark', 'This remark already exists in your organization.')
            return self.form_invalid(form)
    
    def get_queryset(self):
        return AssetRemarks.objects.filter(organization=self.request.user.organization)

# AJAX endpoints for new cascading dropdowns
def get_subgroups(request):
    group_id = request.GET.get('group_id')
    if group_id:
        subgroups = SubGroup.objects.filter(group_id=group_id).values('id', 'name')
        return JsonResponse(list(subgroups), safe=False)
    return JsonResponse([], safe=False)

def get_categories_by_subgroup(request):
    sub_group_id = request.GET.get('sub_group_id')
    if sub_group_id:
        categories = Category.objects.filter(
            sub_group_id=sub_group_id,
            organization=request.user.organization
        ).values('id', 'name')
    else:
        categories = Category.objects.filter(
            organization=request.user.organization
        ).values('id', 'name')
    return JsonResponse(list(categories), safe=False)

# ==================== APPROVAL WORKFLOW VIEWS ====================

from .models import ApprovalRequest, ApprovalLog
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.shortcuts import HttpResponseRedirect
from apps.users.views import ApprovalAccessMixin, CheckerRequiredMixin, SeniorManagerRequiredMixin, EmployeeRequiredMixin

class ApprovalListView(ApprovalAccessMixin, TemplateView):
    """Unified approvals dashboard showing both asset approvals and disposal requests"""
    template_name = 'assets/approval_list.html'
    
    def get_disposal_queryset(self):
        """Get relevant disposal approvals based on user role"""
        user = self.request.user
        org = user.organization
        
        # Managers see PENDING disposals to review
        if user.role in [user.Role.SENIOR_MANAGER, user.Role.CHECKER]:
            return AssetDisposal.objects.filter(
                organization=org,
                status=AssetDisposal.Status.PENDING
            ).order_by('-created_at')
        
        # Admins see MANAGER_APPROVED disposals awaiting final approval
        elif user.is_superuser or user.role == user.Role.ADMIN:
            return AssetDisposal.objects.filter(
                organization=org,
                status=AssetDisposal.Status.MANAGER_APPROVED
            ).order_by('-created_at')
        
        # Employees see their own disposal requests
        elif user.role == user.Role.EMPLOYEE:
            return AssetDisposal.objects.filter(
                organization=org,
                requested_by=user
            ).order_by('-created_at')
        
        return AssetDisposal.objects.none()
    
    def get_asset_approval_queryset(self):
        """Get relevant asset approval requests based on user role"""
        user = self.request.user
        
        # Checkers see pending requests
        if user.is_checker:
            return ApprovalRequest.objects.filter(
                organization=user.organization,
                status=ApprovalRequest.Status.PENDING
            ).order_by('-created_at')
        
        # Senior managers see checker-approved requests
        elif user.is_senior_manager:
            return ApprovalRequest.objects.filter(
                organization=user.organization,
                status=ApprovalRequest.Status.CHECKER_APPROVED
            ).order_by('-created_at')
        
        # Employees can see only their own
        elif user.role == user.Role.EMPLOYEE:
            return ApprovalRequest.objects.filter(
                organization=user.organization,
                requester=user
            ).order_by('-created_at')
        
        # Admins see all
        elif user.is_superuser or user.role == user.Role.ADMIN:
            return ApprovalRequest.objects.filter(
                organization=user.organization
            ).order_by('-created_at')
        
        return ApprovalRequest.objects.none()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        # Get disposals
        disposals = self.get_disposal_queryset()
        
        # Get asset approvals
        asset_approvals = self.get_asset_approval_queryset()
        
        # Combine all for unified list (sort by created_at)
        all_approvals = []
        for disposal in disposals:
            all_approvals.append({
                'type': 'disposal',
                'id': disposal.id,
                'pk': disposal.pk,
                'asset': disposal.asset,
                'status': disposal.status,
                'get_status_display': disposal.get_status_display(),
                'created_at': disposal.created_at,
                'requested_by': disposal.requested_by,
                'get_disposal_method_display': disposal.get_disposal_method_display(),
                'disposal_method': disposal.disposal_method,
                'manager_approved_by': disposal.manager_approved_by,
                'can_approve': (user.role in [user.Role.SENIOR_MANAGER, user.Role.CHECKER] and disposal.status == AssetDisposal.Status.PENDING) or 
                               (user.is_superuser or user.role == user.Role.ADMIN) and disposal.status == AssetDisposal.Status.MANAGER_APPROVED
            })
        
        for approval in asset_approvals:
            all_approvals.append({
                'type': 'asset',
                'id': approval.id,
                'pk': approval.pk,
                'asset': approval.asset,
                'asset_name': approval.asset.name if approval.asset else 'N/A',
                'status': approval.status,
                'get_status_display': approval.get_status_display(),
                'created_at': approval.created_at,
                'requester': approval.requester,
                'request_type': approval.get_request_type_display(),
                'can_approve': approval.status in [ApprovalRequest.Status.PENDING, ApprovalRequest.Status.CHECKER_APPROVED]
            })
        
        # Sort by created_at
        all_approvals.sort(key=lambda x: x['created_at'], reverse=True)
        
        # Calculate stats
        disposal_pending = AssetDisposal.objects.filter(
            organization=user.organization,
            status=AssetDisposal.Status.PENDING
        ).count()
        
        asset_pending = ApprovalRequest.objects.filter(
            organization=user.organization,
            status=ApprovalRequest.Status.PENDING
        ).count()
        
        # Add to context
        context['disposals'] = disposals
        context['asset_approvals'] = asset_approvals
        context['all_approvals'] = all_approvals
        context['disposal_pending'] = disposals.filter(status=AssetDisposal.Status.PENDING).count()
        context['asset_pending'] = asset_approvals.filter(status=ApprovalRequest.Status.PENDING).count()
        context['total_pending'] = context['disposal_pending'] + context['asset_pending']
        context['total_approved'] = disposals.filter(status=AssetDisposal.Status.APPROVED).count() + \
                                   asset_approvals.filter(status=ApprovalRequest.Status.APPROVED).count()
        context['total_rejected'] = disposals.filter(status=AssetDisposal.Status.REJECTED).count() + \
                                   asset_approvals.filter(status=ApprovalRequest.Status.PENDING).count()
        
        return context


class ApprovalDetailView(ApprovalAccessMixin, DetailView):
    """View approval request details"""
    model = ApprovalRequest
    template_name = 'assets/approval_detail.html'
    context_object_name = 'approval'
    
    def get_queryset(self):
        user = self.request.user
        qs = ApprovalRequest.objects.filter(organization=user.organization)
        # Admins and approvers may view all requests
        if user.is_superuser or user.role == user.Role.ADMIN or user.is_checker or user.is_senior_manager:
            return qs
        # Employees may view only their own requests
        if user.role == user.Role.EMPLOYEE:
            return qs.filter(requester=user)
        return ApprovalRequest.objects.none()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        approval = self.object
        context['approval_logs'] = approval.approval_logs.all().order_by('-created_at')
        # Only admins (or superusers) can perform approval actions
        is_admin = self.request.user.is_superuser or self.request.user.role == self.request.user.Role.ADMIN
        context['can_approve_checker'] = is_admin and approval.needs_checker_approval
        context['can_approve_senior'] = is_admin and approval.needs_senior_approval
        return context


class ApprovalApproveView(LoginRequiredMixin, View):
    """Approve an approval request"""
    
    def post(self, request, pk):
        approval = get_object_or_404(ApprovalRequest, pk=pk, organization=request.user.organization)
        user = request.user
        
        decision = request.POST.get('decision')  # 'APPROVED' or 'REJECTED'
        comments = request.POST.get('comments', '')
        
        # Only admins (or superusers) can approve; they advance the request through the workflow
        if (user.is_superuser or user.role == user.Role.ADMIN) and approval.status == ApprovalRequest.Status.PENDING:
            if decision == 'APPROVED':
                approval.status = ApprovalRequest.Status.CHECKER_APPROVED
                messages.success(request, 'Request approved. Pending senior manager approval.')
            elif decision == 'REJECTED':
                approval.status = ApprovalRequest.Status.CHECKER_REJECTED
                messages.warning(request, 'Request rejected.')
            approval.save()
            ApprovalLog.objects.create(
                approval_request=approval,
                approver=user,
                decision=decision,
                approval_level='CHECKER',
                comments=comments,
                organization=user.organization
            )
        elif (user.is_superuser or user.role == user.Role.ADMIN) and approval.status == ApprovalRequest.Status.CHECKER_APPROVED:
            if decision == 'APPROVED':
                approval.status = ApprovalRequest.Status.APPROVED
                messages.success(request, 'Request fully approved!')
            elif decision == 'REJECTED':
                approval.status = ApprovalRequest.Status.SENIOR_REJECTED
                messages.warning(request, 'Request rejected.')
            approval.save()
            ApprovalLog.objects.create(
                approval_request=approval,
                approver=user,
                decision=decision,
                approval_level='SENIOR_MANAGER',
                comments=comments,
                organization=user.organization
            )
        else:
            messages.error(request, 'You do not have permission to approve this request.')
            return HttpResponseRedirect(reverse('approval_detail', args=[pk]))
        
        return HttpResponseRedirect(reverse('approval_detail', args=[pk]))


class CreateApprovalRequestView(LoginRequiredMixin, CreateView):
    """Create approval request for new asset (any authenticated user)."""
    model = ApprovalRequest
    fields = ['request_type', 'data', 'comments']
    template_name = 'assets/approval_request_form.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        form.instance.requester = self.request.user
        form.instance.organization = self.request.user.organization
        messages.success(self.request, 'Approval request submitted for review.')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('approval_list')


class ReportsListView(LoginRequiredMixin, View):
    """Display list of available reports"""
    template_name = 'assets/reports_list.html'
    
    def get(self, request):
        user = request.user
        # Define available reports
        reports = []
        
        # Only show Depreciation Report to non-EMPLOYEE users
        if user.role != user.Role.EMPLOYEE:
            reports.append({
                'name': 'Depreciation Report',
                'description': 'View asset depreciation analysis and financial depreciation data.',
                'icon': 'trending-down',
                'url': reverse('asset-list') + '?view=depreciation',
                'color': 'info'
            })
        
        # Asset Inventory - visible to all
        reports.append({
            'name': 'Asset Inventory',
            'description': 'Complete list of all assets in your inventory with details.',
            'icon': 'package',
            'url': reverse('asset-list'),
            'color': 'primary'
        })
        
        # Asset Approvals - only for users with approval permissions
        if user.can_approve:
            reports.append({
                'name': 'Asset Approvals',
                'description': 'View and manage pending and completed asset approval requests.',
                'icon': 'check-circle',
                'url': reverse('approval_list'),
                'color': 'success'
            })

        # Asset Reconciliation Report - non-EMPLOYEE users
        if user.role != user.Role.EMPLOYEE:
            reports.append({
                'name': 'Asset Reconciliation Report',
                'description': 'Full-picture summary of the entire asset register with opening balance, additions, disposals, and closing values.',
                'icon': 'clipboard-check',
                'url': reverse('reconciliation-report'),
                'color': 'warning'
            })

        from django.contrib.auth import get_user_model
        User = get_user_model()

        context = {
            'reports': reports,
            'total_assets': Asset.objects.filter(
                organization=request.user.organization,
                is_deleted=False
            ).count(),
        }
        
        return render(request, self.template_name, context)


class MastersListView(LoginRequiredMixin, View):
    """Display list of all assets with complete details - with pagination and filters"""
    template_name = 'assets/masters_list.html'
    paginate_by = 50
    
    def get(self, request):
        from django.core.paginator import Paginator
        from django.db.models import Q
        organization = request.user.organization
        
        # Get search and filter parameters
        search_query = request.GET.get('q', '').strip()
        category_filter = request.GET.get('category', '').strip()
        sub_category_filter = request.GET.get('sub_category', '').strip()
        company_filter = request.GET.get('company', '').strip()
        department_filter = request.GET.get('department', '').strip()
        condition_filter = request.GET.get('condition', '').strip()
        status_filter = request.GET.get('status', '').strip()
        brand_filter = request.GET.get('brand', '').strip()
        vendor_filter = request.GET.get('vendor', '').strip()
        supplier_filter = request.GET.get('supplier', '').strip()
        group_filter = request.GET.get('group', '').strip()
        subgroup_filter = request.GET.get('subgroup', '').strip()
        custodian_filter = request.GET.get('custodian', '').strip()
        purchase_date_from = request.GET.get('purchase_date_from', '').strip()
        purchase_date_to = request.GET.get('purchase_date_to', '').strip()
        sort_by = request.GET.get('sort_by', '-created_at').strip()
        sort_order = request.GET.get('sort_order', 'desc').strip()
        
        # Fetch assets with optimization
        assets_qs = Asset.objects.filter(
            organization=organization,
            is_deleted=False
        ).select_related(
            'category', 'sub_category', 'group', 'sub_group', 'brand_new',
            'company', 'supplier', 'custodian', 'department', 'assigned_to',
            'branch', 'building', 'floor', 'room', 'region', 'site', 
            'location', 'sub_location', 'vendor', 'asset_remarks', 'parent'
        )
        
        # Apply sorting
        if sort_by:
            if sort_order == 'asc':
                assets_qs = assets_qs.order_by(sort_by)
            else:
                assets_qs = assets_qs.order_by(f'-{sort_by}' if not sort_by.startswith('-') else sort_by)
        else:
            assets_qs = assets_qs.order_by('-created_at')
        
        # Apply search filter
        if search_query:
            assets_qs = assets_qs.filter(
                Q(asset_tag__icontains=search_query) |
                Q(name__icontains=search_query) |
                Q(serial_number__icontains=search_query) |
                Q(category__name__icontains=search_query)
            )
        
        # Apply category and sub-category filters
        if category_filter:
            assets_qs = assets_qs.filter(category_id=category_filter)
        if sub_category_filter:
            assets_qs = assets_qs.filter(sub_category_id=sub_category_filter)
        
        # Apply company filter
        if company_filter:
            assets_qs = assets_qs.filter(company_id=company_filter)
        
        # Apply department filter
        if department_filter:
            assets_qs = assets_qs.filter(department_id=department_filter)
        
        # Apply condition filter
        if condition_filter:
            assets_qs = assets_qs.filter(condition=condition_filter)
        
        # Apply status filter
        if status_filter:
            assets_qs = assets_qs.filter(status=status_filter)
            
        
        # Apply brand filter
        if brand_filter:
            assets_qs = assets_qs.filter(brand_new_id=brand_filter)
        
        # Apply vendor filter
        if vendor_filter:
            assets_qs = assets_qs.filter(vendor_id=vendor_filter)
        
        # Apply supplier filter
        if supplier_filter:
            assets_qs = assets_qs.filter(supplier_id=supplier_filter)
        
        # Apply group filter
        if group_filter:
            assets_qs = assets_qs.filter(group_id=group_filter)
        
        # Apply subgroup filter
        if subgroup_filter:
            assets_qs = assets_qs.filter(sub_group_id=subgroup_filter)
        
        # Apply custodian filter
        if custodian_filter:
            assets_qs = assets_qs.filter(custodian_id=custodian_filter)
        
        # Apply purchase date range filters
        if purchase_date_from:
            try:
                from datetime import datetime
                from_date = datetime.strptime(purchase_date_from, '%Y-%m-%d').date()
                assets_qs = assets_qs.filter(purchase_date__gte=from_date)
            except:
                pass
        
        if purchase_date_to:
            try:
                from datetime import datetime
                to_date = datetime.strptime(purchase_date_to, '%Y-%m-%d').date()
                assets_qs = assets_qs.filter(purchase_date__lte=to_date)
            except:
                pass
        
        # Paginate results
        paginator = Paginator(assets_qs, self.paginate_by)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
        # Get filter options
        categories = Category.objects.filter(organization=organization).order_by('name')
        sub_categories = SubCategory.objects.filter(organization=organization).order_by('name')
        companies = Company.objects.filter(organization=organization).order_by('name')
        departments = Department.objects.filter(organization=organization).order_by('name')
        brands = Brand.objects.filter(organization=organization).order_by('name')
        vendors = Vendor.objects.filter(organization=organization).order_by('name')
        suppliers = Supplier.objects.filter(organization=organization).order_by('name')
        groups = Group.objects.filter(organization=organization).order_by('name')
        subgroups = SubGroup.objects.filter(organization=organization).order_by('name')
        custodians = Custodian.objects.filter(organization=organization).order_by('employee_id')
        
        context = {
            'page_obj': page_obj,
            'assets': page_obj.object_list,
            'is_paginated': page_obj.has_other_pages(),
            'total_assets': paginator.count,
            'search_query': search_query,
            'category_filter': category_filter,
            'sub_category_filter': sub_category_filter,
            'company_filter': company_filter,
            'department_filter': department_filter,
            'condition_filter': condition_filter,
            'status_filter': status_filter,
            'brand_filter': brand_filter,
            'vendor_filter': vendor_filter,
            'supplier_filter': supplier_filter,
            'group_filter': group_filter,
            'subgroup_filter': subgroup_filter,
            'custodian_filter': custodian_filter,
            'purchase_date_from': purchase_date_from,
            'purchase_date_to': purchase_date_to,
            'sort_by': sort_by,
            'sort_order': sort_order,
            'categories': categories,
            'sub_categories': sub_categories,
            'companies': companies,
            'departments': departments,
            'brands': brands,
            'vendors': vendors,
            'suppliers': suppliers,
            'groups': groups,
            'subgroups': subgroups,
            'custodians': custodians,
            'condition_choices': Asset._meta.get_field('condition').choices,
            'status_choices': Asset._meta.get_field('status').choices,
        }
        
        return render(request, self.template_name, context)


class MastersExportExcelView(LoginRequiredMixin, View):
    """Export masters data to Excel with all fields and applied filters"""
    
    def get(self, request):
        from django.core.paginator import Paginator
        from django.db.models import Q
        organization = request.user.organization
        
        # Get same filters as MastersListView
        search_query = request.GET.get('q', '').strip()
        category_filter = request.GET.get('category', '').strip()
        sub_category_filter = request.GET.get('sub_category', '').strip()
        company_filter = request.GET.get('company', '').strip()
        department_filter = request.GET.get('department', '').strip()
        condition_filter = request.GET.get('condition', '').strip()
        status_filter = request.GET.get('status', '').strip()
        brand_filter = request.GET.get('brand', '').strip()
        vendor_filter = request.GET.get('vendor', '').strip()
        supplier_filter = request.GET.get('supplier', '').strip()
        group_filter = request.GET.get('group', '').strip()
        subgroup_filter = request.GET.get('subgroup', '').strip()
        custodian_filter = request.GET.get('custodian', '').strip()
        purchase_date_from = request.GET.get('purchase_date_from', '').strip()
        purchase_date_to = request.GET.get('purchase_date_to', '').strip()
        sort_by = request.GET.get('sort_by', '-created_at').strip()
        
        # Build queryset with same logic as MastersListView
        assets_qs = Asset.objects.filter(
            organization=organization
        ).select_related(
            'category', 'sub_category', 'group', 'sub_group', 'brand_new',
            'company', 'supplier', 'custodian', 'department', 'assigned_to',
            'branch', 'building', 'floor', 'room', 'region', 'site', 
            'location', 'sub_location', 'vendor', 'asset_remarks', 'parent'
        )
        
        # Apply filters
        if search_query:
            assets_qs = assets_qs.filter(
                Q(asset_tag__icontains=search_query) |
                Q(name__icontains=search_query) |
                Q(serial_number__icontains=search_query) |
                Q(category__name__icontains=search_query)
            )
        if category_filter:
            assets_qs = assets_qs.filter(category_id=category_filter)
        if sub_category_filter:
            assets_qs = assets_qs.filter(sub_category_id=sub_category_filter)
        if company_filter:
            assets_qs = assets_qs.filter(company_id=company_filter)
        if department_filter:
            assets_qs = assets_qs.filter(department_id=department_filter)
        if condition_filter:
            assets_qs = assets_qs.filter(condition=condition_filter)
        if status_filter:
            assets_qs = assets_qs.filter(status=status_filter)
        if brand_filter:
            assets_qs = assets_qs.filter(brand_new_id=brand_filter)
        if vendor_filter:
            assets_qs = assets_qs.filter(vendor_id=vendor_filter)
        if supplier_filter:
            assets_qs = assets_qs.filter(supplier_id=supplier_filter)
        if group_filter:
            assets_qs = assets_qs.filter(group_id=group_filter)
        if subgroup_filter:
            assets_qs = assets_qs.filter(sub_group_id=subgroup_filter)
        if custodian_filter:
            assets_qs = assets_qs.filter(custodian_id=custodian_filter)
        if purchase_date_from:
            try:
                from_date = datetime.strptime(purchase_date_from, '%Y-%m-%d').date()
                assets_qs = assets_qs.filter(purchase_date__gte=from_date)
            except:
                pass
        if purchase_date_to:
            try:
                to_date = datetime.strptime(purchase_date_to, '%Y-%m-%d').date()
                assets_qs = assets_qs.filter(purchase_date__lte=to_date)
            except:
                pass
        
        # Apply sorting
        if sort_by:
            assets_qs = assets_qs.order_by(sort_by)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asset Masters"
        
        # Add export information at the top
        current_row = 1
        ws.merge_cells(f'A{current_row}:E{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = "ASSET MASTER DATA EXPORT"
        cell.font = openpyxl.styles.Font(size=14, bold=True)
        cell.alignment = openpyxl.styles.Alignment(horizontal='left')
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Export Date:")
        ws.cell(row=current_row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True)
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Total Records:")
        ws.cell(row=current_row, column=2, value=assets_qs.count())
        ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True)
        current_row += 1
        
        # Add filter information if any filters are applied
        filters_applied = []
        if search_query:
            filters_applied.append(f"Search: {search_query}")
        if category_filter:
            try:
                cat = Category.objects.get(id=category_filter)
                filters_applied.append(f"Category: {cat.name}")
            except:
                pass
        if sub_category_filter:
            try:
                subcat = SubCategory.objects.get(id=sub_category_filter)
                filters_applied.append(f"Sub Category: {subcat.name}")
            except:
                pass
        if company_filter:
            try:
                comp = Company.objects.get(id=company_filter)
                filters_applied.append(f"Company: {comp.name}")
            except:
                pass
        if department_filter:
            try:
                dept = Department.objects.get(id=department_filter)
                filters_applied.append(f"Department: {dept.name}")
            except:
                pass
        if brand_filter:
            try:
                brand = Brand.objects.get(id=brand_filter)
                filters_applied.append(f"Brand: {brand.name}")
            except:
                pass
        if vendor_filter:
            try:
                vendor = Vendor.objects.get(id=vendor_filter)
                filters_applied.append(f"Vendor: {vendor.name}")
            except:
                pass
        if supplier_filter:
            try:
                supplier = Supplier.objects.get(id=supplier_filter)
                filters_applied.append(f"Supplier: {supplier.name}")
            except:
                pass
        if group_filter:
            try:
                group = Group.objects.get(id=group_filter)
                filters_applied.append(f"Group: {group.name}")
            except:
                pass
        if subgroup_filter:
            try:
                subgrp = SubGroup.objects.get(id=subgroup_filter)
                filters_applied.append(f"Sub Group: {subgrp.name}")
            except:
                pass
        if custodian_filter:
            try:
                custodian = Custodian.objects.get(id=custodian_filter)
                cust_name = custodian.user.get_full_name() if custodian.user else custodian.employee_id
                filters_applied.append(f"Custodian: {cust_name}")
            except:
                pass
        if condition_filter:
            cond_display = dict(Asset._meta.get_field('condition').choices).get(condition_filter, condition_filter)
            filters_applied.append(f"Condition: {cond_display}")
        if status_filter:
            status_display = dict(Asset._meta.get_field('status').choices).get(status_filter, status_filter)
            filters_applied.append(f"Status: {status_display}")
        if purchase_date_from:
            filters_applied.append(f"Purchase Date From: {purchase_date_from}")
        if purchase_date_to:
            filters_applied.append(f"Purchase Date To: {purchase_date_to}")
        
        if filters_applied:
            current_row += 1
            ws.cell(row=current_row, column=1, value="Applied Filters:")
            ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(bold=True)
            current_row += 1
            for filter_text in filters_applied:
                ws.cell(row=current_row, column=1, value=f"  â€¢ {filter_text}")
                current_row += 1
        else:
            current_row += 1
            ws.cell(row=current_row, column=1, value="Filters: None (All assets)")
            ws.cell(row=current_row, column=1).font = openpyxl.styles.Font(italic=True)
            current_row += 1
        
        current_row += 1  # Empty row before data
        
        # Define headers (matching the table columns)
        headers = [
            'Asset Tag', 'Name', 'Category', 'Sub Category', 'Asset Type', 'Group', 'Sub Group',
            'Brand', 'Model', 'Condition', 'Serial Number', 'Supplier', 'Vendor', 'Company',
            'Department', 'Custodian', 'Employee Number', 'Region', 'Branch', 'Building', 
            'Floor', 'Room', 'Location', 'Site', 'Purchase Date', 'Purchase Price', 'Currency',
            'Invoice Number', 'PO Number', 'GRN Number', 'Warranty Start', 'Warranty End',
            'Tagged Date', 'Maintenance Start', 'Maintenance Frequency (Days)', 'Next Maintenance',
            'Useful Life (Years)', 'Salvage Value', 'Depreciation Method', 'Status', 'Remarks'
        ]
        
        # Add headers at current_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        
        current_row += 1
        
        # Add data rows
        for asset in assets_qs:
            row = [
                asset.asset_tag or '',
                asset.name or '',
                asset.category.name if asset.category else '',
                asset.sub_category.name if asset.sub_category else '',
                asset.asset_type or '',
                asset.group.name if asset.group else '',
                asset.sub_group.name if asset.sub_group else '',
                asset.brand_new.name if asset.brand_new else '',
                asset.model or '',
                asset.get_condition_display() if asset.condition else '',
                asset.serial_number or '',
                asset.supplier.name if asset.supplier else '',
                asset.vendor.name if asset.vendor else '',
                asset.company.name if asset.company else '',
                asset.department.name if asset.department else '',
                asset.custodian.user.get_full_name() if asset.custodian and asset.custodian.user else (asset.custodian.employee_id if asset.custodian else ''),
                asset.employee_number or '',
                asset.region.name if asset.region else '',
                asset.branch.name if asset.branch else '',
                asset.building.name if asset.building else '',
                asset.floor.name if asset.floor else '',
                asset.room.name if asset.room else '',
                asset.location.name if asset.location else '',
                asset.site.name if asset.site else '',
                asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '',
                float(asset.purchase_price) if asset.purchase_price else '',
                asset.currency or '',
                asset.invoice_number or '',
                asset.po_number or '',
                asset.grn_number or '',
                asset.warranty_start.strftime('%Y-%m-%d') if asset.warranty_start else '',
                asset.warranty_end.strftime('%Y-%m-%d') if asset.warranty_end else '',
                asset.tagged_date.strftime('%Y-%m-%d') if asset.tagged_date else '',
                asset.maintenance_start_date.strftime('%Y-%m-%d') if asset.maintenance_start_date else '',
                asset.maintenance_frequency_days or '',
                asset.next_maintenance_date.strftime('%Y-%m-%d') if asset.next_maintenance_date else '',
                asset.useful_life_years or '',
                float(asset.salvage_value) if asset.salvage_value else '',
                asset.get_depreciation_method_display() if asset.depreciation_method else '',
                asset.get_status_display() if asset.status else '',
                asset.notes or (asset.asset_remarks.name if asset.asset_remarks else ''),
            ]
            
            # Add row at current position
            for col_num, value in enumerate(row, 1):
                ws.cell(row=current_row, column=col_num, value=value)
            current_row += 1
        
        # Auto-adjust column widths
        # Use explicit row/column indexing to avoid MergedCell issues from merged title cells.
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)

            for row_idx in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is None:
                    continue
                cell_length = len(str(cell_value))
                if cell_length > max_length:
                    max_length = cell_length

            adjusted_width = min(max_length + 2, 50) if max_length > 0 else 12
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="asset_masters_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response


# ========================
# Asset Transfer Views
# ========================

class AssetTransferListView(LoginRequiredMixin, ListView):
    """List all asset transfers with filtering"""
    model = AssetTransfer
    template_name = 'assets/transfer_list.html'
    context_object_name = 'transfers'
    paginate_by = 50

    _FROM_SNAPSHOT_FIELDS = [
        'transferred_from_user',
        'transferred_from_department',
        'transferred_from_location',
        'transferred_from_region',
        'transferred_from_site',
        'transferred_from_building',
        'transferred_from_floor',
        'transferred_from_room',
        'transferred_from_company',
        'transferred_from_custodian',
    ]

    def _hydrate_missing_from_snapshots(self, queryset):
        """Backfill legacy open transfers that were created without from-snapshot data."""
        transfers_to_fix = queryset.filter(
            status__in=[AssetTransfer.Status.PENDING, AssetTransfer.Status.IN_TRANSIT],
            transferred_from_location__isnull=True,
        )

        for transfer in transfers_to_fix:
            transfer.snapshot_from_asset()
            transfer.save(update_fields=self._FROM_SNAPSHOT_FIELDS + ['updated_at'])

    def get_filtered_queryset(self):
        org = self.request.user.organization
        queryset = AssetTransfer.objects.filter(organization=org).select_related(
            'asset',
            'transferred_from_user',
            'transferred_from_department',
            'transferred_from_location',
            'transferred_from_company',
            'transferred_from_custodian',
            'transferred_from_region',
            'transferred_from_site',
            'transferred_from_building',
            'transferred_from_floor',
            'transferred_from_room',
            'transferred_to_user',
            'transferred_to_department',
            'transferred_to_location',
            'transferred_to_company',
            'transferred_to_custodian',
            'transferred_to_region',
            'transferred_to_site',
            'transferred_to_building',
            'transferred_to_floor',
            'transferred_to_room',
            'created_by'
        )

        # Self-heal open legacy rows where from snapshot was not persisted at creation.
        self._hydrate_missing_from_snapshots(queryset)

        # Filter by status
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # Filter by asset
        asset_id = self.request.GET.get('asset')
        if asset_id:
            queryset = queryset.filter(asset_id=asset_id)

        # Filter by date range
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        if date_from:
            queryset = queryset.filter(transfer_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(transfer_date__lte=date_to)

        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(asset__asset_tag__icontains=search) |
                Q(asset__asset_code__icontains=search) |
                Q(asset__name__icontains=search) |
                Q(transfer_no__icontains=search) |
                Q(transfer_reason__icontains=search)
            )

        # Employees can create transfers but should only view transfers they initiated or are involved in
        user = self.request.user
        if user.role == user.Role.EMPLOYEE:
            queryset = queryset.filter(
                Q(created_by=user) | Q(transferred_to_user=user) | Q(transferred_from_user=user)
            )

        return queryset.order_by('-transfer_date')
    
    def get_queryset(self):
        return self.get_filtered_queryset()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        org = self.request.user.organization
        
        context['status_choices'] = AssetTransfer.Status.choices
        context['assets'] = Asset.objects.filter(organization=org, is_deleted=False)
        context['status_filter'] = self.request.GET.get('status', '')
        context['asset_filter'] = self.request.GET.get('asset', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['search'] = self.request.GET.get('search', '')
        
        return context


class AssetTransferExportExcelView(AssetTransferListView):
    """Export filtered asset transfers to Excel."""

    def get(self, request, *args, **kwargs):
        transfers = self.get_filtered_queryset()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Asset Transfers"

        headers = [
            'Transfer No', 'Asset Tag', 'Asset Name', 
            'From - User', 'From - Department', 'From - Location', 'From - Company', 'From - Custodian', 'From - Region', 'From - Site', 'From - Building', 'From - Floor', 'From - Room',
            'To - User', 'To - Department', 'To - Location', 'To - Company', 'To - Custodian', 'To - Region', 'To - Site', 'To - Building', 'To - Floor', 'To - Room',
            'Transfer Reason', 'Movement Reason', 'Status', 'Transfer Date', 'Expected Receipt', 'Actual Receipt', 'Created By'
        ]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for transfer in transfers:
            # FROM details
            from_user = transfer.transferred_from_user.get_full_name() if transfer.transferred_from_user else '-'
            from_dept = transfer.transferred_from_department.name if transfer.transferred_from_department else '-'
            from_location = transfer.transferred_from_location.name if transfer.transferred_from_location else '-'
            from_company = transfer.transferred_from_company.name if transfer.transferred_from_company else '-'
            from_custodian = str(transfer.transferred_from_custodian) if transfer.transferred_from_custodian else '-'
            from_region = transfer.transferred_from_region.name if transfer.transferred_from_region else '-'
            from_site = transfer.transferred_from_site.name if transfer.transferred_from_site else '-'
            from_building = transfer.transferred_from_building.name if transfer.transferred_from_building else '-'
            from_floor = transfer.transferred_from_floor.name if transfer.transferred_from_floor else '-'
            from_room = transfer.transferred_from_room.name if transfer.transferred_from_room else '-'

            # TO details
            to_user = transfer.transferred_to_user.get_full_name() if transfer.transferred_to_user else '-'
            to_dept = transfer.transferred_to_department.name if transfer.transferred_to_department else '-'
            to_location = transfer.transferred_to_location.name if transfer.transferred_to_location else '-'
            to_company = transfer.transferred_to_company.name if transfer.transferred_to_company else '-'
            to_custodian = str(transfer.transferred_to_custodian) if transfer.transferred_to_custodian else '-'
            to_region = transfer.transferred_to_region.name if transfer.transferred_to_region else '-'
            to_site = transfer.transferred_to_site.name if transfer.transferred_to_site else '-'
            to_building = transfer.transferred_to_building.name if transfer.transferred_to_building else '-'
            to_floor = transfer.transferred_to_floor.name if transfer.transferred_to_floor else '-'
            to_room = transfer.transferred_to_room.name if transfer.transferred_to_room else '-'

            created_by_value = transfer.created_by.get_full_name() if transfer.created_by else '-'

            ws.append([
                transfer.transfer_no or '',
                transfer.asset.asset_tag if transfer.asset else '',
                transfer.asset.name if transfer.asset else '',
                from_user, from_dept, from_location, from_company, from_custodian, from_region, from_site, from_building, from_floor, from_room,
                to_user, to_dept, to_location, to_company, to_custodian, to_region, to_site, to_building, to_floor, to_room,
                transfer.transfer_reason or '',
                transfer.movement_reason or '',
                transfer.get_status_display(),
                transfer.transfer_date.strftime('%Y-%m-%d %H:%M:%S') if transfer.transfer_date else '',
                transfer.expected_receipt_date.strftime('%Y-%m-%d') if transfer.expected_receipt_date else '',
                transfer.actual_receipt_date.strftime('%Y-%m-%d %H:%M:%S') if transfer.actual_receipt_date else '',
                created_by_value,
            ])

        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is None:
                    continue
                max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 45) if max_length else 12

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="asset_transfers_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response


class AssetTransferCreateView(LoginRequiredMixin, CreateView):
    """Create a new asset transfer"""
    model = AssetTransfer
    form_class = AssetTransferForm
    template_name = 'assets/transfer_form.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def form_valid(self, form):
        organization = self.request.user.organization
        form.instance.organization = organization
        form.instance.created_by = self.request.user
        # Asset IDs come from the hidden asset_ids CharField (plain string, possibly comma-separated UUIDs)
        asset_raw = form.cleaned_data.get('asset_ids') or self.request.POST.get('asset_ids', '')
        submitted_asset_ids = [s.strip() for s in str(asset_raw).split(',') if s.strip()]

        # Preserve order but ignore duplicate IDs submitted by the client.
        seen_asset_ids = set()
        asset_ids = []
        for aid in submitted_asset_ids:
            if aid in seen_asset_ids:
                continue
            seen_asset_ids.add(aid)
            asset_ids.append(aid)

        if not asset_ids:
            form.add_error(None, 'Please add at least one asset before submitting.')
            return self.form_invalid(form)

        created = []
        skipped_duplicates = []
        for aid in asset_ids:
            try:
                _active_disposal_statuses = [
                    AssetDisposal.Status.PENDING,
                    AssetDisposal.Status.MANAGER_APPROVED,
                    AssetDisposal.Status.APPROVED,
                    AssetDisposal.Status.COMPLETED,
                ]
                asset_obj = Asset.objects.filter(is_deleted=False).exclude(
                    disposals__status__in=_active_disposal_statuses
                ).get(pk=aid, organization=organization)
            except (Asset.DoesNotExist, Exception):
                continue

            # Block duplicate transfer requests for the same asset while one is still open.
            duplicate_exists = AssetTransfer.objects.filter(
                organization=organization,
                asset=asset_obj,
                status__in=[AssetTransfer.Status.PENDING, AssetTransfer.Status.IN_TRANSIT],
            ).exists()
            if duplicate_exists:
                skipped_duplicates.append(asset_obj.asset_tag)
                continue

            from_snapshot = {
                'transferred_from_user': getattr(asset_obj, 'assigned_to', None),
                'transferred_from_department': getattr(asset_obj, 'department', None),
                'transferred_from_location': getattr(asset_obj, 'location', None),
                'transferred_from_region': getattr(asset_obj, 'region', None),
                'transferred_from_site': getattr(asset_obj, 'site', None),
                'transferred_from_building': getattr(asset_obj, 'building', None),
                'transferred_from_floor': getattr(asset_obj, 'floor', None),
                'transferred_from_room': getattr(asset_obj, 'room', None),
                'transferred_from_company': getattr(asset_obj, 'company', None),
                'transferred_from_custodian': getattr(asset_obj, 'custodian', None),
            }

            at = AssetTransfer.objects.create(
                organization=organization,
                created_by=self.request.user,
                asset=asset_obj,
                **from_snapshot,
                transfer_no=form.cleaned_data.get('transfer_no'),
                transfer_description=form.cleaned_data.get('transfer_description'),
                transferred_to_region=form.cleaned_data.get('transferred_to_region'),
                transferred_to_site=form.cleaned_data.get('transferred_to_site'),
                transferred_to_building=form.cleaned_data.get('transferred_to_building'),
                transferred_to_floor=form.cleaned_data.get('transferred_to_floor'),
                transferred_to_room=form.cleaned_data.get('transferred_to_room'),
                transferred_to_location=form.cleaned_data.get('transferred_to_location'),
                transferred_to_company=form.cleaned_data.get('transferred_to_company'),
                transferred_to_department=form.cleaned_data.get('transferred_to_department'),
                transferred_to_custodian=form.cleaned_data.get('transferred_to_custodian'),
                movement_reason=form.cleaned_data.get('movement_reason'),
                requester_name=form.cleaned_data.get('requester_name'),
            )
            created.append(at)

        if created:
            if skipped_duplicates:
                messages.warning(
                    self.request,
                    f'Created {len(created)} transfer request(s). '
                    f'Skipped {len(skipped_duplicates)} duplicate request(s) already pending/in transit.'
                )
            else:
                messages.success(self.request, f'Asset transfer created for {len(created)} asset(s).')
            return redirect(reverse('transfer-detail', kwargs={'pk': created[0].pk}))

        if skipped_duplicates:
            form.add_error(
                None,
                'No new transfer requests were created because selected asset(s) already have pending or in-transit transfer requests.'
            )
            return self.form_invalid(form)

        form.add_error(None, 'No valid assets were found. Please check the asset tag/ID and try again.')
        return self.form_invalid(form)
    
    def get_success_url(self):
        return reverse('transfer-detail', kwargs={'pk': self.object.pk})


class AssetTransferDetailView(LoginRequiredMixin, DetailView):
    """View details of a specific asset transfer"""
    model = AssetTransfer
    template_name = 'assets/transfer_detail.html'
    context_object_name = 'transfer'
    
    def get_queryset(self):
        org = self.request.user.organization
        qs = AssetTransfer.objects.filter(organization=org).select_related(
            'asset',
            'transferred_from_user',
            'transferred_from_department',
            'transferred_from_location',
            'transferred_to_user',
            'transferred_to_department',
            'transferred_to_location',
            'created_by',
            'received_by'
        )
        # Employees may view only transfers they created or where they're involved
        user = self.request.user
        if user.role == user.Role.EMPLOYEE:
            qs = qs.filter(Q(created_by=user) | Q(transferred_to_user=user) | Q(transferred_from_user=user))
        return qs

    def get_object(self, queryset=None):
        transfer = super().get_object(queryset=queryset)
        if (
            transfer.status in [AssetTransfer.Status.PENDING, AssetTransfer.Status.IN_TRANSIT]
            and transfer.transferred_from_location is None
        ):
            transfer.snapshot_from_asset()
            transfer.save(update_fields=[
                'transferred_from_user',
                'transferred_from_department',
                'transferred_from_location',
                'transferred_from_region',
                'transferred_from_site',
                'transferred_from_building',
                'transferred_from_floor',
                'transferred_from_room',
                'transferred_from_company',
                'transferred_from_custodian',
                'updated_at',
            ])
        return transfer


class AssetTransferUpdateView(LoginRequiredMixin, UpdateView):
    """Update an asset transfer (mainly for status changes)"""
    model = AssetTransfer
    form_class = AssetTransferForm
    template_name = 'assets/transfer_form.html'

    def _is_transfer_approver(self):
        user = self.request.user
        return user.is_superuser or user.role in [user.Role.ADMIN, user.Role.SENIOR_MANAGER]

    def dispatch(self, request, *args, **kwargs):
        if not self._is_transfer_approver():
            return HttpResponseForbidden('Only senior manager or admin can approve or update asset transfers.')
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        org = self.request.user.organization
        return AssetTransfer.objects.filter(organization=org)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs
    
    def form_valid(self, form):
        messages.success(self.request, 'Asset transfer updated successfully')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('transfer-detail', kwargs={'pk': self.object.pk})




class AssetTransferApproveView(LoginRequiredMixin, View):
    """Approve a PENDING transfer -> IN_TRANSIT (Manager / Admin only)."""

    def _is_approver(self, user):
        return user.is_superuser or user.role in [user.Role.ADMIN, user.Role.SENIOR_MANAGER, user.Role.CHECKER]

    def post(self, request, pk):
        if not self._is_approver(request.user):
            return HttpResponseForbidden('Only a Manager or Admin can approve transfers.')
        transfer = get_object_or_404(
            AssetTransfer,
            pk=pk,
            organization=request.user.organization,
            status=AssetTransfer.Status.PENDING,
        )
        transfer.status = AssetTransfer.Status.IN_TRANSIT
        transfer.save(update_fields=['status', 'updated_at'])
        messages.success(request, f'Transfer approved â€” {transfer.asset.asset_tag} is now In Transit.')
        return redirect(reverse('transfer-detail', kwargs={'pk': transfer.pk}))

class AssetTransferReceiveView(LoginRequiredMixin, UpdateView):
    """Mark an asset transfer as received"""
    model = AssetTransfer
    form_class = AssetTransferReceiveForm
    template_name = 'assets/transfer_receive.html'

    def _is_transfer_approver(self):
        user = self.request.user
        return user.is_superuser or user.role in [user.Role.ADMIN, user.Role.SENIOR_MANAGER, user.Role.CHECKER]

    def dispatch(self, request, *args, **kwargs):
        if not self._is_transfer_approver():
            return HttpResponseForbidden('Only manager, senior manager, or admin can approve asset transfer requests.')
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        org = self.request.user.organization
        return AssetTransfer.objects.filter(organization=org)
    
    def form_valid(self, form):
        form.instance.received_by = self.request.user
        user = self.request.user
        is_final_approver = user.is_superuser or user.role in [user.Role.ADMIN, user.Role.SENIOR_MANAGER, user.Role.CHECKER]

        if form.instance.status == AssetTransfer.Status.RECEIVED:
            if not form.instance.actual_receipt_date:
                form.instance.actual_receipt_date = date.today()

            if is_final_approver and form.instance.asset:
                asset = form.instance.asset
                transfer = form.instance
                fields_to_update = []

                if transfer.transferred_to_user is not None:
                    asset.assigned_to = transfer.transferred_to_user
                    fields_to_update.append('assigned_to')
                if transfer.transferred_to_department is not None:
                    asset.department = transfer.transferred_to_department
                    fields_to_update.append('department')
                    if getattr(transfer.transferred_to_department, 'branch', None) is not None:
                        asset.branch = transfer.transferred_to_department.branch
                        fields_to_update.append('branch')
                if transfer.transferred_to_location is not None:
                    asset.location = transfer.transferred_to_location
                    fields_to_update.append('location')
                if transfer.transferred_to_region is not None:
                    asset.region = transfer.transferred_to_region
                    fields_to_update.append('region')
                if transfer.transferred_to_site is not None:
                    asset.site = transfer.transferred_to_site
                    fields_to_update.append('site')
                if transfer.transferred_to_building is not None:
                    asset.building = transfer.transferred_to_building
                    fields_to_update.append('building')
                    if getattr(transfer.transferred_to_building, 'branch', None) is not None:
                        asset.branch = transfer.transferred_to_building.branch
                        fields_to_update.append('branch')
                if transfer.transferred_to_floor is not None:
                    asset.floor = transfer.transferred_to_floor
                    fields_to_update.append('floor')
                if transfer.transferred_to_room is not None:
                    asset.room = transfer.transferred_to_room
                    fields_to_update.append('room')
                if transfer.transferred_to_company is not None:
                    asset.company = transfer.transferred_to_company
                    fields_to_update.append('company')
                if transfer.transferred_to_custodian is not None:
                    asset.custodian = transfer.transferred_to_custodian
                    fields_to_update.append('custodian')

                if fields_to_update:
                    asset.save(update_fields=list(dict.fromkeys(fields_to_update)))

            messages.success(self.request, f'Asset transfer marked as received: {form.instance.asset.asset_tag}')
        elif form.instance.status == AssetTransfer.Status.REJECTED:
            messages.warning(self.request, f'Asset transfer rejected: {form.instance.asset.asset_tag}')
        
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('transfer-detail', kwargs={'pk': self.object.pk})


# ========================
# Asset Disposal Views
# ========================

class AssetDisposalListView(LoginRequiredMixin, ListView):
    """List asset disposal requests"""
    model = AssetDisposal
    template_name = 'assets/disposal_list.html'
    context_object_name = 'disposals'
    paginate_by = 50

    def get_filtered_queryset(self):
        org = self.request.user.organization
        qs = AssetDisposal.objects.filter(organization=org).select_related(
            'asset', 'requested_by', 'approved_by'
        )
        
        # Filter by status
        status = self.request.GET.get('status')
        if status:
            qs = qs.filter(status=status)
        
        # Search by asset tag or name
        search = self.request.GET.get('search')
        if search:
            qs = qs.filter(
                Q(asset__asset_tag__icontains=search) |
                Q(asset__name__icontains=search) |
                Q(batch_reference__icontains=search)
            )
        
        # Employees can see only their own disposal requests
        user = self.request.user
        if user.role == user.Role.EMPLOYEE:
            qs = qs.filter(requested_by=user)

        return qs.order_by('-created_at')

    def get_queryset(self):
        return self.get_filtered_queryset()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        context['status_choices'] = AssetDisposal.Status.choices
        context['status_filter'] = self.request.GET.get('status', '')
        context['search'] = self.request.GET.get('search', '')
        context['can_manager_bulk_disposal_action'] = user.role in [user.Role.SENIOR_MANAGER, user.Role.ADMIN] or user.is_superuser
        context['can_admin_bulk_disposal_action'] = user.role == user.Role.ADMIN or user.is_superuser

        batch_rows = list(
            AssetDisposal.objects.filter(organization=user.organization)
            .exclude(batch_reference='')
            .values('batch_reference')
            .annotate(total=models.Count('id'))
        )
        batch_counts = {row['batch_reference']: row['total'] for row in batch_rows}
        for disposal in context.get('disposals', []):
            ref = (disposal.batch_reference or '').strip()
            disposal.batch_count = batch_counts.get(ref, 1) if ref else 1

        return context


class AssetDisposalBulkActionView(LoginRequiredMixin, View):
    """Batch approve/reject disposal requests with role-aware transitions."""

    def post(self, request, *args, **kwargs):
        user = request.user
        org = user.organization

        disposal_ids = request.POST.getlist('disposal_ids')
        action = (request.POST.get('bulk_action') or '').strip().lower()
        reason = (request.POST.get('bulk_reason') or '').strip()

        if not disposal_ids:
            messages.warning(request, 'Please select at least one disposal request.')
            return redirect(request.META.get('HTTP_REFERER') or reverse('disposal-list'))

        can_admin = user.is_superuser or user.role == user.Role.ADMIN
        can_manager = can_admin or user.role == user.Role.SENIOR_MANAGER

        if not can_manager:
            messages.error(request, 'You do not have permission to perform batch disposal actions.')
            return redirect(request.META.get('HTTP_REFERER') or reverse('disposal-list'))

        if action not in ['approve', 'reject']:
            messages.warning(request, 'Please choose a valid batch action.')
            return redirect(request.META.get('HTTP_REFERER') or reverse('disposal-list'))

        qs = AssetDisposal.objects.filter(organization=org, id__in=disposal_ids).select_related('asset')

        processed_count = 0
        skipped_count = 0
        approved_assets_count = 0

        with transaction.atomic():
            for disposal in qs:
                if can_admin:
                    eligible = disposal.status in [AssetDisposal.Status.PENDING, AssetDisposal.Status.MANAGER_APPROVED]
                    if not eligible:
                        skipped_count += 1
                        continue

                    if action == 'approve':
                        disposal.status = AssetDisposal.Status.APPROVED
                        disposal.approved_by = user
                        disposal.approved_at = datetime.now()
                        if reason:
                            disposal.notes = ((disposal.notes or '').strip() + ('\n' if disposal.notes else '') + f'Batch approval note: {reason}').strip()
                        disposal.save(update_fields=['status', 'approved_by', 'approved_at', 'notes', 'updated_at'])

                        if disposal.asset and not disposal.asset.is_deleted:
                            disposal.asset.status = Asset.Status.RETIRED
                            disposal.asset.is_deleted = True
                            disposal.asset.save(update_fields=['status', 'is_deleted'])
                            approved_assets_count += 1
                    else:
                        disposal.status = AssetDisposal.Status.REJECTED
                        disposal.approved_by = user
                        disposal.approved_at = datetime.now()
                        if reason:
                            disposal.rejection_reason = reason
                        disposal.save(update_fields=['status', 'approved_by', 'approved_at', 'rejection_reason', 'updated_at'])

                    processed_count += 1
                    continue

                # Senior Manager flow
                eligible = disposal.status == AssetDisposal.Status.PENDING
                if not eligible:
                    skipped_count += 1
                    continue

                if action == 'approve':
                    disposal.status = AssetDisposal.Status.MANAGER_APPROVED
                    disposal.manager_approved_by = user
                    disposal.manager_approved_at = datetime.now()
                    if reason:
                        disposal.notes = ((disposal.notes or '').strip() + ('\n' if disposal.notes else '') + f'Batch manager note: {reason}').strip()
                    disposal.save(update_fields=['status', 'manager_approved_by', 'manager_approved_at', 'notes', 'updated_at'])
                else:
                    disposal.status = AssetDisposal.Status.REJECTED
                    if reason:
                        disposal.manager_rejection_reason = reason
                    disposal.save(update_fields=['status', 'manager_rejection_reason', 'updated_at'])

                processed_count += 1

        if approved_assets_count > 0:
            invalidate_dashboard_cache_for_org(org)

        if processed_count > 0:
            if skipped_count > 0:
                messages.success(request, f'Batch action completed: {processed_count} processed, {skipped_count} skipped (status not eligible).')
            else:
                messages.success(request, f'Batch action completed: {processed_count} processed.')
        else:
            messages.warning(request, 'No disposal requests were processed. Check status eligibility for your role.')

        return redirect(request.META.get('HTTP_REFERER') or reverse('disposal-list'))


class AssetDisposalExportPDFView(AssetDisposalListView):
    """Export filtered asset disposal requests to PDF."""

    def get(self, request, *args, **kwargs):
        disposals = self.get_filtered_queryset()

        from fpdf import FPDF

        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=10)
        pdf.set_margins(10, 10, 10)
        pdf.add_page()

        def safe_text(value):
            text = str(value or '')
            # Keep PDF output robust for non-latin punctuation/symbols.
            return text.encode('latin-1', 'replace').decode('latin-1')

        generated_on = timezone.now().strftime('%d %b %Y %H:%M')
        status_filter = request.GET.get('status', '')
        search_filter = request.GET.get('search', '')

        pdf.set_fill_color(48, 84, 150)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 13)
        pdf.cell(0, 9, safe_text('Asset Disposal Report'), ln=1, fill=True)

        pdf.set_font('Helvetica', '', 9)
        pdf.cell(0, 6, safe_text(f"Generated: {generated_on}    Total Records: {disposals.count()}"), ln=1)
        pdf.cell(
            0,
            6,
            safe_text(f"Filters - Status: {status_filter or 'All'} | Search: {search_filter or 'None'}"),
            ln=1
        )
        pdf.ln(2)

        headers = ['Tag', 'Asset Name', 'Method', 'Status', 'Requested By', 'Requested Date']
        col_widths = [35, 95, 35, 35, 45, 30]

        pdf.set_fill_color(48, 84, 150)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 8)
        for header, width in zip(headers, col_widths):
            pdf.cell(width, 7, safe_text(header), border=1, align='L', fill=True)
        pdf.ln()

        pdf.set_text_color(0, 0, 0)
        pdf.set_font('Helvetica', '', 8)
        row_fill = False

        for disposal in disposals:
            if row_fill:
                pdf.set_fill_color(245, 247, 251)
            else:
                pdf.set_fill_color(255, 255, 255)

            requested_by = '-'
            if disposal.requested_by:
                requested_by = disposal.requested_by.get_full_name() or disposal.requested_by.username

            row_values = [
                disposal.asset.asset_tag if disposal.asset else '-',
                disposal.asset.name if disposal.asset else '-',
                disposal.get_disposal_method_display(),
                disposal.get_status_display(),
                requested_by,
                disposal.created_at.strftime('%Y-%m-%d') if disposal.created_at else '-',
            ]

            for idx, (value, width) in enumerate(zip(row_values, col_widths)):
                align = 'L'
                if idx == 5:
                    align = 'C'
                pdf.cell(width, 6, safe_text(value)[:70], border=1, align=align, fill=True)
            pdf.ln()
            row_fill = not row_fill

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = (
            f'attachment; filename="asset_disposals_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        )
        response.write(bytes(pdf.output()))
        return response


class AssetDisposalCreateView(LoginRequiredMixin, CreateView):
    """Create a new asset disposal request (any authenticated user)."""
    model = AssetDisposal
    form_class = AssetDisposalForm
    template_name = 'assets/disposal_form.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('login')
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def _parse_asset_ids(self, raw=None):
        if raw is None:
            raw = self.request.POST.get('asset_ids') or self.request.GET.get('asset_ids', '')
        return [s.strip() for s in str(raw).split(',') if s.strip()]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        selected_ids = self._parse_asset_ids()
        if selected_ids:
            context['selected_assets'] = list(
                Asset.objects.filter(
                    id__in=selected_ids,
                    organization=self.request.user.organization,
                    is_deleted=False
                ).order_by('asset_tag')
            )
            context['selected_asset_ids'] = ','.join([str(a.id) for a in context['selected_assets']])
        else:
            context['selected_assets'] = []
            context['selected_asset_ids'] = ''
        return context
    
    def form_valid(self, form):
        org = self.request.user.organization
        selected_ids = self._parse_asset_ids(form.cleaned_data.get('asset_ids'))
        form_assets = list(form.cleaned_data.get('selected_assets') or [])
        skipped_ineligible_count = 0

        if selected_ids:
            requested_ids = list(dict.fromkeys(selected_ids))
            assets = list(
                Asset.objects.filter(
                    id__in=requested_ids,
                    organization=org,
                    is_deleted=False,
                    status__in=[Asset.Status.ACTIVE, Asset.Status.IN_STORAGE, Asset.Status.UNDER_MAINTENANCE]
                )
            )
            skipped_ineligible_count = max(len(requested_ids) - len(assets), 0)
        else:
            assets = [a for a in form_assets if a.organization_id == org.id and not a.is_deleted]

        if not assets:
            form.add_error(None, 'No eligible assets found for disposal. Please reselect assets and try again.')
            return self.form_invalid(form)

        created_count = 0
        skipped_count = 0
        batch_reference = f"DSP-{uuid4().hex[:8].upper()}" if len(assets) > 1 else ''

        for asset in assets:
            already_pending = AssetDisposal.objects.filter(
                organization=org,
                asset=asset,
                status__in=[AssetDisposal.Status.PENDING, AssetDisposal.Status.MANAGER_APPROVED]
            ).exists()

            if already_pending:
                skipped_count += 1
                continue

            AssetDisposal.objects.create(
                organization=org,
                requested_by=self.request.user,
                asset=asset,
                batch_reference=batch_reference,
                disposal_method=form.cleaned_data.get('disposal_method') or AssetDisposal.DisposalMethod.SCRAP,
                reason=form.cleaned_data.get('reason') or '',
                disposal_date=form.cleaned_data.get('disposal_date'),
                estimated_salvage_value=form.cleaned_data.get('estimated_salvage_value'),
                notes=form.cleaned_data.get('notes') or '',
            )
            created_count += 1

        if created_count > 0:
            total_skipped = skipped_count + skipped_ineligible_count
            batch_note = f' Batch ref: {batch_reference}.' if batch_reference else ''
            if total_skipped > 0:
                notes = []
                if skipped_count > 0:
                    notes.append(f'{skipped_count} already pending')
                if skipped_ineligible_count > 0:
                    notes.append(f'{skipped_ineligible_count} ineligible or not accessible')
                messages.success(self.request, f"Disposal requests created for {created_count} asset(s). {total_skipped} skipped ({', '.join(notes)}).{batch_note}")
            else:
                messages.success(self.request, f'Disposal requests created for {created_count} asset(s).{batch_note}')
            return redirect('disposal-list')

        if skipped_ineligible_count > 0 and skipped_count == 0:
            messages.warning(self.request, 'No disposal requests were created. Selected assets were ineligible or not accessible.')
        else:
            messages.error(self.request, 'No new disposal requests were created. Selected assets may already have pending requests or be ineligible.')
        return redirect('disposal-list')
    
    def get_success_url(self):
        return reverse('disposal-detail', kwargs={'pk': self.object.pk})


class AssetDisposalDetailView(LoginRequiredMixin, DetailView):
    """View details of an asset disposal request"""
    model = AssetDisposal
    template_name = 'assets/disposal_detail.html'
    context_object_name = 'disposal'
    
    def get_queryset(self):
        org = self.request.user.organization
        qs = AssetDisposal.objects.filter(organization=org).select_related(
            'asset', 'requested_by', 'approved_by'
        )
        
        # Employees can view only their own disposals
        user = self.request.user
        if user.role == user.Role.EMPLOYEE:
            qs = qs.filter(requested_by=user)
        
        return qs


class AssetDisposalManagerApproveView(LoginRequiredMixin, UpdateView):
    """Manager approval of asset disposal request (step 1)"""
    model = AssetDisposal
    form_class = AssetDisposalManagerApprovalForm
    template_name = 'assets/disposal_manager_approve.html'
    
    def test_func(self):
        """Only senior manager or admin can approve disposal requests."""
        user = self.request.user
        return user.is_superuser or user.role in [user.Role.SENIOR_MANAGER, user.Role.ADMIN]
    
    def dispatch(self, request, *args, **kwargs):
        if not self.test_func():
            messages.error(request, 'You do not have permission to review disposal requests.')
            return redirect('disposal-list')
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        org = self.request.user.organization
        return AssetDisposal.objects.filter(organization=org, status=AssetDisposal.Status.PENDING)
    
    def form_valid(self, form):
        form.instance.manager_approved_by = self.request.user
        form.instance.manager_approved_at = datetime.now()
        
        if form.instance.status == AssetDisposal.Status.MANAGER_APPROVED:
            messages.success(self.request, f'Asset disposal request approved by manager: {form.instance.asset.asset_tag}. Pending admin approval.')
        elif form.instance.status == AssetDisposal.Status.REJECTED:
            messages.warning(self.request, f'Asset disposal request rejected by manager: {form.instance.asset.asset_tag}')
        
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('disposal-detail', kwargs={'pk': self.object.pk})


class AssetDisposalApproveView(LoginRequiredMixin, UpdateView):
    """Final admin approval of asset disposal request (step 2)"""
    model = AssetDisposal
    form_class = AssetDisposalApprovalForm
    template_name = 'assets/disposal_approve.html'
    
    def test_func(self):
        """Only admins can give final approval"""
        return self.request.user.is_superuser or self.request.user.role == self.request.user.Role.ADMIN
    
    def dispatch(self, request, *args, **kwargs):
        if not self.test_func():
            messages.error(request, 'You do not have permission to approve disposal requests.')
            return redirect('disposal-list')
        return super().dispatch(request, *args, **kwargs)
    
    def get_queryset(self):
        org = self.request.user.organization
        # Admins can approve both PENDING (direct) and MANAGER_APPROVED disposals
        return AssetDisposal.objects.filter(
            organization=org,
            status__in=[AssetDisposal.Status.PENDING, AssetDisposal.Status.MANAGER_APPROVED]
        )

    def form_valid(self, form):
        form.instance.approved_by = self.request.user
        form.instance.approved_at = datetime.now()

        # Capture the decision before saving
        new_status = form.cleaned_data.get('status')

        response = super().form_valid(form)

        if new_status == AssetDisposal.Status.APPROVED:
            # Capture full asset snapshot before soft-deleting
            asset = self.object.asset
            snapshot = {
                'asset_tag': asset.asset_tag,
                'asset_code': asset.asset_code,
                'name': asset.name,
                'serial_number': asset.serial_number,
                'status': asset.status,
                'category': str(asset.category) if asset.category else None,
                'sub_category': str(asset.sub_category) if asset.sub_category else None,
                'brand': asset.brand,
                'model': asset.model,
                'purchase_price': str(asset.purchase_price) if asset.purchase_price else None,
                'purchase_date': str(asset.purchase_date) if asset.purchase_date else None,
                'useful_life_years': asset.useful_life_years,
                'salvage_value': str(asset.salvage_value) if asset.salvage_value else None,
                'depreciation_method': asset.depreciation_method,
                'location': str(asset.location) if asset.location else None,
                'sub_location': str(asset.sub_location) if asset.sub_location else None,
                'department': str(asset.department) if asset.department else None,
                'branch': str(asset.branch) if asset.branch else None,
                'site': str(asset.site) if asset.site else None,
                'building': str(asset.building) if asset.building else None,
                'room': str(asset.room) if asset.room else None,
                'assigned_to': str(asset.assigned_to) if asset.assigned_to else None,
                'custodian': str(asset.custodian) if asset.custodian else None,
                'vendor': str(asset.vendor) if asset.vendor else None,
                'supplier': str(asset.supplier) if asset.supplier else None,
                'warranty_start': str(asset.warranty_start) if asset.warranty_start else None,
                'warranty_end': str(asset.warranty_end) if asset.warranty_end else None,
                'notes': asset.notes,
                'condition': asset.condition if hasattr(asset, 'condition') else None,
            }

            # Mark the asset as retired and soft-deleted
            asset.status = Asset.Status.RETIRED
            asset.is_deleted = True
            asset.save(update_fields=['status', 'is_deleted'])
            invalidate_dashboard_cache_for_org(self.request.user.organization)
            messages.success(self.request, f'Asset disposal approved and asset {asset.asset_tag} removed from inventory.')
        elif new_status == AssetDisposal.Status.REJECTED:
            messages.warning(self.request, f'Asset disposal request rejected: {self.object.asset.asset_tag}')

        return response

    def get_success_url(self):
        return reverse('disposal-detail', kwargs={'pk': self.object.pk})


# --- DEPRECIATION REPORT VIEWS ---
class DepreciationReportCategoryView(LoginRequiredMixin, ListView):
    """Dedicated view for category-based depreciation report"""
    model = Asset
    template_name = 'assets/depreciation_report_category.html'
    context_object_name = 'assets'
    paginate_by = 50

    def get_queryset(self):
        queryset = Asset.objects.filter(
            organization=self.request.user.organization,
            is_deleted=False
        ).select_related(
            'category', 'sub_category', 'branch', 'assigned_to', 
            'site', 'building', 'brand_new', 'room', 'department',
            'region', 'location', 'sub_location', 'vendor', 
            'supplier', 'company', 'group', 'custodian'
        )

        queryset = queryset.filter(purchase_price__isnull=False, purchase_price__gt=0)

        queryset = queryset.filter(purchase_price__isnull=False, purchase_price__gt=0)

        queryset = queryset.filter(purchase_price__isnull=False, purchase_price__gt=0)

        # Only assets with financial value
        queryset = queryset.filter(purchase_price__isnull=False, purchase_price__gt=0)

        # Search filter
        query = self.request.GET.get('q')
        if query:
            q = (
                Q(name__icontains=query) |
                Q(asset_tag__icontains=query) |
                Q(asset_code__icontains=query) |
                Q(serial_number__icontains=query) |
                Q(category__name__icontains=query)
            )
            queryset = queryset.filter(q)
        
        # Date range filters (opening/closing style)
        self._opening_date = None
        self._closing_date = None
        depr_date_from = self.request.GET.get('depr_date_from')
        depr_date_to = self.request.GET.get('depr_date_to')

        if depr_date_from:
            try:
                self._opening_date = datetime.strptime(depr_date_from, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                self._opening_date = None

        if depr_date_to:
            try:
                self._closing_date = datetime.strptime(depr_date_to, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                self._closing_date = None

        # Include only assets that existed up to closing date
        if self._closing_date:
            queryset = queryset.filter(Q(purchase_date__lte=self._closing_date) | Q(purchase_date__isnull=True))
        
        # Dimension filters
        depr_filters = {
            'depr_category': 'category_id',
            'depr_group': 'group_id',
            'depr_department': 'department_id',
            'depr_site': 'site_id',
            'depr_branch': 'branch_id',
            'depr_building': 'building_id',
            'depr_location': 'location_id',
            'depr_tagging_status': 'tagging_status',
        }
        
        for param, field in depr_filters.items():
            val = self.request.GET.get(param)
            if val:
                queryset = queryset.filter(**{field: val})
        
        return queryset.order_by('-purchase_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        org = self.request.user.organization
        queryset = self.get_queryset()
        
        # Summary totals
        from django.db.models import Sum, Count
        from django.db.models.functions import Coalesce
        
        agg = queryset.aggregate(
            total_cost=Coalesce(Sum('purchase_price'), Decimal('0')),
            total_count=Count('id')
        )
        
        total_cost = agg['total_cost']
        total_count = agg['total_count']
        
        # Calculate exact values by iterating in batches
        BATCH_SIZE = 1000
        total_acc_dep = Decimal('0')
        total_nbv = Decimal('0')
        _cd = getattr(self, '_closing_date', None)
        for i in range(0, total_count, BATCH_SIZE):
            batch = list(queryset[i:i+BATCH_SIZE])
            for asset in batch:
                total_nbv += asset.current_value
                total_acc_dep += asset.get_accumulated_dep_at_date(_cd) if _cd else asset.accumulated_depreciation
        
        context['total_cost'] = total_cost
        context['total_acc_dep'] = total_acc_dep
        context['total_nbv'] = total_nbv
        context['total_assets_report'] = total_count

        opening_date = getattr(self, '_opening_date', None)
        closing_date = getattr(self, '_closing_date', None)

        if opening_date or closing_date:
            all_for_period = list(queryset)
            if opening_date:
                total_opening_value = sum(a.get_value_at_date(opening_date) for a in all_for_period) if all_for_period else Decimal('0')
            else:
                total_opening_value = total_nbv

            if closing_date:
                total_closing_value = sum(a.get_value_at_date(closing_date) for a in all_for_period) if all_for_period else Decimal('0')
            else:
                total_closing_value = total_nbv
        else:
            total_opening_value = total_nbv
            total_closing_value = total_nbv

        context['total_opening_value'] = total_opening_value
        context['total_closing_value'] = total_closing_value
        context['period_depreciation'] = total_opening_value - total_closing_value
        context['opening_date'] = opening_date
        context['closing_date'] = closing_date
        
        # Category grouping
        grouped_data = queryset.values('category', 'category__name').annotate(
            count=Count('id'),
            total_cost=Sum('purchase_price')
        ).order_by('-total_cost')[:100]
        
        grouped_list = []
        for group in grouped_data:
            cat_id = group['category']
            cat_assets = list(queryset.filter(category_id=cat_id))
            total_cat_dep = sum(a.get_accumulated_dep_at_date(closing_date) if closing_date else a.accumulated_depreciation for a in cat_assets) if cat_assets else Decimal('0')
            
            grouped_list.append({
                'id': cat_id,
                'name': group['category__name'] or 'Uncategorized',
                'total_cost': group['total_cost'] or Decimal('0'),
                'total_acc_dep': total_cat_dep,
                'total_nbv': (group['total_cost'] or Decimal('0')) - total_cat_dep,
                'count': group['count'],
            })
        
        context['grouped_data'] = grouped_list
        context['categories'] = Category.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['departments'] = Department.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['locations'] = Location.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['branches'] = Branch.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['sites'] = Site.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['buildings'] = Building.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['groups'] = Group.objects.filter(organization=org).only('id', 'name').order_by('name')
        
        # Filter persistence
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['query_params'] = query_params.urlencode()
        
        # Store filter values
        context['depr_date_from'] = self.request.GET.get('depr_date_from', '')
        context['depr_date_to'] = self.request.GET.get('depr_date_to', '')
        context['depr_category'] = self.request.GET.get('depr_category', '')
        context['depr_group'] = self.request.GET.get('depr_group', '')
        context['depr_department'] = self.request.GET.get('depr_department', '')
        context['depr_site'] = self.request.GET.get('depr_site', '')
        context['depr_branch'] = self.request.GET.get('depr_branch', '')
        context['depr_building'] = self.request.GET.get('depr_building', '')
        context['depr_location'] = self.request.GET.get('depr_location', '')
        context['search_query'] = self.request.GET.get('q', '')
        context['depr_tagging_status'] = self.request.GET.get('depr_tagging_status', '')
        context['tagging_status_choices'] = [('TAGGED', 'Tagged'), ('UNTAGGED', 'Untagged')]

        # Enrich paginated assets with opening/closing values
        if context.get('page_obj'):
            enriched_assets = []
            for asset in context['page_obj'].object_list:
                asset.opening_value = asset.get_value_at_date(opening_date) if opening_date else asset.current_value
                asset.closing_value = asset.get_value_at_date(closing_date) if closing_date else asset.current_value
                asset.period_depreciation = asset.opening_value - asset.closing_value
                enriched_assets.append(asset)
            context['assets'] = enriched_assets
        
        return context


class DepreciationReportDepartmentView(LoginRequiredMixin, ListView):
    """Dedicated view for department-based depreciation report"""
    model = Asset
    template_name = 'assets/depreciation_report_department.html'
    context_object_name = 'assets'
    paginate_by = 50

    def get_queryset(self):
        queryset = Asset.objects.filter(
            organization=self.request.user.organization,
            is_deleted=False
        ).select_related(
            'category', 'sub_category', 'branch', 'assigned_to', 
            'site', 'building', 'brand_new', 'room', 'department',
            'region', 'location', 'sub_location', 'vendor', 
            'supplier', 'company', 'group', 'custodian'
        )

        # Search filter
        query = self.request.GET.get('q')
        if query:
            q = (
                Q(name__icontains=query) |
                Q(asset_tag__icontains=query) |
                Q(asset_code__icontains=query) |
                Q(serial_number__icontains=query) |
                Q(department__name__icontains=query)
            )
            queryset = queryset.filter(q)
        
        # Date range filters (opening/closing style)
        self._opening_date = None
        self._closing_date = None
        depr_date_from = self.request.GET.get('depr_date_from')
        depr_date_to = self.request.GET.get('depr_date_to')

        if depr_date_from:
            try:
                self._opening_date = datetime.strptime(depr_date_from, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                self._opening_date = None

        if depr_date_to:
            try:
                self._closing_date = datetime.strptime(depr_date_to, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                self._closing_date = None

        if self._closing_date:
            queryset = queryset.filter(Q(purchase_date__lte=self._closing_date) | Q(purchase_date__isnull=True))
        
        # Dimension filters
        depr_filters = {
            'depr_category': 'category_id',
            'depr_group': 'group_id',
            'depr_department': 'department_id',
            'depr_site': 'site_id',
            'depr_branch': 'branch_id',
            'depr_building': 'building_id',
            'depr_location': 'location_id',
            'depr_tagging_status': 'tagging_status',
        }
        
        for param, field in depr_filters.items():
            val = self.request.GET.get(param)
            if val:
                queryset = queryset.filter(**{field: val})
        
        return queryset.order_by('-purchase_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        org = self.request.user.organization
        queryset = self.get_queryset()
        
        # Summary totals
        from django.db.models import Sum, Count
        from django.db.models.functions import Coalesce
        
        agg = queryset.aggregate(
            total_cost=Coalesce(Sum('purchase_price'), Decimal('0')),
            total_count=Count('id')
        )
        
        total_cost = agg['total_cost']
        total_count = agg['total_count']
        
        # Calculate exact values by iterating in batches
        BATCH_SIZE = 1000
        total_acc_dep = Decimal('0')
        total_nbv = Decimal('0')
        _cd = getattr(self, '_closing_date', None)
        for i in range(0, total_count, BATCH_SIZE):
            batch = list(queryset[i:i+BATCH_SIZE])
            for asset in batch:
                total_nbv += asset.current_value
                total_acc_dep += asset.get_accumulated_dep_at_date(_cd) if _cd else asset.accumulated_depreciation
        
        context['total_cost'] = total_cost
        context['total_acc_dep'] = total_acc_dep
        context['total_nbv'] = total_nbv
        context['total_assets_report'] = total_count

        opening_date = getattr(self, '_opening_date', None)
        closing_date = getattr(self, '_closing_date', None)
        all_for_period = list(queryset)

        if opening_date:
            total_opening_value = sum(a.get_value_at_date(opening_date) for a in all_for_period) if all_for_period else Decimal('0')
        else:
            total_opening_value = total_nbv

        if closing_date:
            total_closing_value = sum(a.get_value_at_date(closing_date) for a in all_for_period) if all_for_period else Decimal('0')
        else:
            total_closing_value = total_nbv

        context['total_opening_value'] = total_opening_value
        context['total_closing_value'] = total_closing_value
        context['period_depreciation'] = total_opening_value - total_closing_value
        context['opening_date'] = opening_date
        context['closing_date'] = closing_date
        
        # Department grouping
        grouped_data = queryset.values('department', 'department__name').annotate(
            count=Count('id'),
            total_cost=Sum('purchase_price')
        ).order_by('-total_cost')[:100]
        
        grouped_list = []
        for department in grouped_data:
            department_id = department['department']
            department_assets = list(queryset.filter(department_id=department_id))
            total_department_dep = sum(a.get_accumulated_dep_at_date(closing_date) if closing_date else a.accumulated_depreciation for a in department_assets) if department_assets else Decimal('0')
            
            grouped_list.append({
                'id': department_id,
                'name': department['department__name'] or 'Uncategorized',
                'total_cost': department['total_cost'] or Decimal('0'),
                'total_acc_dep': total_department_dep,
                'total_nbv': (department['total_cost'] or Decimal('0')) - total_department_dep,
                'count': department['count'],
            })
        
        context['grouped_data'] = grouped_list
        context['categories'] = Category.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['departments'] = Department.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['locations'] = Location.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['branches'] = Branch.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['sites'] = Site.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['buildings'] = Building.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['groups'] = Group.objects.filter(organization=org).only('id', 'name').order_by('name')
        
        # Filter persistence
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['query_params'] = query_params.urlencode()
        
        # Store filter values
        context['depr_date_from'] = self.request.GET.get('depr_date_from', '')
        context['depr_date_to'] = self.request.GET.get('depr_date_to', '')
        context['depr_category'] = self.request.GET.get('depr_category', '')
        context['depr_group'] = self.request.GET.get('depr_group', '')
        context['depr_department'] = self.request.GET.get('depr_department', '')
        context['depr_site'] = self.request.GET.get('depr_site', '')
        context['depr_branch'] = self.request.GET.get('depr_branch', '')
        context['depr_building'] = self.request.GET.get('depr_building', '')
        context['depr_location'] = self.request.GET.get('depr_location', '')
        context['search_query'] = self.request.GET.get('q', '')
        context['depr_tagging_status'] = self.request.GET.get('depr_tagging_status', '')
        context['tagging_status_choices'] = [('TAGGED', 'Tagged'), ('UNTAGGED', 'Untagged')]

        if context.get('page_obj'):
            enriched_assets = []
            for asset in context['page_obj'].object_list:
                asset.opening_value = asset.get_value_at_date(opening_date) if opening_date else asset.current_value
                asset.closing_value = asset.get_value_at_date(closing_date) if closing_date else asset.current_value
                asset.period_depreciation = asset.opening_value - asset.closing_value
                enriched_assets.append(asset)
            context['assets'] = enriched_assets
        
        return context

class DepreciationReportLocationView(LoginRequiredMixin, ListView):
    """Dedicated view for location-based depreciation report"""
    model = Asset
    template_name = 'assets/depreciation_report_location.html'
    context_object_name = 'assets'
    paginate_by = 50

    def get_queryset(self):
        queryset = Asset.objects.filter(
            organization=self.request.user.organization,
            is_deleted=False
        ).select_related(
            'category', 'sub_category', 'branch', 'assigned_to', 
            'site', 'building', 'brand_new', 'room', 'department',
            'region', 'location', 'sub_location', 'vendor', 
            'supplier', 'company', 'group', 'custodian'
        )

        # Search filter
        query = self.request.GET.get('q')
        if query:
            q = (
                Q(name__icontains=query) |
                Q(asset_tag__icontains=query) |
                Q(asset_code__icontains=query) |
                Q(serial_number__icontains=query) |
                Q(location__name__icontains=query)
            )
            queryset = queryset.filter(q)
        
        # Date range filters (opening/closing style)
        self._opening_date = None
        self._closing_date = None
        depr_date_from = self.request.GET.get('depr_date_from')
        depr_date_to = self.request.GET.get('depr_date_to')

        if depr_date_from:
            try:
                self._opening_date = datetime.strptime(depr_date_from, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                self._opening_date = None

        if depr_date_to:
            try:
                self._closing_date = datetime.strptime(depr_date_to, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                self._closing_date = None

        if self._closing_date:
            queryset = queryset.filter(Q(purchase_date__lte=self._closing_date) | Q(purchase_date__isnull=True))
        
        # Dimension filters
        depr_filters = {
            'depr_category': 'category_id',
            'depr_group': 'group_id',
            'depr_department': 'department_id',
            'depr_site': 'site_id',
            'depr_branch': 'branch_id',
            'depr_building': 'building_id',
            'depr_location': 'location_id',
            'depr_tagging_status': 'tagging_status',
        }
        
        for param, field in depr_filters.items():
            val = self.request.GET.get(param)
            if val:
                queryset = queryset.filter(**{field: val})
        
        return queryset.order_by('-purchase_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        org = self.request.user.organization
        queryset = self.get_queryset()
        
        # Summary totals
        from django.db.models import Sum, Count
        from django.db.models.functions import Coalesce
        
        agg = queryset.aggregate(
            total_cost=Coalesce(Sum('purchase_price'), Decimal('0')),
            total_count=Count('id')
        )
        
        total_cost = agg['total_cost']
        total_count = agg['total_count']
        
        # Calculate exact values by iterating in batches
        BATCH_SIZE = 1000
        total_acc_dep = Decimal('0')
        total_nbv = Decimal('0')
        _cd = getattr(self, '_closing_date', None)
        for i in range(0, total_count, BATCH_SIZE):
            batch = list(queryset[i:i+BATCH_SIZE])
            for asset in batch:
                total_nbv += asset.current_value
                total_acc_dep += asset.get_accumulated_dep_at_date(_cd) if _cd else asset.accumulated_depreciation
        
        context['total_cost'] = total_cost
        context['total_acc_dep'] = total_acc_dep
        context['total_nbv'] = total_nbv
        context['total_assets_report'] = total_count

        opening_date = getattr(self, '_opening_date', None)
        closing_date = getattr(self, '_closing_date', None)
        all_for_period = list(queryset)

        if opening_date:
            total_opening_value = sum(a.get_value_at_date(opening_date) for a in all_for_period) if all_for_period else Decimal('0')
        else:
            total_opening_value = total_nbv

        if closing_date:
            total_closing_value = sum(a.get_value_at_date(closing_date) for a in all_for_period) if all_for_period else Decimal('0')
        else:
            total_closing_value = total_nbv

        context['total_opening_value'] = total_opening_value
        context['total_closing_value'] = total_closing_value
        context['period_depreciation'] = total_opening_value - total_closing_value
        context['opening_date'] = opening_date
        context['closing_date'] = closing_date
        
        # Location grouping
        grouped_data = queryset.values('location', 'location__name').annotate(
            count=Count('id'),
            total_cost=Sum('purchase_price')
        ).order_by('-total_cost')[:100]
        
        grouped_list = []
        for location in grouped_data:
            location_id = location['location']
            location_assets = list(queryset.filter(location_id=location_id))
            total_location_dep = sum(a.get_accumulated_dep_at_date(closing_date) if closing_date else a.accumulated_depreciation for a in location_assets) if location_assets else Decimal('0')
            
            grouped_list.append({
                'id': location_id,
                'name': location['location__name'] or 'Uncategorized',
                'total_cost': location['total_cost'] or Decimal('0'),
                'total_acc_dep': total_location_dep,
                'total_nbv': (location['total_cost'] or Decimal('0')) - total_location_dep,
                'count': location['count'],
            })
        
        context['grouped_data'] = grouped_list
        context['categories'] = Category.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['departments'] = Department.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['locations'] = Location.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['branches'] = Branch.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['sites'] = Site.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['buildings'] = Building.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['groups'] = Group.objects.filter(organization=org).only('id', 'name').order_by('name')
        
        # Filter persistence
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['query_params'] = query_params.urlencode()
        
        # Store filter values
        context['depr_date_from'] = self.request.GET.get('depr_date_from', '')
        context['depr_date_to'] = self.request.GET.get('depr_date_to', '')
        context['depr_category'] = self.request.GET.get('depr_category', '')
        context['depr_group'] = self.request.GET.get('depr_group', '')
        context['depr_department'] = self.request.GET.get('depr_department', '')
        context['depr_site'] = self.request.GET.get('depr_site', '')
        context['depr_branch'] = self.request.GET.get('depr_branch', '')
        context['depr_building'] = self.request.GET.get('depr_building', '')
        context['depr_location'] = self.request.GET.get('depr_location', '')
        context['search_query'] = self.request.GET.get('q', '')
        context['depr_tagging_status'] = self.request.GET.get('depr_tagging_status', '')
        context['tagging_status_choices'] = [('TAGGED', 'Tagged'), ('UNTAGGED', 'Untagged')]
        
        if context.get('page_obj'):
            enriched_assets = []
            for asset in context['page_obj'].object_list:
                asset.opening_value = asset.get_value_at_date(opening_date) if opening_date else asset.current_value
                asset.closing_value = asset.get_value_at_date(closing_date) if closing_date else asset.current_value
                asset.period_depreciation = asset.opening_value - asset.closing_value
                enriched_assets.append(asset)
            context['assets'] = enriched_assets

        return context

class DepreciationReportGroupView(LoginRequiredMixin, ListView):
    """Dedicated view for group-based depreciation report"""
    model = Asset
    template_name = 'assets/depreciation_report_group.html'
    context_object_name = 'assets'
    paginate_by = 50

    def get_queryset(self):
        queryset = Asset.objects.filter(
            organization=self.request.user.organization,
            is_deleted=False
        ).select_related(
            'category', 'sub_category', 'branch', 'assigned_to', 
            'site', 'building', 'brand_new', 'room', 'department',
            'region', 'location', 'sub_location', 'vendor', 
            'supplier', 'company', 'group', 'custodian'
        )

        # Search filter
        query = self.request.GET.get('q')
        if query:
            q = (
                Q(name__icontains=query) |
                Q(asset_tag__icontains=query) |
                Q(asset_code__icontains=query) |
                Q(serial_number__icontains=query) |
                Q(group__name__icontains=query)
            )
            queryset = queryset.filter(q)
        
        # Date range filters (opening/closing style)
        self._opening_date = None
        self._closing_date = None
        depr_date_from = self.request.GET.get('depr_date_from')
        depr_date_to = self.request.GET.get('depr_date_to')

        if depr_date_from:
            try:
                self._opening_date = datetime.strptime(depr_date_from, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                self._opening_date = None

        if depr_date_to:
            try:
                self._closing_date = datetime.strptime(depr_date_to, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                self._closing_date = None

        if self._closing_date:
            queryset = queryset.filter(Q(purchase_date__lte=self._closing_date) | Q(purchase_date__isnull=True))
        
        # Dimension filters
        depr_filters = {
            'depr_category': 'category_id',
            'depr_group': 'group_id',
            'depr_department': 'department_id',
            'depr_site': 'site_id',
            'depr_branch': 'branch_id',
            'depr_building': 'building_id',
            'depr_location': 'location_id',
            'depr_tagging_status': 'tagging_status',
        }
        
        for param, field in depr_filters.items():
            val = self.request.GET.get(param)
            if val:
                queryset = queryset.filter(**{field: val})
        
        return queryset.order_by('-purchase_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        org = self.request.user.organization
        queryset = self.get_queryset()
        
        # Summary totals
        from django.db.models import Sum, Count
        from django.db.models.functions import Coalesce
        
        agg = queryset.aggregate(
            total_cost=Coalesce(Sum('purchase_price'), Decimal('0')),
            total_count=Count('id')
        )
        
        total_cost = agg['total_cost']
        total_count = agg['total_count']
        
        # Calculate exact values by iterating in batches
        BATCH_SIZE = 1000
        total_acc_dep = Decimal('0')
        total_nbv = Decimal('0')
        _cd = getattr(self, '_closing_date', None)
        for i in range(0, total_count, BATCH_SIZE):
            batch = list(queryset[i:i+BATCH_SIZE])
            for asset in batch:
                total_nbv += asset.current_value
                total_acc_dep += asset.get_accumulated_dep_at_date(_cd) if _cd else asset.accumulated_depreciation
        
        context['total_cost'] = total_cost
        context['total_acc_dep'] = total_acc_dep
        context['total_nbv'] = total_nbv
        context['total_assets_report'] = total_count

        opening_date = getattr(self, '_opening_date', None)
        closing_date = getattr(self, '_closing_date', None)
        all_for_period = list(queryset)

        if opening_date:
            total_opening_value = sum(a.get_value_at_date(opening_date) for a in all_for_period) if all_for_period else Decimal('0')
        else:
            total_opening_value = total_nbv

        if closing_date:
            total_closing_value = sum(a.get_value_at_date(closing_date) for a in all_for_period) if all_for_period else Decimal('0')
        else:
            total_closing_value = total_nbv

        context['total_opening_value'] = total_opening_value
        context['total_closing_value'] = total_closing_value
        context['period_depreciation'] = total_opening_value - total_closing_value
        context['opening_date'] = opening_date
        context['closing_date'] = closing_date
        
        # Group grouping
        grouped_data = queryset.values('group', 'group__name').annotate(
            count=Count('id'),
            total_cost=Sum('purchase_price')
        ).order_by('-total_cost')[:100]
        
        grouped_list = []
        for group in grouped_data:
            group_id = group['group']
            group_assets = list(queryset.filter(group_id=group_id))
            total_group_dep = sum(a.accumulated_depreciation for a in group_assets) if group_assets else Decimal('0')
            
            grouped_list.append({
                'id': group_id,
                'name': group['group__name'] or 'Uncategorized',
                'total_cost': group['total_cost'] or Decimal('0'),
                'total_acc_dep': total_group_dep,
                'total_nbv': (group['total_cost'] or Decimal('0')) - total_group_dep,
                'count': group['count'],
            })
        
        context['grouped_data'] = grouped_list
        context['categories'] = Category.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['departments'] = Department.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['locations'] = Location.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['branches'] = Branch.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['sites'] = Site.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['buildings'] = Building.objects.filter(organization=org).only('id', 'name').order_by('name')
        context['groups'] = Group.objects.filter(organization=org).only('id', 'name').order_by('name')
        
        # Filter persistence
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['query_params'] = query_params.urlencode()
        
        # Store filter values
        context['depr_date_from'] = self.request.GET.get('depr_date_from', '')
        context['depr_date_to'] = self.request.GET.get('depr_date_to', '')
        context['depr_category'] = self.request.GET.get('depr_category', '')
        context['depr_group'] = self.request.GET.get('depr_group', '')
        context['depr_department'] = self.request.GET.get('depr_department', '')
        context['depr_site'] = self.request.GET.get('depr_site', '')
        context['depr_branch'] = self.request.GET.get('depr_branch', '')
        context['depr_building'] = self.request.GET.get('depr_building', '')
        context['depr_location'] = self.request.GET.get('depr_location', '')
        context['search_query'] = self.request.GET.get('q', '')
        context['depr_tagging_status'] = self.request.GET.get('depr_tagging_status', '')
        context['tagging_status_choices'] = [('TAGGED', 'Tagged'), ('UNTAGGED', 'Untagged')]

        if context.get('page_obj'):
            enriched_assets = []
            for asset in context['page_obj'].object_list:
                asset.opening_value = asset.get_value_at_date(opening_date) if opening_date else asset.current_value
                asset.closing_value = asset.get_value_at_date(closing_date) if closing_date else asset.current_value
                asset.period_depreciation = asset.opening_value - asset.closing_value
                enriched_assets.append(asset)
            context['assets'] = enriched_assets
        
        return context

# AJAX View to create category inline
@login_required
def ajax_create_category(request):
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            useful_life_years = request.POST.get('useful_life_years', '5')
            depreciation_method = request.POST.get('depreciation_method', 'straight_line')
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Category name is required'}, status=400)
            
            # Check if category already exists
            if Category.objects.filter(organization=request.user.organization, name__iexact=name).exists():
                return JsonResponse({'success': False, 'error': 'Category with this name already exists'}, status=400)
            
            category = Category.objects.create(
                organization=request.user.organization,
                name=name,
                useful_life_years=int(useful_life_years),
                depreciation_method=depreciation_method
            )
            
            return JsonResponse({
                'success': True,
                'id': category.id,
                'name': category.name
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)


# AJAX View to create subcategory inline
@login_required
def ajax_create_subcategory(request):
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            category_id = request.POST.get('category_id', '')
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Subcategory name is required'}, status=400)
            
            if not category_id:
                return JsonResponse({'success': False, 'error': 'Category is required'}, status=400)
            
            # Get category and verify it belongs to user's organization
            try:
                category = Category.objects.get(id=category_id, organization=request.user.organization)
            except Category.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Invalid category'}, status=400)
            
            # Check if subcategory already exists
            if SubCategory.objects.filter(category=category, name__iexact=name).exists():
                return JsonResponse({'success': False, 'error': 'Subcategory with this name already exists in this category'}, status=400)
            
            subcategory = SubCategory.objects.create(
                category=category,
                name=name
            )
            
            return JsonResponse({
                'success': True,
                'id': subcategory.id,
                'name': subcategory.name
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)

# Barcode & QR Code Views
@login_required
def generate_asset_codes(request, pk):
    """Generate or regenerate barcode/QR/label for an asset."""
    try:
        asset = Asset.objects.get(id=pk, organization=request.user.organization)
        from .code_generators import AssetCodeGenerator
        
        barcode_path = AssetCodeGenerator.save_barcode_to_file(asset.asset_tag)
        qr_path = AssetCodeGenerator.save_qr_to_file(asset.asset_tag)
        label_path = AssetCodeGenerator.save_label_to_file(asset.asset_tag)
        
        asset.barcode_image = barcode_path
        asset.qr_code_image = qr_path
        asset.label_image = label_path
        asset.save(update_fields=['barcode_image', 'qr_code_image', 'label_image'])
        
        return JsonResponse({
            'success': True,
            'barcode_url': asset.barcode_image.url if asset.barcode_image else None,
            'qr_url': asset.qr_code_image.url if asset.qr_code_image else None,
            'label_url': asset.label_image.url if asset.label_image else None,
        })
    except Asset.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Asset not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def download_asset_barcode(request, pk):
    """Download barcode image for asset"""
    try:
        asset = Asset.objects.get(id=pk, organization=request.user.organization)
        if not asset.barcode_image:
            from .code_generators import AssetCodeGenerator
            path = AssetCodeGenerator.save_barcode_to_file(asset.asset_tag)
            if path:
                asset.barcode_image = path
                asset.save(update_fields=['barcode_image'])
        
        if asset.barcode_image:
            return redirect(asset.barcode_image.url)
        return JsonResponse({'error': 'No barcode available'}, status=404)
    except Asset.DoesNotExist:
        return JsonResponse({'error': 'Asset not found'}, status=404)


@login_required
def download_asset_qr(request, pk):
    """Download QR code image for asset"""
    try:
        asset = Asset.objects.get(id=pk, organization=request.user.organization)
        if not asset.qr_code_image:
            from .code_generators import AssetCodeGenerator
            path = AssetCodeGenerator.save_qr_to_file(asset.asset_tag)
            if path:
                asset.qr_code_image = path
                asset.save(update_fields=['qr_code_image'])
        
        if asset.qr_code_image:
            return redirect(asset.qr_code_image.url)
        return JsonResponse({'error': 'No QR code available'}, status=404)
    except Asset.DoesNotExist:
        return JsonResponse({'error': 'Asset not found'}, status=404)


@login_required
def download_asset_label(request, pk):
    """Download combined label image for asset"""
    try:
        asset = Asset.objects.get(id=pk, organization=request.user.organization)
        if not asset.label_image:
            from .code_generators import AssetCodeGenerator
            path = AssetCodeGenerator.save_label_to_file(asset.asset_tag)
            if path:
                asset.label_image = path
                asset.save(update_fields=['label_image'])
        
        if asset.label_image:
            return redirect(asset.label_image.url)
        return JsonResponse({'error': 'No label available'}, status=404)
    except Asset.DoesNotExist:
        return JsonResponse({'error': 'Asset not found'}, status=404)


@login_required
def download_barcode_batch(request):
    """Download barcodes for multiple assets as ZIP."""
    import tempfile
    import zipfile
    from pathlib import Path
    from django.http import FileResponse
    
    try:
        asset_ids = request.GET.get('asset_ids', '').split(',')
        if not asset_ids or not asset_ids[0]:
            return JsonResponse({'error': 'No assets specified'}, status=400)
        
        assets = Asset.objects.filter(
            id__in=asset_ids,
            organization=request.user.organization
        )
        
        if not assets.exists():
            return JsonResponse({'error': 'No assets found'}, status=404)
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as tmp:
            with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zf:
                for asset in assets:
                    if asset.barcode_image:
                        barcode_path = Path(asset.barcode_image.path)
                        if barcode_path.exists():
                            zf.write(barcode_path, arcname=f'{asset.asset_tag}_barcode.png')
            tmp_path = tmp.name
        
        response = FileResponse(
            open(tmp_path, 'rb'),
            content_type='application/zip'
        )
        response['Content-Disposition'] = f'attachment; filename=\"asset_barcodes_{request.user.organization.code}.zip\"'
        return response
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)




class AssetReconciliationReportView(LoginRequiredMixin, View):
    """Comprehensive Asset Reconciliation Report - full-picture summary of the asset register."""
    template_name = 'assets/reconciliation_report.html'

    def get(self, request):
        from django.db.models import Sum, Count, Q
        from django.db.models.functions import Coalesce

        org = request.user.organization
        date_from_str = request.GET.get('date_from', '')
        date_to_str = request.GET.get('date_to', '')

        date_from = None
        date_to = None
        try:
            if date_from_str:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            pass
        try:
            if date_to_str:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            pass

        base_qs = Asset.objects.filter(organization=org, is_deleted=False).select_related(
            'category', 'sub_category', 'group', 'sub_group', 'brand_new', 'supplier',
            'site', 'branch', 'building', 'floor', 'location', 'room', 'sub_location',
            'department', 'created_by'
        )

        # Report-like filters (matching depreciation/advanced report filters)
        q = (request.GET.get('q') or '').strip()
        if q:
            base_qs = base_qs.filter(
                Q(asset_tag__icontains=q) |
                Q(name__icontains=q) |
                Q(asset_code__icontains=q) |
                Q(serial_number__icontains=q)
            )

        depr_product_name = (request.GET.get('depr_product_name') or '').strip()
        if depr_product_name:
            base_qs = base_qs.filter(name__icontains=depr_product_name)

        fk_filters = {
            'depr_category': 'category_id',
            'depr_subcategory': 'sub_category_id',
            'depr_group': 'group_id',
            'depr_sub_group': 'sub_group_id',
            'depr_brand': 'brand_new_id',
            'depr_supplier': 'supplier_id',
            'depr_site': 'site_id',
            'depr_branch': 'branch_id',
            'depr_building': 'building_id',
            'depr_floor': 'floor_id',
            'depr_location': 'location_id',
            'depr_room': 'room_id',
            'depr_sub_location': 'sub_location_id',
            'depr_department': 'department_id',
            'depr_created_by': 'created_by_id',
        }
        for param, field in fk_filters.items():
            val = (request.GET.get(param) or '').strip()
            if val:
                base_qs = base_qs.filter(**{field: val})

        depr_status = (request.GET.get('depr_status') or '').strip()
        if depr_status:
            base_qs = base_qs.filter(status=depr_status)

        depr_condition = (request.GET.get('depr_condition') or '').strip()
        if depr_condition:
            base_qs = base_qs.filter(condition=depr_condition)

        depr_label_type = (request.GET.get('depr_label_type') or '').strip()
        if depr_label_type:
            base_qs = base_qs.filter(label_type=depr_label_type)

        depr_tagging_status = (request.GET.get('depr_tagging_status') or '').strip()
        if depr_tagging_status:
            base_qs = base_qs.filter(tagging_status=depr_tagging_status)

        depr_purchase_date_from = (request.GET.get('depr_purchase_date_from') or '').strip()
        depr_purchase_date_to = (request.GET.get('depr_purchase_date_to') or '').strip()
        if depr_purchase_date_from:
            try:
                base_qs = base_qs.filter(purchase_date__gte=datetime.strptime(depr_purchase_date_from, '%Y-%m-%d').date())
            except ValueError:
                pass
        if depr_purchase_date_to:
            try:
                base_qs = base_qs.filter(purchase_date__lte=datetime.strptime(depr_purchase_date_to, '%Y-%m-%d').date())
            except ValueError:
                pass

        depr_registered_date_from = (request.GET.get('depr_registered_date_from') or '').strip()
        depr_registered_date_to = (request.GET.get('depr_registered_date_to') or '').strip()
        if depr_registered_date_from:
            try:
                base_qs = base_qs.filter(created_at__date__gte=datetime.strptime(depr_registered_date_from, '%Y-%m-%d').date())
            except ValueError:
                pass
        if depr_registered_date_to:
            try:
                base_qs = base_qs.filter(created_at__date__lte=datetime.strptime(depr_registered_date_to, '%Y-%m-%d').date())
            except ValueError:
                pass

        # --- Period slicing ---
        if date_from:
            period_qs = base_qs.filter(purchase_date__gte=date_from)
        else:
            period_qs = base_qs

        if date_to:
            period_qs = period_qs.filter(purchase_date__lte=date_to)

        # Additions in period (purchased within range)
        additions_qs = period_qs if (date_from or date_to) else base_qs.none()

        # Opening balance: purchased before date_from
        if date_from:
            opening_qs = base_qs.filter(Q(purchase_date__lt=date_from) | Q(purchase_date__isnull=True))
        else:
            opening_qs = base_qs.none()

        # --- Helper: aggregate financials from a queryset ---
        def agg_financials(qs):
            result = qs.aggregate(
                count=Count('id'),
                total_cost=Coalesce(Sum('purchase_price'), Decimal('0')),
            )
            assets_list = list(qs)
            acc_dep = sum(a.accumulated_depreciation for a in assets_list)
            nbv = sum(a.current_value for a in assets_list)
            result['acc_dep'] = acc_dep
            result['nbv'] = nbv
            return result

        # Overall totals (all assets, no date filter)
        all_assets = list(base_qs.select_related('category', 'department', 'site'))
        total_count = len(all_assets)
        total_cost = sum((a.purchase_price or Decimal('0')) for a in all_assets)
        total_acc_dep = sum(a.accumulated_depreciation for a in all_assets)
        total_nbv = sum(a.current_value for a in all_assets)

        # Opening / additions / closing
        opening_data = agg_financials(opening_qs) if date_from else None
        additions_data = agg_financials(additions_qs) if (date_from or date_to) else None
        closing_data = agg_financials(base_qs.filter(purchase_date__lte=date_to) if date_to else base_qs)

        # --- By Category ---
        by_category = []
        cat_groups = base_qs.values('category__id', 'category__name').annotate(
            count=Count('id'),
            total_cost=Coalesce(Sum('purchase_price'), Decimal('0'))
        ).order_by('-total_cost')
        for row in cat_groups:
            cat_assets = [a for a in all_assets if a.category_id == row['category__id']]
            acc = sum(a.accumulated_depreciation for a in cat_assets)
            nbv = sum(a.current_value for a in cat_assets)
            by_category.append({
                'name': row['category__name'] or 'Uncategorized',
                'count': row['count'],
                'cost': row['total_cost'] or Decimal('0'),
                'acc_dep': acc,
                'nbv': nbv,
            })

        # --- By Status ---
        by_status = []
        for code, label in Asset.Status.choices:
            status_assets = [a for a in all_assets if a.status == code]
            cost = sum((a.purchase_price or Decimal('0')) for a in status_assets)
            acc = sum(a.accumulated_depreciation for a in status_assets)
            nbv = sum(a.current_value for a in status_assets)
            by_status.append({
                'label': label, 'count': len(status_assets),
                'cost': cost, 'acc_dep': acc, 'nbv': nbv,
            })

        # --- By Condition ---
        by_condition = []
        for code, label in Asset.Condition.choices:
            c_assets = [a for a in all_assets if a.condition == code]
            cost = sum((a.purchase_price or Decimal('0')) for a in c_assets)
            acc = sum(a.accumulated_depreciation for a in c_assets)
            nbv = sum(a.current_value for a in c_assets)
            by_condition.append({
                'label': label, 'count': len(c_assets),
                'cost': cost, 'acc_dep': acc, 'nbv': nbv,
            })

        # --- Tagged vs Untagged ---
        tagged_assets = [a for a in all_assets if a.tagging_status == 'TAGGED']
        untagged_assets = [a for a in all_assets if a.tagging_status != 'TAGGED']
        by_tagged = [
            {
                'label': 'Tagged',
                'count': len(tagged_assets),
                'cost': sum((a.purchase_price or Decimal('0')) for a in tagged_assets),
                'acc_dep': sum(a.accumulated_depreciation for a in tagged_assets),
                'nbv': sum(a.current_value for a in tagged_assets),
            },
            {
                'label': 'Untagged',
                'count': len(untagged_assets),
                'cost': sum((a.purchase_price or Decimal('0')) for a in untagged_assets),
                'acc_dep': sum(a.accumulated_depreciation for a in untagged_assets),
                'nbv': sum(a.current_value for a in untagged_assets),
            },
        ]

        # --- By Department ---
        by_department = []
        dept_map = {}
        for a in all_assets:
            key = a.department_id
            label = a.department.name if a.department else 'No Department'
            if key not in dept_map:
                dept_map[key] = {'label': label, 'assets': []}
            dept_map[key]['assets'].append(a)
        for entry in sorted(dept_map.values(), key=lambda x: -sum((a.purchase_price or 0) for a in x['assets'])):
            a_list = entry['assets']
            by_department.append({
                'label': entry['label'],
                'count': len(a_list),
                'cost': sum((a.purchase_price or Decimal('0')) for a in a_list),
                'acc_dep': sum(a.accumulated_depreciation for a in a_list),
                'nbv': sum(a.current_value for a in a_list),
            })

        # --- By Site ---
        by_site = []
        site_map = {}
        for a in all_assets:
            key = a.site_id
            label = a.site.name if a.site else 'No Site'
            if key not in site_map:
                site_map[key] = {'label': label, 'assets': []}
            site_map[key]['assets'].append(a)
        for entry in sorted(site_map.values(), key=lambda x: -sum((a.purchase_price or 0) for a in x['assets'])):
            a_list = entry['assets']
            by_site.append({
                'label': entry['label'],
                'count': len(a_list),
                'cost': sum((a.purchase_price or Decimal('0')) for a in a_list),
                'acc_dep': sum(a.accumulated_depreciation for a in a_list),
                'nbv': sum(a.current_value for a in a_list),
            })

        context = {
            # Totals
            'total_count': total_count,
            'total_cost': total_cost,
            'total_acc_dep': total_acc_dep,
            'total_nbv': total_nbv,
            # Period
            'date_from': date_from_str,
            'date_to': date_to_str,
            'opening_data': opening_data,
            'additions_data': additions_data,
            'closing_data': closing_data,
            # Breakdowns
            'by_category': by_category,
            'by_status': by_status,
            'by_condition': by_condition,
            'by_tagged': by_tagged,
            'by_department': by_department,
            'by_site': by_site,
            # Currency
            'currency': 'AED',
            # Filter options
            'categories': Category.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'subcategories': SubCategory.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'groups': Group.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'subgroups': SubGroup.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'brands': Brand.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'suppliers': Supplier.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'sites': Site.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'branches': Branch.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'buildings': Building.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'floors': Floor.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'locations': Location.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'rooms': Room.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'sublocations': SubLocation.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'departments': Department.objects.filter(organization=org).only('id', 'name').order_by('name'),
            'creators': User.objects.filter(organization=org).only('id', 'first_name', 'last_name', 'username').order_by('first_name', 'last_name', 'username'),
            'statuses': Asset.Status.choices,
            'conditions': Asset.Condition.choices,
            'label_types': Asset.LabelType.choices,
            # Selected filter values
            'depr_category': request.GET.get('depr_category', ''),
            'depr_subcategory': request.GET.get('depr_subcategory', ''),
            'depr_group': request.GET.get('depr_group', ''),
            'depr_sub_group': request.GET.get('depr_sub_group', ''),
            'depr_status': request.GET.get('depr_status', ''),
            'depr_condition': request.GET.get('depr_condition', ''),
            'depr_label_type': request.GET.get('depr_label_type', ''),
            'depr_tagging_status': request.GET.get('depr_tagging_status', ''),
            'depr_brand': request.GET.get('depr_brand', ''),
            'depr_supplier': request.GET.get('depr_supplier', ''),
            'depr_product_name': request.GET.get('depr_product_name', ''),
            'depr_site': request.GET.get('depr_site', ''),
            'depr_branch': request.GET.get('depr_branch', ''),
            'depr_building': request.GET.get('depr_building', ''),
            'depr_floor': request.GET.get('depr_floor', ''),
            'depr_location': request.GET.get('depr_location', ''),
            'depr_room': request.GET.get('depr_room', ''),
            'depr_sub_location': request.GET.get('depr_sub_location', ''),
            'depr_department': request.GET.get('depr_department', ''),
            'depr_purchase_date_from': request.GET.get('depr_purchase_date_from', ''),
            'depr_purchase_date_to': request.GET.get('depr_purchase_date_to', ''),
            'depr_registered_date_from': request.GET.get('depr_registered_date_from', ''),
            'depr_registered_date_to': request.GET.get('depr_registered_date_to', ''),
            'depr_created_by': request.GET.get('depr_created_by', ''),
            'has_advanced_filters': any((request.GET.get(k) or '').strip() for k in [
                'q', 'depr_category', 'depr_subcategory', 'depr_group', 'depr_sub_group', 'depr_status', 'depr_condition',
                'depr_label_type', 'depr_tagging_status', 'depr_brand', 'depr_supplier', 'depr_product_name',
                'depr_site', 'depr_branch', 'depr_building', 'depr_floor', 'depr_location', 'depr_room',
                'depr_sub_location', 'depr_department', 'depr_purchase_date_from', 'depr_purchase_date_to',
                'depr_registered_date_from', 'depr_registered_date_to', 'depr_created_by', 'date_from', 'date_to'
            ]),
        }
        return render(request, self.template_name, context)


class AssetReconciliationReportPDFView(LoginRequiredMixin, View):
    """PDF export of Asset Reconciliation Report."""

    def get(self, request):
        from django.db.models import Sum, Count, Q
        from django.db.models.functions import Coalesce

        org = request.user.organization
        date_from_str = request.GET.get('date_from', '')
        date_to_str = request.GET.get('date_to', '')

        date_from = None
        date_to = None
        try:
            if date_from_str:
                date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        except ValueError:
            pass
        try:
            if date_to_str:
                date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        except ValueError:
            pass

        base_qs = Asset.objects.filter(organization=org, is_deleted=False)

        q = (request.GET.get('q') or '').strip()
        if q:
            base_qs = base_qs.filter(
                Q(asset_tag__icontains=q) |
                Q(name__icontains=q) |
                Q(asset_code__icontains=q) |
                Q(serial_number__icontains=q)
            )

        depr_product_name = (request.GET.get('depr_product_name') or '').strip()
        if depr_product_name:
            base_qs = base_qs.filter(name__icontains=depr_product_name)

        fk_filters = {
            'depr_category': 'category_id',
            'depr_subcategory': 'sub_category_id',
            'depr_group': 'group_id',
            'depr_sub_group': 'sub_group_id',
            'depr_brand': 'brand_new_id',
            'depr_supplier': 'supplier_id',
            'depr_site': 'site_id',
            'depr_branch': 'branch_id',
            'depr_building': 'building_id',
            'depr_floor': 'floor_id',
            'depr_location': 'location_id',
            'depr_room': 'room_id',
            'depr_sub_location': 'sub_location_id',
            'depr_department': 'department_id',
            'depr_created_by': 'created_by_id',
        }
        for param, field in fk_filters.items():
            val = (request.GET.get(param) or '').strip()
            if val:
                base_qs = base_qs.filter(**{field: val})

        depr_status = (request.GET.get('depr_status') or '').strip()
        if depr_status:
            base_qs = base_qs.filter(status=depr_status)

        depr_condition = (request.GET.get('depr_condition') or '').strip()
        if depr_condition:
            base_qs = base_qs.filter(condition=depr_condition)

        depr_label_type = (request.GET.get('depr_label_type') or '').strip()
        if depr_label_type:
            base_qs = base_qs.filter(label_type=depr_label_type)

        depr_tagging_status = (request.GET.get('depr_tagging_status') or '').strip()
        if depr_tagging_status:
            base_qs = base_qs.filter(tagging_status=depr_tagging_status)

        depr_purchase_date_from = (request.GET.get('depr_purchase_date_from') or '').strip()
        depr_purchase_date_to = (request.GET.get('depr_purchase_date_to') or '').strip()
        if depr_purchase_date_from:
            try:
                base_qs = base_qs.filter(purchase_date__gte=datetime.strptime(depr_purchase_date_from, '%Y-%m-%d').date())
            except ValueError:
                pass
        if depr_purchase_date_to:
            try:
                base_qs = base_qs.filter(purchase_date__lte=datetime.strptime(depr_purchase_date_to, '%Y-%m-%d').date())
            except ValueError:
                pass

        depr_registered_date_from = (request.GET.get('depr_registered_date_from') or '').strip()
        depr_registered_date_to = (request.GET.get('depr_registered_date_to') or '').strip()
        if depr_registered_date_from:
            try:
                base_qs = base_qs.filter(created_at__date__gte=datetime.strptime(depr_registered_date_from, '%Y-%m-%d').date())
            except ValueError:
                pass
        if depr_registered_date_to:
            try:
                base_qs = base_qs.filter(created_at__date__lte=datetime.strptime(depr_registered_date_to, '%Y-%m-%d').date())
            except ValueError:
                pass

        if date_from:
            period_qs = base_qs.filter(purchase_date__gte=date_from)
        else:
            period_qs = base_qs
        if date_to:
            period_qs = period_qs.filter(purchase_date__lte=date_to)
        additions_qs = period_qs if (date_from or date_to) else base_qs.none()
        if date_from:
            opening_qs = base_qs.filter(Q(purchase_date__lt=date_from) | Q(purchase_date__isnull=True))
        else:
            opening_qs = base_qs.none()

        def agg_financials(qs):
            result = qs.aggregate(count=Count('id'), total_cost=Coalesce(Sum('purchase_price'), Decimal('0')))
            assets_list = list(qs)
            result['acc_dep'] = sum(a.accumulated_depreciation for a in assets_list)
            result['nbv'] = sum(a.current_value for a in assets_list)
            return result

        all_assets = list(base_qs.select_related('category', 'department', 'site'))
        total_count = len(all_assets)
        total_cost = sum((a.purchase_price or Decimal('0')) for a in all_assets)
        total_acc_dep = sum(a.accumulated_depreciation for a in all_assets)
        total_nbv = sum(a.current_value for a in all_assets)

        opening_data = agg_financials(opening_qs) if date_from else None
        additions_data = agg_financials(additions_qs) if (date_from or date_to) else None
        closing_data = agg_financials(base_qs.filter(purchase_date__lte=date_to) if date_to else base_qs)

        by_category = []
        cat_groups = base_qs.values('category__id', 'category__name').annotate(
            count=Count('id'), total_cost=Coalesce(Sum('purchase_price'), Decimal('0'))
        ).order_by('-total_cost')
        for row in cat_groups:
            cat_assets = [a for a in all_assets if a.category_id == row['category__id']]
            by_category.append({
                'name': row['category__name'] or 'Uncategorized',
                'count': row['count'],
                'cost': row['total_cost'] or Decimal('0'),
                'acc_dep': sum(a.accumulated_depreciation for a in cat_assets),
                'nbv': sum(a.current_value for a in cat_assets),
            })

        by_status = []
        for code, label in Asset.Status.choices:
            s_assets = [a for a in all_assets if a.status == code]
            by_status.append({
                'label': label, 'count': len(s_assets),
                'cost': sum((a.purchase_price or Decimal('0')) for a in s_assets),
                'acc_dep': sum(a.accumulated_depreciation for a in s_assets),
                'nbv': sum(a.current_value for a in s_assets),
            })

        by_condition = []
        for code, label in Asset.Condition.choices:
            c_assets = [a for a in all_assets if a.condition == code]
            by_condition.append({
                'label': label, 'count': len(c_assets),
                'cost': sum((a.purchase_price or Decimal('0')) for a in c_assets),
                'acc_dep': sum(a.accumulated_depreciation for a in c_assets),
                'nbv': sum(a.current_value for a in c_assets),
            })

        tagged_assets = [a for a in all_assets if a.tagging_status == 'TAGGED']
        untagged_assets = [a for a in all_assets if a.tagging_status != 'TAGGED']
        by_tagged = [
            {'label': 'Tagged', 'count': len(tagged_assets),
             'cost': sum((a.purchase_price or Decimal('0')) for a in tagged_assets),
             'acc_dep': sum(a.accumulated_depreciation for a in tagged_assets),
             'nbv': sum(a.current_value for a in tagged_assets)},
            {'label': 'Untagged', 'count': len(untagged_assets),
             'cost': sum((a.purchase_price or Decimal('0')) for a in untagged_assets),
             'acc_dep': sum(a.accumulated_depreciation for a in untagged_assets),
             'nbv': sum(a.current_value for a in untagged_assets)},
        ]

        by_department = []
        dept_map = {}
        for a in all_assets:
            key = a.department_id
            label = a.department.name if a.department else 'No Department'
            if key not in dept_map:
                dept_map[key] = {'label': label, 'assets': []}
            dept_map[key]['assets'].append(a)
        for entry in sorted(dept_map.values(), key=lambda x: -sum((a.purchase_price or 0) for a in x['assets'])):
            a_list = entry['assets']
            by_department.append({
                'label': entry['label'], 'count': len(a_list),
                'cost': sum((a.purchase_price or Decimal('0')) for a in a_list),
                'acc_dep': sum(a.accumulated_depreciation for a in a_list),
                'nbv': sum(a.current_value for a in a_list),
            })

        by_site = []
        site_map = {}
        for a in all_assets:
            key = a.site_id
            label = a.site.name if a.site else 'No Site'
            if key not in site_map:
                site_map[key] = {'label': label, 'assets': []}
            site_map[key]['assets'].append(a)
        for entry in sorted(site_map.values(), key=lambda x: -sum((a.purchase_price or 0) for a in x['assets'])):
            a_list = entry['assets']
            by_site.append({
                'label': entry['label'], 'count': len(a_list),
                'cost': sum((a.purchase_price or Decimal('0')) for a in a_list),
                'acc_dep': sum(a.accumulated_depreciation for a in a_list),
                'nbv': sum(a.current_value for a in a_list),
            })

        context = {
            'organization': str(org),
            'generated_on': timezone.now(),
            'currency': 'AED',
            'date_from': date_from_str,
            'date_to': date_to_str,
            'total_count': total_count,
            'total_cost': total_cost,
            'total_acc_dep': total_acc_dep,
            'total_nbv': total_nbv,
            'opening_data': opening_data,
            'additions_data': additions_data,
            'closing_data': closing_data,
            'by_category': by_category,
            'by_status': by_status,
            'by_condition': by_condition,
            'by_tagged': by_tagged,
            'by_department': by_department,
            'by_site': by_site,
        }
        from fpdf import FPDF
        pdf = FPDF(orientation='L', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.set_margins(10, 10, 10)
        pdf.add_page()
        W = pdf.epw

        def fmt(val):
            try:
                return f'{float(val):,.0f}'
            except Exception:
                return '0'

        generated_str = timezone.now().strftime('%d %b %Y, %H:%M')
        period_str = 'All time'
        if date_from_str or date_to_str:
            period_str = f"{date_from_str or 'Beginning'} to {date_to_str or 'Today'}"

        # Header
        pdf.set_fill_color(30, 58, 95)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Helvetica', 'B', 14)
        pdf.cell(W, 8, 'Asset Reconciliation Report', border=0, align='L', fill=True)
        pdf.ln(8)
        pdf.set_font('Helvetica', '', 8)
        pdf.cell(W, 5, f'Org: {org}   |   Generated: {generated_str}   |   Period: {period_str}', border=0, align='L', fill=True)
        pdf.ln(10)

        # KPI Boxes
        pdf.set_text_color(0, 0, 0)
        kw = (W - 6) / 4
        kpis = [
            ('TOTAL ASSETS', str(total_count), 'items in register'),
            ('ORIGINAL COST', f'AED {fmt(total_cost)}', 'purchase value'),
            ('ACC. DEPRECIATION', f'AED {fmt(total_acc_dep)}', 'total written down'),
            ('NET BOOK VALUE', f'AED {fmt(total_nbv)}', 'current carry value'),
        ]
        kpi_y = pdf.get_y()
        for i, (lbl, val, sub) in enumerate(kpis):
            x = 10 + i * (kw + 2)
            pdf.set_fill_color(247, 250, 253)
            pdf.set_draw_color(224, 232, 240)
            pdf.rect(x, kpi_y, kw, 16, style='FD')
            pdf.set_xy(x + 2, kpi_y + 1)
            pdf.set_font('Helvetica', '', 6)
            pdf.set_text_color(107, 114, 128)
            pdf.cell(kw - 4, 4, lbl)
            pdf.set_xy(x + 2, kpi_y + 6)
            pdf.set_font('Helvetica', 'B', 10)
            pdf.set_text_color(30, 58, 95)
            pdf.cell(kw - 4, 6, val)
            pdf.set_xy(x + 2, kpi_y + 12)
            pdf.set_font('Helvetica', '', 6)
            pdf.set_text_color(156, 163, 175)
            pdf.cell(kw - 4, 4, sub)
        pdf.set_xy(10, kpi_y + 20)

        def section_title(title):
            pdf.set_font('Helvetica', 'B', 9)
            pdf.set_text_color(30, 58, 95)
            pdf.set_draw_color(37, 99, 235)
            pdf.cell(W, 6, title, border='B', align='L')
            pdf.ln(7)
            pdf.set_draw_color(0, 0, 0)

        def fin_table(rows, col1='Name'):
            c0 = 90
            c1 = 22
            cn = (W - c0 - c1) / 3
            cols = [(col1, c0, 'L'), ('Count', c1, 'C'),
                    ('Cost (AED)', cn, 'R'), ('Acc. Dep (AED)', cn, 'R'), ('NBV (AED)', cn, 'R')]
            pdf.set_fill_color(30, 58, 95)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font('Helvetica', 'B', 7)
            for cname, cw, calign in cols:
                pdf.cell(cw, 6, cname, border=0, align=calign, fill=True)
            pdf.ln()
            for i, row in enumerate(rows):
                if row.get('_total'):
                    pdf.set_fill_color(232, 240, 250)
                    pdf.set_text_color(30, 58, 95)
                    pdf.set_font('Helvetica', 'B', 7)
                elif i % 2:
                    pdf.set_fill_color(240, 245, 251)
                    pdf.set_text_color(55, 65, 81)
                    pdf.set_font('Helvetica', '', 7)
                else:
                    pdf.set_fill_color(255, 255, 255)
                    pdf.set_text_color(55, 65, 81)
                    pdf.set_font('Helvetica', '', 7)
                name = row.get('label') or row.get('name', '')
                vals = [name, str(row.get('count', '')),
                        fmt(row.get('cost', 0)), fmt(row.get('acc_dep', 0)), fmt(row.get('nbv', 0))]
                for (_, cw, calign), v in zip(cols, vals):
                    pdf.cell(cw, 5, v, border='B', align=calign, fill=True)
                pdf.ln()
            pdf.ln(3)

        section_title('By Category')
        fin_table(list(by_category) + [{'name': 'TOTAL', 'count': total_count, 'cost': total_cost, 'acc_dep': total_acc_dep, 'nbv': total_nbv, '_total': True}], 'Category')

        section_title('By Status')
        fin_table([r for r in by_status if r['count'] > 0], 'Status')

        section_title('By Condition')
        fin_table([r for r in by_condition if r['count'] > 0], 'Condition')

        if pdf.get_y() > 150:
            pdf.add_page()
        section_title('By Department')
        fin_table(by_department, 'Department')

        if pdf.get_y() > 150:
            pdf.add_page()
        section_title('By Site')
        fin_table(by_site, 'Site')

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="asset_reconciliation_report.pdf"'
        response.write(bytes(pdf.output()))
        return response


# AJAX View to create category inline
@login_required
def ajax_create_category(request):
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            useful_life_years = request.POST.get('useful_life_years', '5')
            depreciation_method = request.POST.get('depreciation_method', 'straight_line')
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Category name is required'}, status=400)
            
            # Check if category already exists
            if Category.objects.filter(organization=request.user.organization, name__iexact=name).exists():
                return JsonResponse({'success': False, 'error': 'Category with this name already exists'}, status=400)
            
            category = Category.objects.create(
                organization=request.user.organization,
                name=name,
                useful_life_years=int(useful_life_years),
                depreciation_method=depreciation_method
            )
            
            return JsonResponse({
                'success': True,
                'id': category.id,
                'name': category.name
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)


# AJAX View to create subcategory inline
@login_required
def ajax_create_subcategory(request):
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            category_id = request.POST.get('category_id', '')
            
            if not name:
                return JsonResponse({'success': False, 'error': 'Subcategory name is required'}, status=400)
            
            if not category_id:
                return JsonResponse({'success': False, 'error': 'Category is required'}, status=400)
            
            # Get category and verify it belongs to user's organization
            try:
                category = Category.objects.get(id=category_id, organization=request.user.organization)
            except Category.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Invalid category'}, status=400)
            
            # Check if subcategory already exists
            if SubCategory.objects.filter(category=category, name__iexact=name).exists():
                return JsonResponse({'success': False, 'error': 'Subcategory with this name already exists in this category'}, status=400)
            
            subcategory = SubCategory.objects.create(
                category=category,
                name=name
            )
            
            return JsonResponse({
                'success': True,
                'id': subcategory.id,
                'name': subcategory.name
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)

# Barcode & QR Code Views
@login_required
def generate_asset_codes(request, pk):
    """Generate or regenerate barcode/QR/label for an asset."""
    try:
        asset = Asset.objects.get(id=pk, organization=request.user.organization)
        from .code_generators import AssetCodeGenerator
        
        barcode_path = AssetCodeGenerator.save_barcode_to_file(asset.asset_tag)
        qr_path = AssetCodeGenerator.save_qr_to_file(asset.asset_tag)
        label_path = AssetCodeGenerator.save_label_to_file(asset.asset_tag)
        
        asset.barcode_image = barcode_path
        asset.qr_code_image = qr_path
        asset.label_image = label_path
        asset.save(update_fields=['barcode_image', 'qr_code_image', 'label_image'])
        
        return JsonResponse({
            'success': True,
            'barcode_url': asset.barcode_image.url if asset.barcode_image else None,
            'qr_url': asset.qr_code_image.url if asset.qr_code_image else None,
            'label_url': asset.label_image.url if asset.label_image else None,
        })
    except Asset.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Asset not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def download_asset_barcode(request, pk):
    """Download barcode image for asset"""
    try:
        asset = Asset.objects.get(id=pk, organization=request.user.organization)
        if not asset.barcode_image:
            from .code_generators import AssetCodeGenerator
            path = AssetCodeGenerator.save_barcode_to_file(asset.asset_tag)
            if path:
                asset.barcode_image = path
                asset.save(update_fields=['barcode_image'])
        
        if asset.barcode_image:
            return redirect(asset.barcode_image.url)
        return JsonResponse({'error': 'No barcode available'}, status=404)
    except Asset.DoesNotExist:
        return JsonResponse({'error': 'Asset not found'}, status=404)


@login_required
def download_asset_qr(request, pk):
    """Download QR code image for asset"""
    try:
        asset = Asset.objects.get(id=pk, organization=request.user.organization)
        if not asset.qr_code_image:
            from .code_generators import AssetCodeGenerator
            path = AssetCodeGenerator.save_qr_to_file(asset.asset_tag)
            if path:
                asset.qr_code_image = path
                asset.save(update_fields=['qr_code_image'])
        
        if asset.qr_code_image:
            return redirect(asset.qr_code_image.url)
        return JsonResponse({'error': 'No QR code available'}, status=404)
    except Asset.DoesNotExist:
        return JsonResponse({'error': 'Asset not found'}, status=404)


@login_required
def download_asset_label(request, pk):
    """Download combined label image for asset"""
    try:
        asset = Asset.objects.get(id=pk, organization=request.user.organization)
        if not asset.label_image:
            from .code_generators import AssetCodeGenerator
            path = AssetCodeGenerator.save_label_to_file(asset.asset_tag)
            if path:
                asset.label_image = path
                asset.save(update_fields=['label_image'])
        
        if asset.label_image:
            return redirect(asset.label_image.url)
        return JsonResponse({'error': 'No label available'}, status=404)
    except Asset.DoesNotExist:
        return JsonResponse({'error': 'Asset not found'}, status=404)


@login_required
def download_barcode_batch(request):
    """Download barcodes for multiple assets as ZIP."""
    import tempfile
    import zipfile
    from pathlib import Path
    from django.http import FileResponse
    
    try:
        asset_ids = request.GET.get('asset_ids', '').split(',')
        if not asset_ids or not asset_ids[0]:
            return JsonResponse({'error': 'No assets specified'}, status=400)
        
        assets = Asset.objects.filter(
            id__in=asset_ids,
            organization=request.user.organization
        )
        
        if not assets.exists():
            return JsonResponse({'error': 'No assets found'}, status=404)
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.zip') as tmp:
            with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zf:
                for asset in assets:
                    if asset.barcode_image:
                        barcode_path = Path(asset.barcode_image.path)
                        if barcode_path.exists():
                            zf.write(barcode_path, arcname=f'{asset.asset_tag}_barcode.png')
            tmp_path = tmp.name
        
        response = FileResponse(
            open(tmp_path, 'rb'),
            content_type='application/zip'
        )
        response['Content-Disposition'] = f'attachment; filename=\"asset_barcodes_{request.user.organization.code}.zip\"'
        return response
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def asset_attachment_upload(request, pk):
    """AJAX: Upload a file attachment to an asset."""
    try:
        asset = Asset.objects.get(pk=pk, organization=request.user.organization)
    except Asset.DoesNotExist:
        return JsonResponse({'error': 'Asset not found'}, status=404)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return JsonResponse({'error': 'No file provided'}, status=400)

    # Basic validation: max 1 MB
    if uploaded_file.size > 1 * 1024 * 1024:
        return JsonResponse({'error': 'File too large (max 1 MB)'}, status=400)

    attachment_type = request.POST.get('attachment_type', 'OTHER')
    valid_types = {t[0] for t in AssetAttachment.Type.choices}
    if attachment_type not in valid_types:
        attachment_type = 'OTHER'

    description = request.POST.get('description', '')[:255]

    attachment = AssetAttachment.objects.create(
        asset=asset,
        organization=request.user.organization,
        file=uploaded_file,
        attachment_type=attachment_type,
        description=description,
    )

    return JsonResponse({
        'success': True,
        'id': attachment.pk,
        'name': uploaded_file.name,
        'url': attachment.file.url,
        'type': attachment.get_attachment_type_display(),
        'description': attachment.description,
        'size': uploaded_file.size,
    })


@login_required
def asset_attachment_delete(request, pk, attachment_id):
    """AJAX: Delete an attachment from an asset."""
    try:
        asset = Asset.objects.get(pk=pk, organization=request.user.organization)
        attachment = AssetAttachment.objects.get(pk=attachment_id, asset=asset)
    except (Asset.DoesNotExist, AssetAttachment.DoesNotExist):
        return JsonResponse({'error': 'Not found'}, status=404)

    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    attachment.file.delete(save=False)
    attachment.delete()
    return JsonResponse({'success': True})


@login_required
def label_print_center(request):
    """Label Print Center â€” lets the user pick a design and select assets to print."""
    org = request.user.organization
    categories = Category.objects.filter(organization=org).order_by('name')
    branches = Branch.objects.filter(organization=org).order_by('name')
    all_assets = Asset.objects.filter(organization=org, is_deleted=False).order_by('asset_tag').values('asset_tag')
    return render(request, 'assets/label_print_center.html', {
        'org': org,
        'categories': categories,
        'branches': branches,
        'all_assets': all_assets,
    })


@login_required
def print_asset_labels_bulk(request):
    """Renders the printable label sheet based on filters passed via GET params."""
    org = request.user.organization
    design = request.GET.get('design', 'CLASSIC').upper()
    VALID_DESIGNS = {'CLASSIC', 'COMPACT', 'DETAILED', 'BARCODE_ONLY'}
    if design not in VALID_DESIGNS:
        design = 'CLASSIC'

    DESIGN_CHOICES = [
        ('CLASSIC', 'Classic'),
        ('COMPACT', 'Compact'),
        ('DETAILED', 'Standard'),
        ('BARCODE_ONLY', 'Barcode Only'),
    ]

    qs = Asset.objects.filter(organization=org, is_deleted=False)

    # Filter by specific IDs (from asset-list checkbox selection)
    ids = request.GET.get('ids', '')
    if ids:
        id_list = [i.strip() for i in ids.split(',') if i.strip()]
        qs = qs.filter(id__in=id_list)
    else:
        # Filter by tag range
        tag_from = request.GET.get('tag_from', '')
        tag_to = request.GET.get('tag_to', '')
        category = request.GET.get('category', '')
        branch = request.GET.get('branch', '')
        specific_tags = request.GET.get('specific_tags', '')

        if tag_from and tag_to:
            qs = qs.filter(asset_tag__gte=tag_from, asset_tag__lte=tag_to)
        elif category or branch:
            if category:
                qs = qs.filter(category_id=category)
            if branch:
                qs = qs.filter(branch_id=branch)
        elif specific_tags:
            tags = [t.strip() for t in specific_tags.replace('\n', ',').split(',') if t.strip()]
            qs = qs.filter(asset_tag__in=tags)

    qs = qs.order_by('asset_tag')

    # Batch pagination â€” 100 labels per page
    BATCH_SIZE = 100
    total_count = qs.count()
    total_batches = max(1, (total_count + BATCH_SIZE - 1) // BATCH_SIZE)
    try:
        batch = max(1, min(int(request.GET.get('batch', 1)), total_batches))
    except (ValueError, TypeError):
        batch = 1

    batch_start = (batch - 1) * BATCH_SIZE
    batch_end = batch_start + BATCH_SIZE
    assets = list(qs[batch_start:batch_end])

    # Refresh mode: when ?refresh=1 is passed (via the "Refresh codes" button),
    # scan ALL matching assets (not just the current batch) and regenerate
    # ONLY the ones with missing files. Existing files are left untouched.
    refresh_requested = request.GET.get('refresh') == '1'
    refresh_report = None

    from django.core.files.storage import default_storage
    from PIL import Image
    from .code_generators import generate_codes_for_asset, AssetCodeGenerator

    def _file_missing_or_low_res(field, kind):
        if not field:
            return True
        try:
            if not default_storage.exists(field.name):
                return True

            min_dims = {
                'barcode': (700, 120),
                'qr': (220, 220),
                'label': (700, 300),
            }
            min_w, min_h = min_dims.get(kind, (1, 1))

            with default_storage.open(field.name, 'rb') as f:
                with Image.open(f) as img:
                    w, h = img.size

            return w < min_w or h < min_h
        except Exception:
            return True

    def _heal(target_assets, refresh_after=False):
        regenerated = 0
        failed = 0
        for asset in target_assets:
            try:
                missing_b = _file_missing_or_low_res(asset.barcode_image, 'barcode')
                missing_q = _file_missing_or_low_res(asset.qr_code_image, 'qr')
                missing_l = _file_missing_or_low_res(asset.label_image, 'label')
                if not (missing_b or missing_q or missing_l):
                    continue
                if missing_b:
                    asset.barcode_image = None
                if missing_q:
                    asset.qr_code_image = None
                if missing_l:
                    asset.label_image = None
                generate_codes_for_asset(asset)
                if refresh_after:
                    asset.refresh_from_db(fields=['barcode_image', 'qr_code_image', 'label_image'])
                regenerated += 1
            except Exception as e:
                failed += 1
                print(f"[print_asset_labels_bulk] failed to regenerate codes for {asset.asset_tag}: {e}")
        return regenerated, failed

    if refresh_requested:
        # Scan everything matching the current filter
        try:
            scanned = total_count
            regenerated, failed = _heal(qs.iterator(chunk_size=200), refresh_after=False)
            # Reload the current batch so URLs reflect newly generated files
            assets = list(qs[batch_start:batch_end])
            refresh_report = {
                'scanned': scanned,
                'regenerated': regenerated,
                'failed': failed,
            }
        except Exception as e:
            print(f"[print_asset_labels_bulk] refresh failed: {e}")
            refresh_report = {'scanned': 0, 'regenerated': 0, 'failed': 0, 'error': str(e)}
    else:
        # Default behaviour: auto-heal only the current batch on every load
        try:
            _heal(assets, refresh_after=True)
        except Exception as e:
            print(f"[print_asset_labels_bulk] auto-heal skipped: {e}")

    # Prefer SVG (vector) codes on print page for scanner-grade sharpness.
    for asset in assets:
        asset.barcode_svg_data = None
        asset.qr_svg_data = None
        try:
            asset.barcode_svg_data = AssetCodeGenerator.generate_barcode_svg_data_uri(asset.asset_tag)
        except Exception as e:
            print(f"[print_asset_labels_bulk] barcode SVG fallback for {asset.asset_tag}: {e}")
        try:
            asset.qr_svg_data = AssetCodeGenerator.generate_qr_svg_data_uri(asset.asset_tag)
        except Exception as e:
            print(f"[print_asset_labels_bulk] QR SVG fallback for {asset.asset_tag}: {e}")
    return render(request, 'assets/print_label.html', {
        'org': org,
        'assets': assets,
        'design': design,
        'designs': DESIGN_CHOICES,
        'batch': batch,
        'total_batches': total_batches,
        'total_count': total_count,
        'batch_start': batch_start + 1,
        'batch_end': min(batch_end, total_count),
        'remaining': max(0, total_count - batch_end),
        'refresh_report': refresh_report,
        'asset_ids_json': json.dumps([str(a.id) for a in assets]),
    })


@login_required
@require_POST
def mark_assets_tagged(request):
    """Mark a list of assets as TAGGED (tagging_status='TAGGED')."""
    try:
        body = json.loads(request.body)
        asset_ids = body.get('asset_ids', [])
        if not asset_ids:
            return JsonResponse({'success': False, 'error': 'No asset IDs provided.'}, status=400)
        updated = Asset.objects.filter(
            id__in=asset_ids,
            organization=request.user.organization,
            is_deleted=False,
        ).update(tagging_status='TAGGED')
        return JsonResponse({'success': True, 'updated': updated})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

class ApprovalListView(ApprovalAccessMixin, TemplateView):
    """Unified approvals dashboard showing both asset approvals and disposal requests."""
    template_name = 'assets/approval_list.html'

    def get_disposal_queryset(self):
        """Get all disposal requests for the org. Action permissions are enforced in context/template."""
        user = self.request.user
        return AssetDisposal.objects.filter(
            organization=user.organization
        ).select_related('asset', 'requested_by', 'manager_approved_by').order_by('-created_at')

    def get_asset_approval_queryset(self):
        """Get relevant asset approval requests based on user role."""
        user = self.request.user

        if user.is_checker:
            return ApprovalRequest.objects.filter(
                organization=user.organization,
                status=ApprovalRequest.Status.PENDING,
            ).order_by('-created_at')

        if user.is_senior_manager:
            return ApprovalRequest.objects.filter(
                organization=user.organization,
                status__in=[ApprovalRequest.Status.PENDING, ApprovalRequest.Status.CHECKER_APPROVED],
            ).order_by('-created_at')

        if user.is_superuser or user.role == user.Role.ADMIN:
            return ApprovalRequest.objects.filter(
                organization=user.organization
            ).order_by('-created_at')

        return ApprovalRequest.objects.none()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        disposals = self.get_disposal_queryset()
        asset_approvals = self.get_asset_approval_queryset()

        all_approvals = []
        for disposal in disposals:
            disposal_detail_url = reverse('disposal-detail', kwargs={'pk': disposal.pk})
            disposal_review_url = None
            if user.role in [user.Role.SENIOR_MANAGER, user.Role.CHECKER] and disposal.status == AssetDisposal.Status.PENDING:
                disposal_review_url = reverse('disposal-manager-approve', kwargs={'pk': disposal.pk})
            elif (user.is_superuser or user.role == user.Role.ADMIN) and disposal.status == AssetDisposal.Status.MANAGER_APPROVED:
                disposal_review_url = reverse('disposal-approve', kwargs={'pk': disposal.pk})

            all_approvals.append({
                'type': 'disposal',
                'id': disposal.id,
                'pk': disposal.pk,
                'asset': disposal.asset,
                'status': disposal.status,
                'get_status_display': disposal.get_status_display(),
                'created_at': disposal.created_at,
                'requested_by': disposal.requested_by,
                'get_disposal_method_display': disposal.get_disposal_method_display(),
                'disposal_method': disposal.disposal_method,
                'manager_approved_by': disposal.manager_approved_by,
                'detail_url': disposal_detail_url,
                'review_url': disposal_review_url,
                'can_approve': disposal_review_url is not None,
            })

        for approval in asset_approvals:
            approval_detail_url = reverse('approval-request-detail', kwargs={'pk': approval.pk})
            can_review_asset = False
            if approval.status == ApprovalRequest.Status.PENDING:
                can_review_asset = user.is_checker or user.is_senior_manager or user.is_superuser or user.role == user.Role.ADMIN
            elif approval.status == ApprovalRequest.Status.CHECKER_APPROVED:
                can_review_asset = user.is_senior_manager or user.is_superuser or user.role == user.Role.ADMIN

            all_approvals.append({
                'type': 'asset',
                'id': approval.id,
                'pk': approval.pk,
                'asset': approval.asset,
                'asset_name': approval.asset.name if approval.asset else 'N/A',
                'status': approval.status,
                'get_status_display': approval.get_status_display(),
                'created_at': approval.created_at,
                'requester': approval.requester,
                'request_type': approval.get_request_type_display(),
                'detail_url': approval_detail_url,
                'review_url': approval_detail_url if can_review_asset else None,
                'can_approve': can_review_asset,
            })

        all_approvals.sort(key=lambda x: x['created_at'], reverse=True)

        if user.role in [user.Role.SENIOR_MANAGER, user.Role.CHECKER]:
            disposal_pending = disposals.filter(status=AssetDisposal.Status.PENDING).count()
        elif user.is_superuser or user.role == user.Role.ADMIN:
            disposal_pending = disposals.filter(status=AssetDisposal.Status.MANAGER_APPROVED).count()
        else:
            disposal_pending = disposals.filter(status=AssetDisposal.Status.PENDING).count()

        if user.is_checker:
            asset_pending = asset_approvals.filter(status=ApprovalRequest.Status.PENDING).count()
        elif user.is_senior_manager or user.is_superuser or user.role == user.Role.ADMIN:
            asset_pending = asset_approvals.filter(
                status__in=[ApprovalRequest.Status.PENDING, ApprovalRequest.Status.CHECKER_APPROVED]
            ).count()
        else:
            asset_pending = 0

        context['disposals'] = disposals
        context['asset_approvals'] = asset_approvals
        context['all_approvals'] = all_approvals
        context['disposal_pending'] = disposal_pending
        context['asset_pending'] = asset_pending
        context['total_pending'] = context['disposal_pending'] + context['asset_pending']
        context['total_approved'] = disposals.filter(status=AssetDisposal.Status.APPROVED).count() + \
            asset_approvals.filter(status=ApprovalRequest.Status.APPROVED).count()
        context['total_rejected'] = disposals.filter(status=AssetDisposal.Status.REJECTED).count() + \
            asset_approvals.filter(
                status__in=[
                    ApprovalRequest.Status.CHECKER_REJECTED,
                    ApprovalRequest.Status.SENIOR_REJECTED,
                    ApprovalRequest.Status.REJECTED,
                ]
            ).count()

        return context


